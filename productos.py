# productos.py
import os
import random
import string
import uuid
import requests
from io import BytesIO
from datetime import datetime, date, timedelta
from flask import Blueprint, request, render_template, session, redirect, url_for, jsonify, flash, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import or_
from werkzeug.utils import secure_filename
from decimal import Decimal

from database import db
from models.models import Producto, CatalogoGlobal
from models.modelos_inventario import MovimientoInventario, LoteInventario, LoteMovimientoRelacion
from auth import login_requerido
from utils import (
    process_image, parse_money, normalize_categoria_if_needed,
    find_similar_product_image, find_similar_catalog_image
)
from category_colors import get_category_color, normalize_category
from ajuste_stock import crear_lote_registro

# Importar configuraciones necesarias
from config import UPLOAD_FOLDER, BASE_DIR, ALLOWED_EXTENSIONS

# Crear el blueprint
productos_bp = Blueprint('productos', __name__)

# ==============================
# FUNCIONES AUXILIARES
# ==============================
def generar_codigo_unico():
    """
    Genera un código único para productos sin código de barras con formato 1901XXXXXXXX.
    """
    prefix = "1901"
    caracteres = string.digits
    codigo_aleatorio = ''.join(random.choice(caracteres) for _ in range(8))
    return prefix + codigo_aleatorio

def generar_codigo_a_granel():
    """
    Genera un código único para productos a granel con formato GRA-XXXXXXXX.
    """
    prefix = "GRA-"
    caracteres = string.ascii_uppercase + string.digits
    codigo_aleatorio = ''.join(random.choice(caracteres) for _ in range(8))
    return prefix + codigo_aleatorio

def truncar_url(url, max_length=95):
    """
    Trunca una URL para que quepa en un campo de longitud limitada.
    Preserva el protocolo y dominio para URLs externas.
    """
    if not url or len(url) <= max_length:
        return url
    
    # Convertir a string por si acaso
    url = str(url)
    
    # Verificar si es una URL externa
    if url.startswith(('http://', 'https://')):
        try:
            # Dividir la URL en partes
            from urllib.parse import urlparse
            parsed_url = urlparse(url)
            base = f"{parsed_url.scheme}://{parsed_url.netloc}"
            path = parsed_url.path
            query = parsed_url.query
            
            # Si solo la base ya es demasiado larga
            if len(base) >= max_length - 5:
                return base[:max_length-5] + "..."
            
            # Calcular cuánto espacio queda para el path y query
            remaining = max_length - len(base) - 4  # -4 para "/..."
            
            # Extraer el nombre del archivo de forma segura
            if '/' in path:
                path_parts = path.split('/')
                filename = path_parts[-1] if path_parts else ""
            else:
                filename = path
            
            # Si hay espacio para el nombre completo y no hay query
            if filename and len(filename) <= remaining and not query:
                return f"{base}/.../{filename}"
            
            # Si hay query, incluirla parcialmente si hay espacio
            if query and filename and len(filename) + len(query) + 1 <= remaining:  # +1 por el "?"
                return f"{base}/.../{filename}?{query}"
            
            # Si no hay espacio suficiente para todo, truncar adecuadamente
            if filename and len(filename) <= remaining:
                return f"{base}/.../{filename}"
            
            # Si el nombre es demasiado largo, truncarlo
            if filename:
                return f"{base}/.../{filename[:remaining]}"
            else:
                return f"{base}/..."
                
        except Exception as e:
            print(f"Error al parsear URL {url}: {e}")
            # Si hay algún error, caer al método simple
    
    # Método simple para URLs locales o en caso de error
    try:
        # Intentar preservar al menos el dominio para URLs externas
        if url.startswith(('http://', 'https://')):
            # Buscar el tercer slash
            count = 0
            pos = -1
            for i, char in enumerate(url):
                if char == '/':
                    count += 1
                    if count == 3:
                        pos = i
                        break
            
            if pos > 0 and pos < max_length - 4:
                domain_part = url[:pos]
                return f"{domain_part}/..."
    except:
        pass
    
    # Si todo falla, usar el método más simple y seguro
    if '/' in url:
        parts = url.split('/')
        nombre_archivo = parts[-1] if parts else url
    else:
        nombre_archivo = url
        
    if len(nombre_archivo) > max_length:
        # Tomar los últimos caracteres para preservar la extensión
        return "..." + nombre_archivo[-(max_length-3):]
    elif len(url) > max_length:
        # Si el nombre es corto pero la URL completa es larga
        return ".../" + nombre_archivo
    else:
        return url

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==============================
# RUTAS DE PRODUCTOS
# ==============================
@productos_bp.route('/productos', methods=['GET'])
@login_requerido
def ver_productos():
    empresa_id = session['user_id']
    filtro_aprobacion = request.args.get('filtroAprobacion', 'aprobados')
    termino_busqueda = request.args.get('q', '').strip()

    # Actualizamos automáticamente los lotes sin fecha de caducidad
    try:
        # Obtener productos con caducidad activada
        productos_caducidad = Producto.query.filter(
            Producto.empresa_id == empresa_id,
            Producto.has_caducidad == True,
            Producto.metodo_caducidad.isnot(None)
        ).all()
        
        fixed_count = 0
        for producto in productos_caducidad:
            # Obtener lotes activos sin fecha de caducidad
            lotes = LoteInventario.query.filter(
                LoteInventario.producto_id == producto.id,
                LoteInventario.esta_activo == True,
                LoteInventario.stock > 0,
                LoteInventario.fecha_caducidad.is_(None)
            ).all()
            
            if not lotes:
                continue
                
            # Calcular fecha de caducidad basada en metodo_caducidad
            caducidad_lapso = producto.metodo_caducidad
            fecha_actual = datetime.now()
            fecha_caducidad = None
            
            if caducidad_lapso == '1 día':
                fecha_caducidad = fecha_actual + timedelta(days=1)
            elif caducidad_lapso == '3 días':
                fecha_caducidad = fecha_actual + timedelta(days=3)
            elif caducidad_lapso == '1 semana':
                fecha_caducidad = fecha_actual + timedelta(days=7)
            elif caducidad_lapso == '2 semanas':
                fecha_caducidad = fecha_actual + timedelta(days=14)
            elif caducidad_lapso == '1 mes':
                fecha_caducidad = fecha_actual + timedelta(days=30)
            elif caducidad_lapso == '3 meses':
                fecha_caducidad = fecha_actual + timedelta(days=90)
            elif caducidad_lapso == '6 meses':
                fecha_caducidad = fecha_actual + timedelta(days=180)
            elif caducidad_lapso == '1 año':
                fecha_caducidad = fecha_actual + timedelta(days=365)
            elif caducidad_lapso == '2 años':
                fecha_caducidad = fecha_actual + timedelta(days=730)
            elif caducidad_lapso == '3 años':
                fecha_caducidad = fecha_actual + timedelta(days=1460)  # 4 años (considerados como +3 años)
            
            # Convertir a date
            if fecha_caducidad:
                fecha_caducidad = fecha_caducidad.date()
                
                # Actualizar los lotes
                for lote in lotes:
                    lote.fecha_caducidad = fecha_caducidad
                    fixed_count += 1
                    
        # Guardar cambios si se hicieron modificaciones
        if fixed_count > 0:
            db.session.commit()
    except Exception as e:
        db.session.rollback()

    query = Producto.query.filter_by(empresa_id=empresa_id)
    if filtro_aprobacion == 'aprobados':
        query = query.filter(Producto.is_approved == True)
    elif filtro_aprobacion == 'pendientes':
        query = query.filter(Producto.is_approved == False)
    elif filtro_aprobacion == 'todos':
        pass
    else:
        query = query.filter(Producto.is_approved == True)

    if termino_busqueda:
        query = query.filter(
            or_(
                Producto.nombre.ilike(f"%{termino_busqueda}%"),
                Producto.categoria.ilike(f"%{termino_busqueda}%"),
                Producto.codigo_barras_externo.ilike(f"%{termino_busqueda}%")
            )
        )
    productos = query.all()
    total_productos = Producto.query.filter_by(empresa_id=empresa_id).count()

    # Calcular costo promedio y días hasta caducidad para cada producto
    for producto in productos:
        # Inicializar los atributos temporales
        producto.costo_promedio = producto.costo  # Valor predeterminado
        producto.proximo_lote_dias = None  # Valor predeterminado
        
        try:
            # Buscar lotes activos con stock positivo
            lotes_activos = LoteInventario.query.filter(
                LoteInventario.producto_id == producto.id,
                LoteInventario.esta_activo == True,
                LoteInventario.stock > 0
            ).all()
            
            if lotes_activos:
                # Cálculo del costo promedio ponderado por cantidad en stock
                costo_total = 0
                stock_total = 0
                
                for lote in lotes_activos:
                    costo_total += lote.costo_unitario * lote.stock
                    stock_total += lote.stock
                
                if stock_total > 0:
                    producto.costo_promedio = costo_total / stock_total
                
                # Buscar el próximo lote a caducar entre los que tienen fecha
                lotes_con_caducidad = []
                for lote in lotes_activos:
                    if lote.fecha_caducidad is not None:
                        lotes_con_caducidad.append(lote)
                
                if lotes_con_caducidad:
                    # Ordenar por fecha de caducidad (más cercana primero)
                    lotes_con_caducidad.sort(key=lambda x: x.fecha_caducidad)
                    
                    # MODIFICACIÓN: Buscar el primer lote que no haya caducado aún
                    hoy = date.today()
                    proximo_lote = None
                    lote_caducado = None
                    
                    # Primero intentamos encontrar un lote que no haya caducado
                    for lote in lotes_con_caducidad:
                        if lote.fecha_caducidad >= hoy:
                            proximo_lote = lote
                            break
                        elif lote_caducado is None or lote.fecha_caducidad > lote_caducado.fecha_caducidad:
                            lote_caducado = lote  # Guardamos el lote caducado más reciente
                    
                    # Si no encontramos ningún lote sin caducar, usamos el más recientemente caducado
                    if proximo_lote is None and lote_caducado is not None:
                        proximo_lote = lote_caducado
                    elif proximo_lote is None and len(lotes_con_caducidad) > 0:
                        proximo_lote = lotes_con_caducidad[0]  # Último recurso
                    
                    # Si encontramos un lote para mostrar, calculamos los días
                    if proximo_lote:
                        # Calcular días hasta caducidad directamente
                        if proximo_lote.fecha_caducidad < hoy:
                            dias_calculados = -1 * (hoy - proximo_lote.fecha_caducidad).days
                        else:
                            dias_calculados = (proximo_lote.fecha_caducidad - hoy).days
                        
                        # Asignar el valor a la propiedad del producto
                        producto.proximo_lote_dias = dias_calculados

        except Exception as e:
            # No interrumpir el proceso por un error en un producto
            continue
    
    # Crear diccionarios con todos los campos necesarios
    productos_dict = []
    for p in productos:
        # Calcular días hasta caducidad directamente
        dias_calculados = None
        hoy = date.today()
        
        lotes_activos = LoteInventario.query.filter(
            LoteInventario.producto_id == p.id,
            LoteInventario.esta_activo == True,
            LoteInventario.stock > 0
        ).all()
        
        # Buscar lotes con fecha de caducidad
        proximo_lote = None
        lote_caducado = None

        # Primero buscamos un lote que NO haya caducado
        for lote in lotes_activos:
            if lote.fecha_caducidad is not None:
                # Primero buscamos lotes que no hayan caducado
                if lote.fecha_caducidad >= hoy:
                    if proximo_lote is None or lote.fecha_caducidad < proximo_lote.fecha_caducidad:
                        proximo_lote = lote
                # También guardamos el lote caducado más reciente como respaldo
                elif lote_caducado is None or lote.fecha_caducidad > lote_caducado.fecha_caducidad:
                    lote_caducado = lote

        # Si no encontramos lote sin caducar, usamos el caducado más reciente
        if proximo_lote is None and lote_caducado is not None:
            proximo_lote = lote_caducado
        
        # Calcular días si hay lote con fecha
        if proximo_lote and proximo_lote.fecha_caducidad:
            if proximo_lote.fecha_caducidad < hoy:
                dias_calculados = -1 * (hoy - proximo_lote.fecha_caducidad).days
            else:
                dias_calculados = (proximo_lote.fecha_caducidad - hoy).days
        
        # Crear diccionario con datos seguros - INCLUYENDO NUEVOS CAMPOS
        p_dict = {
            'id': p.id,
            'nombre': p.nombre,
            'stock': p.stock,
            'costo': p.costo,
            'precio_venta': p.precio_venta,
            'precio_final': p.precio_final,  # ✅ Ahora usa directamente el campo de la BD
            'categoria': p.categoria,
            'categoria_color': p.categoria_color,
            'foto': p.foto,
            'codigo_barras_externo': p.codigo_barras_externo,
            'marca': p.marca,
            'es_favorito': p.es_favorito,
            'esta_a_la_venta': p.esta_a_la_venta,
            'unidad': getattr(p, 'unidad', None),
            'costo_promedio': getattr(p, 'costo_promedio', p.costo),
            'proximo_lote_dias': dias_calculados,  # Usar valor calculado directamente
            'tiene_descuento': getattr(p, 'tiene_descuento', False),
            'tipo_descuento': getattr(p, 'tipo_descuento', None),
            'valor_descuento': getattr(p, 'valor_descuento', 0.0),
            # NUEVOS CAMPOS DE RASTREO
            'origen_descuento': getattr(p, 'origen_descuento', None),
            'descuento_grupo_id': getattr(p, 'descuento_grupo_id', None),
            'fecha_aplicacion_descuento': getattr(p, 'fecha_aplicacion_descuento', None)
        }
        
        productos_dict.append(p_dict)

    return render_template(
        'productos.html',
        productos=productos_dict,
        filtro_aprobacion=filtro_aprobacion,
        termino_busqueda=termino_busqueda,
        total_productos=total_productos
    )

@productos_bp.route('/nuevo-producto')
@login_requerido
def nuevo_producto():
    return render_template('nuevo_producto.html')

@productos_bp.route('/agregar-producto', methods=['GET','POST'])
@login_requerido
def agregar_producto():
    if request.method == 'POST':
        try:
            # Recoger campos del formulario
            codigo_barras_externo = request.form.get("codigo_barras_externo", "").strip()
            nombre = request.form.get("nombre", "").strip()
            stock_str = request.form.get("stock", "0").strip() 
            costo_str = request.form.get("costo", "$0").strip()
            precio_str = request.form.get("precio_venta", "$0").strip()
            marca = request.form.get("marca", "").strip()
            
            # NUEVO: Verificar duplicados si hay código de barras
            if codigo_barras_externo:
                # Buscar productos con el mismo código (solo en los productos de esta empresa)
                producto_existente = Producto.query.filter_by(
                    empresa_id=session['user_id'],
                    codigo_barras_externo=codigo_barras_externo
                ).first()
                
                if producto_existente:
                    # Si existe, devolver mensaje de error
                    flash(f'Error: Ya existe un producto con el código de barras {codigo_barras_externo}', 'danger')
                    return redirect(url_for('productos.agregar_producto'))
            
            # Categoría
            cat_option = request.form.get('categoria_option', 'existente')
            if cat_option == 'existente':
                categoria = request.form.get('categoria_existente', '').strip()
            else:
                categoria = request.form.get('categoria_nueva', '').strip()
            
            # MODIFICADO: Usar la nueva función de normalización
            categoria_normalizada = normalize_category(categoria)
            
            # Favorito y a la venta
            es_favorito_bool = (request.form.get("es_favorito", "0") == "1")
            esta_a_la_venta_bool = (request.form.get("esta_a_la_venta", "1") == "1")
            
            # Caducidad
            has_caducidad = (request.form.get("toggle_caducidad_estado", "DESACTIVADO") == "ACTIVADO")
            caducidad_lapso = request.form.get("caducidad_lapso", None) if has_caducidad else None
            
            # Parse numéricos
            try:
                stock_int = int(stock_str)
            except:
                stock_int = 0
            
            costo_val = parse_money(costo_str)
            precio_val = parse_money(precio_str)
            
            # ===== MANEJO SIMPLIFICADO DE IMÁGENES =====
            foto_final = None
            
            # 1. Primero intentar usar el archivo subido
            if request.files.get('foto') and request.files['foto'].filename:
                file = request.files['foto']
                if allowed_file(file.filename):
                    # Generar nombre único
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    
                    # Guardar archivo
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    
                    file.save(os.path.join(UPLOAD_FOLDER, filename))
                    foto_final = filename
            
            # 2. Si no hay archivo, usar displayed_image_url o ia_foto_filename
            if not foto_final:
                displayed_url = request.form.get("displayed_image_url", "").strip()
                ia_filename = request.form.get("ia_foto_filename", "").strip()
                
                if ia_filename and os.path.exists(os.path.join(UPLOAD_FOLDER, ia_filename)):
                    # Usar el archivo IA que ya existe
                    foto_final = ia_filename
                elif displayed_url:
                    # Es una URL - intentar descargar o usar nombre de archivo
                    if "/uploads/" in displayed_url:
                        # Es local, extraer nombre
                        foto_final = displayed_url.split("/uploads/")[-1]
                    elif displayed_url.startswith(("http://", "https://")):
                        # Es externa, descargar
                        try:
                            ext = "jpg"  # Default
                            filename = f"{uuid.uuid4().hex}.{ext}"
                            filepath = os.path.join(UPLOAD_FOLDER, filename)
                            
                            # Descarga síncrona para asegurar que se complete
                            response = requests.get(displayed_url, timeout=10)
                            if response.status_code == 200:
                                if not os.path.exists(os.path.dirname(filepath)):
                                    os.makedirs(os.path.dirname(filepath))
                                
                                with open(filepath, "wb") as f:
                                    f.write(response.content)
                                
                                foto_final = filename
                        except Exception as e:
                            print(f"Error descargando imagen: {e}")
            
            # 3. Si todo falla, usar imagen predeterminada
            if not foto_final:
                foto_final = "default_product.jpg"
            
            # Usar la nueva función para truncar la URL para asegurar que quepa en la columna
            url_imagen_truncada = truncar_url(request.form.get("displayed_image_url", "").strip(), 95)
                
            # Crear el producto
            nuevo = Producto(
                nombre=nombre,
                stock=stock_int,
                costo=costo_val, 
                precio_venta=precio_val,
                categoria=categoria_normalizada,
                # MODIFICADO: Usar la nueva función de color de categoría
                categoria_color=get_category_color(categoria_normalizada),
                foto=foto_final,
                url_imagen=url_imagen_truncada,  # Usando la URL truncada
                is_approved=True,
                empresa_id=session['user_id'],
                codigo_barras_externo=codigo_barras_externo,
                marca=marca,
                es_favorito=es_favorito_bool,
                esta_a_la_venta=esta_a_la_venta_bool,
                has_caducidad=has_caducidad,
                metodo_caducidad=caducidad_lapso
            )
            
            # Guardar en la base de datos
            db.session.add(nuevo)
            db.session.commit()
            
            # ===== MODIFICADO: Crear siempre un lote y movimiento inicial si hay stock positivo =====
            if nuevo.stock > 0:
                try:
                    # Determinar la fecha de caducidad según el método
                    fecha_caducidad = None
                    if nuevo.has_caducidad and nuevo.metodo_caducidad:
                        fecha_actual = datetime.now()

                        if nuevo.metodo_caducidad == '1 día':
                            fecha_caducidad = fecha_actual + timedelta(days=1)
                        elif nuevo.metodo_caducidad == '3 días':
                            fecha_caducidad = fecha_actual + timedelta(days=3)
                        elif nuevo.metodo_caducidad == '1 semana':
                            fecha_caducidad = fecha_actual + timedelta(days=7)
                        elif nuevo.metodo_caducidad == '2 semanas':
                            fecha_caducidad = fecha_actual + timedelta(days=14)
                        elif nuevo.metodo_caducidad == '1 mes':
                            fecha_caducidad = fecha_actual + timedelta(days=30)
                        elif nuevo.metodo_caducidad == '3 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=90)
                        elif nuevo.metodo_caducidad == '6 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=180)
                        elif nuevo.metodo_caducidad == '1 año':
                            fecha_caducidad = fecha_actual + timedelta(days=365)
                        elif nuevo.metodo_caducidad == '2 años':
                            fecha_caducidad = fecha_actual + timedelta(days=730)
                        elif nuevo.metodo_caducidad == '3 años' or nuevo.metodo_caducidad == 'más de 3 años':
                            fecha_caducidad = fecha_actual + timedelta(days=1460)

                        # Convertir a date
                        if fecha_caducidad:
                            fecha_caducidad = fecha_caducidad.date()
                    
                    # Crear el lote inicial y el movimiento inicial
                    # Usamos la función crear_lote_registro existente
                    movimiento, lote = crear_lote_registro(
                        producto=nuevo,
                        cantidad=nuevo.stock,
                        costo=nuevo.costo,
                        fecha_caducidad=fecha_caducidad,
                        usuario_id=session['user_id']
                    )
                    
                    db.session.commit()
                    print(f"Lote inicial y movimiento creados para producto {nuevo.id}")

                except Exception as e:
                    db.session.rollback()
                    print(f"Error al crear lote inicial y movimiento: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    # Continuamos aunque haya error en la creación del lote
            
            flash('Producto guardado exitosamente', 'success')
            
            # MODIFICADO: Redireccionar a la página de confirmación
            response = make_response(redirect(url_for('productos.producto_confirmacion', producto_id=nuevo.id)))
            response.set_cookie('pagina_origen', 'agregar_producto', max_age=300)  # Cookie válida por 5 minutos
            return response
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar el producto: {str(e)}', 'danger')
            print(f"ERROR al guardar producto: {str(e)}")
        
        return redirect(url_for('productos.ver_productos'))
        
    else:
        # GET - Cargar página de agregar producto
        categorias_db = (
            db.session.query(Producto.categoria, Producto.categoria_color)
            .filter(Producto.categoria.isnot(None))
            .filter(Producto.categoria != '')
            .distinct()
            .limit(50)  # Limitar cantidad para que sea más rápido
            .all()
        )
        categories_list = [
            {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
            for c in categorias_db
        ]
        return render_template('agregar_producto.html', categories=categories_list)

@productos_bp.route('/agregar-sin-codigo', methods=['GET','POST'])
@login_requerido
def agregar_sin_codigo():
    if request.method == 'POST':
        try:
            # Recoger campos del formulario
            nombre = request.form.get("nombre", "").strip()
            stock_str = request.form.get("stock", "0").strip() 
            costo_str = request.form.get("costo", "$0").strip()
            precio_str = request.form.get("precio_venta", "$0").strip()
            marca = request.form.get("marca", "").strip()
            
            # VALIDACIÓN: Verificar longitud del nombre
            if len(nombre) > 100:
                flash('El nombre del producto no puede exceder 100 caracteres', 'danger')
                return redirect(url_for('productos.agregar_sin_codigo'))
            
            # Obtener el código de barras externo del formulario
            codigo_barras_externo = request.form.get("codigo_barras_externo", "").strip()
            
            # Verificar si el código está presente, si no, generar uno nuevo
            if not codigo_barras_externo:
                codigo_barras_externo = generar_codigo_unico()
                
            print(f"DEBUG: Usando código de barras: {codigo_barras_externo}")
            
            # Categoría
            cat_option = request.form.get('categoria_option', 'existente')
            if cat_option == 'existente':
                categoria = request.form.get('categoria_existente', '').strip()
            else:
                categoria = request.form.get('categoria_nueva', '').strip()
                
            # MODIFICADO: Usar la nueva función de normalización
            categoria_normalizada = normalize_category(categoria)
            
            # Favorito y a la venta
            es_favorito_bool = (request.form.get("es_favorito", "0") == "1")
            esta_a_la_venta_bool = (request.form.get("esta_a_la_venta", "1") == "1")
            
            # Caducidad
            has_caducidad = (request.form.get("toggle_caducidad_estado", "DESACTIVADO") == "ACTIVADO")
            caducidad_lapso = request.form.get("caducidad_lapso", None) if has_caducidad else None
            
            # Parse numéricos con VALIDACIÓN DE RANGOS
            try:
                stock_int = int(stock_str.replace(",", ""))  # Remover comas antes de convertir
                
                # VALIDACIÓN: Verificar rango del stock
                if stock_int < 0:
                    flash('El stock no puede ser negativo', 'danger')
                    return redirect(url_for('productos.agregar_sin_codigo'))
                    
                if stock_int > 99999999:  # Máximo 8 dígitos
                    flash('El stock no puede ser mayor a 99,999,999 unidades', 'danger')
                    return redirect(url_for('productos.agregar_sin_codigo'))
                    
            except ValueError:
                flash('El stock debe ser un número entero válido', 'danger')
                return redirect(url_for('productos.agregar_sin_codigo'))
            
            # Parse y validación de costo
            costo_val = parse_money(costo_str)
            
            # VALIDACIÓN: Verificar rango del costo
            if costo_val < 0:
                flash('El costo no puede ser negativo', 'danger')
                return redirect(url_for('productos.agregar_sin_codigo'))
                
            if costo_val > Decimal('99999999.99'):
                flash('El costo no puede ser mayor a $99,999,999.99', 'danger')
                return redirect(url_for('productos.agregar_sin_codigo'))
            
            # Parse y validación de precio
            precio_val = parse_money(precio_str)
            
            # VALIDACIÓN: Verificar rango del precio
            if precio_val < 0:
                flash('El precio no puede ser negativo', 'danger')
                return redirect(url_for('productos.agregar_sin_codigo'))
                
            if precio_val > Decimal('99999999.99'):
                flash('El precio no puede ser mayor a $99,999,999.99', 'danger')
                return redirect(url_for('productos.agregar_sin_codigo'))
            
            # Manejo de imagen - USANDO LA VERSIÓN MEJORADA
            foto_final = process_image(request, UPLOAD_FOLDER, BASE_DIR)
            if not foto_final:
                foto_final = "default_product.jpg"
            
            # Usar la función para truncar la URL para asegurar que quepa en la columna
            url_imagen_truncada = truncar_url(request.form.get("displayed_image_url", "").strip(), 295)
                
            # Crear el producto con solo los campos válidos
            nuevo = Producto(
                nombre=nombre,
                stock=stock_int,
                costo=float(costo_val),  # Convertir Decimal a float para la BD
                precio_venta=float(precio_val),  # Convertir Decimal a float para la BD
                categoria=categoria_normalizada,
                # MODIFICADO: Usar la nueva función de color de categoría
                categoria_color=get_category_color(categoria_normalizada),
                foto=foto_final,
                url_imagen=url_imagen_truncada,
                is_approved=True,
                empresa_id=session['user_id'],
                codigo_barras_externo=codigo_barras_externo,
                marca=marca,
                es_favorito=es_favorito_bool,
                esta_a_la_venta=esta_a_la_venta_bool,
                has_caducidad=has_caducidad,
                metodo_caducidad=caducidad_lapso
                # Eliminados los campos que no existen en el modelo
                # tipo_medida, unidad_medida, fabricacion, origen
            )
            
            # IMPORTANTE: Actualizar precio_final basado en precio_venta
            nuevo.actualizar_precio_final()
            
            # Guardar en la base de datos
            db.session.add(nuevo)
            db.session.commit()
            
            # ===== MODIFICADO: Crear siempre un lote y movimiento inicial si hay stock positivo =====
            if nuevo.stock > 0:
                try:
                    # Determinar la fecha de caducidad según el método
                    fecha_caducidad = None
                    if nuevo.has_caducidad and nuevo.metodo_caducidad:
                        fecha_actual = datetime.now()

                        if nuevo.metodo_caducidad == '1 día':
                            fecha_caducidad = fecha_actual + timedelta(days=1)
                        elif nuevo.metodo_caducidad == '3 días':
                            fecha_caducidad = fecha_actual + timedelta(days=3)
                        elif nuevo.metodo_caducidad == '1 semana':
                            fecha_caducidad = fecha_actual + timedelta(days=7)
                        elif nuevo.metodo_caducidad == '2 semanas':
                            fecha_caducidad = fecha_actual + timedelta(days=14)
                        elif nuevo.metodo_caducidad == '1 mes':
                            fecha_caducidad = fecha_actual + timedelta(days=30)
                        elif nuevo.metodo_caducidad == '3 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=90)
                        elif nuevo.metodo_caducidad == '6 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=180)
                        elif nuevo.metodo_caducidad == '1 año':
                            fecha_caducidad = fecha_actual + timedelta(days=365)
                        elif nuevo.metodo_caducidad == '2 años':
                            fecha_caducidad = fecha_actual + timedelta(days=730)
                        elif nuevo.metodo_caducidad == '3 años' or nuevo.metodo_caducidad == 'más de 3 años':
                            fecha_caducidad = fecha_actual + timedelta(days=1460)

                        # Convertir a date
                        if fecha_caducidad:
                            fecha_caducidad = fecha_caducidad.date()
                    
                    # Crear el lote inicial y el movimiento inicial
                    movimiento, lote = crear_lote_registro(
                        producto=nuevo,
                        cantidad=nuevo.stock,
                        costo=nuevo.costo,
                        fecha_caducidad=fecha_caducidad,
                        usuario_id=session['user_id']
                    )
                    
                    db.session.commit()
                    print(f"Lote inicial y movimiento creados para producto {nuevo.id}")

                except Exception as e:
                    db.session.rollback()
                    print(f"Error al crear lote inicial y movimiento: {str(e)}")
                    import traceback
                    traceback.print_exc()
                        
            flash('Producto sin código de barras guardado exitosamente', 'success')
            
            # MODIFICADO: Redireccionar a la página de confirmación
            response = make_response(redirect(url_for('productos.producto_confirmacion', producto_id=nuevo.id)))
            response.set_cookie('pagina_origen', 'agregar_sin_codigo', max_age=300)  # Cookie válida por 5 minutos
            return response
            
        except ValueError as ve:
            # Error de conversión de valores
            db.session.rollback()
            flash(f'Error en los valores ingresados: {str(ve)}', 'danger')
            print(f"ERROR de valores al guardar producto: {str(ve)}")
            return redirect(url_for('productos.agregar_sin_codigo'))
            
        except Exception as e:
            db.session.rollback()
            
            # Manejo especial para errores de base de datos
            error_message = str(e)
            if "numeric field overflow" in error_message:
                flash('Error: Los valores ingresados son demasiado grandes. Por favor, verifica los montos.', 'danger')
            elif "value too long for type character varying" in error_message:
                flash('Error: Uno de los textos ingresados es demasiado largo. Por favor, acorta los textos.', 'danger')
            elif "list index out of range" in error_message:
                flash('Error al procesar la imagen. Por favor, intenta con otra imagen o sin imagen.', 'danger')
            else:
                flash(f'Error al guardar el producto: {str(e)}', 'danger')
                
            print(f"ERROR al guardar producto sin código: {str(e)}")
            return redirect(url_for('productos.agregar_sin_codigo'))
        
    else:
        # GET - Cargar página de agregar producto sin código
        categorias_db = (
            db.session.query(Producto.categoria, Producto.categoria_color)
            .filter(Producto.categoria.isnot(None))
            .filter(Producto.categoria != '')
            .distinct()
            .limit(50)
            .all()
        )
        categories_list = [
            {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
            for c in categorias_db
        ]
        return render_template('agregar_sin_codigo.html', categories=categories_list)

@productos_bp.route('/agregar-a-granel', methods=['GET', 'POST'])
@login_requerido
def agregar_a_granel():
    if request.method == 'POST':
        try:
            # Recoger campos del formulario
            nombre = request.form.get("nombre", "").strip()
            stock_str = request.form.get("stock", "0").strip() 
            costo_str = request.form.get("costo", "$0").strip()
            precio_str = request.form.get("precio_venta", "$0").strip()
            marca = request.form.get("marca", "").strip()
            
            # Obtener unidad de medida (ya no usamos tipo_medida ni origen directamente)
            unidad_medida = request.form.get("unidad_medida", "").strip()
            
            # Generar código de barras único para productos a granel
            codigo_barras_externo = request.form.get("codigo_barras_externo", "").strip()
            if not codigo_barras_externo:
                codigo_barras_externo = generar_codigo_a_granel()
            
            # Categoría
            cat_option = request.form.get('categoria_option', 'existente')
            if cat_option == 'existente':
                categoria = request.form.get('categoria_existente', '').strip()
            else:
                categoria = request.form.get('categoria_nueva', '').strip()
                
            # MODIFICADO: Usar la nueva función de normalización
            categoria_normalizada = normalize_category(categoria)
            
            # Favorito y a la venta
            es_favorito_bool = (request.form.get("es_favorito", "0") == "1")
            esta_a_la_venta_bool = (request.form.get("esta_a_la_venta", "1") == "1")
            
            # Caducidad
            has_caducidad = (request.form.get("toggle_caducidad_estado", "DESACTIVADO") == "ACTIVADO")
            caducidad_lapso = request.form.get("caducidad_lapso", None) if has_caducidad else None
            
            # Parse numéricos
            try:
                stock_int = float(stock_str)  # Cambiado a float para soportar decimales
            except:
                stock_int = 0
            
            costo_val = parse_money(costo_str)
            precio_val = parse_money(precio_str)
            
            # ===== MANEJO SIMPLIFICADO DE IMÁGENES =====
            foto_final = None
            
            # 1. Primero intentar usar el archivo subido
            if request.files.get('foto') and request.files['foto'].filename:
                file = request.files['foto']
                if allowed_file(file.filename):
                    # Generar nombre único
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    
                    # Guardar archivo
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    
                    file.save(os.path.join(UPLOAD_FOLDER, filename))
                    foto_final = filename
            
            # 2. Si no hay archivo, usar displayed_image_url o ia_foto_filename
            if not foto_final:
                displayed_url = request.form.get("displayed_image_url", "").strip()
                ia_filename = request.form.get("ia_foto_filename", "").strip()
                
                if ia_filename and os.path.exists(os.path.join(UPLOAD_FOLDER, ia_filename)):
                    # Usar el archivo IA que ya existe
                    foto_final = ia_filename
                elif displayed_url:
                    # Es una URL - intentar descargar o usar nombre de archivo
                    if "/uploads/" in displayed_url:
                        # Es local, extraer nombre
                        foto_final = displayed_url.split("/uploads/")[-1]
                    elif displayed_url.startswith(("http://", "https://")):
                        # Es externa, descargar
                        try:
                            ext = "jpg"  # Default
                            filename = f"{uuid.uuid4().hex}.{ext}"
                            filepath = os.path.join(UPLOAD_FOLDER, filename)
                            
                            # Descarga síncrona para asegurar que se complete
                            response = requests.get(displayed_url, timeout=10)
                            if response.status_code == 200:
                                if not os.path.exists(os.path.dirname(filepath)):
                                    os.makedirs(os.path.dirname(filepath))
                                
                                with open(filepath, "wb") as f:
                                    f.write(response.content)
                                
                                foto_final = filename
                        except Exception as e:
                            print(f"Error descargando imagen: {e}")
            
            # 3. Si todo falla, usar imagen predeterminada
            if not foto_final:
                foto_final = "default_product.jpg"
            
            # Usar la nueva función para truncar la URL para asegurar que quepa en la columna
            url_imagen_truncada = truncar_url(request.form.get("displayed_image_url", "").strip(), 95)
            
            # Crear el producto - Solo usar campos que existen en el modelo
            nuevo = Producto(
                nombre=nombre,
                stock=stock_int,
                costo=costo_val, 
                precio_venta=precio_val,
                categoria=categoria_normalizada,
                # MODIFICADO: Usar la nueva función de color de categoría
                categoria_color=get_category_color(categoria_normalizada),
                foto=foto_final,
                url_imagen=url_imagen_truncada,
                is_approved=True,
                empresa_id=session['user_id'],
                codigo_barras_externo=codigo_barras_externo,
                marca=marca,
                es_favorito=es_favorito_bool,
                esta_a_la_venta=esta_a_la_venta_bool,
                has_caducidad=has_caducidad,
                metodo_caducidad=caducidad_lapso,
                unidad=unidad_medida  # Usar unidad correctamente
                # Ya no usamos "origen" que no existe en el modelo
            )
            
            # Guardar en la base de datos
            db.session.add(nuevo)
            db.session.commit()
            
            # ===== MODIFICADO: Crear siempre un lote y movimiento inicial si hay stock positivo =====
            if nuevo.stock > 0:
                try:
                    # Determinar la fecha de caducidad según el método
                    fecha_caducidad = None
                    if nuevo.has_caducidad and nuevo.metodo_caducidad:
                        fecha_actual = datetime.now()

                        if nuevo.metodo_caducidad == '1 día':
                            fecha_caducidad = fecha_actual + timedelta(days=1)
                        elif nuevo.metodo_caducidad == '3 días':
                            fecha_caducidad = fecha_actual + timedelta(days=3)
                        elif nuevo.metodo_caducidad == '1 semana':
                            fecha_caducidad = fecha_actual + timedelta(days=7)
                        elif nuevo.metodo_caducidad == '2 semanas':
                            fecha_caducidad = fecha_actual + timedelta(days=14)
                        elif nuevo.metodo_caducidad == '1 mes':
                            fecha_caducidad = fecha_actual + timedelta(days=30)
                        elif nuevo.metodo_caducidad == '3 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=90)
                        elif nuevo.metodo_caducidad == '6 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=180)
                        elif nuevo.metodo_caducidad == '1 año':
                            fecha_caducidad = fecha_actual + timedelta(days=365)
                        elif nuevo.metodo_caducidad == '2 años':
                            fecha_caducidad = fecha_actual + timedelta(days=730)
                        elif nuevo.metodo_caducidad == '3 años' or nuevo.metodo_caducidad == 'más de 3 años':
                            fecha_caducidad = fecha_actual + timedelta(days=1460)

                        # Convertir a date
                        if fecha_caducidad:
                            fecha_caducidad = fecha_caducidad.date()
                    
                    # Crear el lote inicial y el movimiento inicial
                    movimiento, lote = crear_lote_registro(
                        producto=nuevo,
                        cantidad=nuevo.stock,
                        costo=nuevo.costo,
                        fecha_caducidad=fecha_caducidad,
                        usuario_id=session['user_id']
                    )
                    
                    db.session.commit()
                    print(f"Lote inicial y movimiento creados para producto a granel {nuevo.id}")

                except Exception as e:
                    db.session.rollback()
                    print(f"Error al crear lote inicial y movimiento: {str(e)}")
                    import traceback
                    traceback.print_exc()
                        
            flash('Producto a granel guardado exitosamente', 'success')
            
            # MODIFICADO: Redireccionar a la página de confirmación
            response = make_response(redirect(url_for('productos.producto_confirmacion', producto_id=nuevo.id)))
            response.set_cookie('pagina_origen', 'agregar_a_granel', max_age=300)  # Cookie válida por 5 minutos
            return response
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar el producto a granel: {str(e)}', 'danger')
            print(f"ERROR al guardar producto a granel: {str(e)}")
        
        return redirect(url_for('productos.ver_productos'))
        
    else:
        # GET - Cargar página de agregar producto a granel
        categorias_db = (
            db.session.query(Producto.categoria, Producto.categoria_color)
            .filter(Producto.categoria.isnot(None))
            .filter(Producto.categoria != '')
            .distinct()
            .limit(50)
            .all()
        )
        categories_list = [
            {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
            for c in categorias_db
        ]
        return render_template('agregar_a_granel.html', categories=categories_list)

@productos_bp.route('/editar-producto/<int:prod_id>', methods=['GET','POST'])
@login_requerido
def editar_producto(prod_id):
    """Edita un producto existente."""
    producto = Producto.query.get_or_404(prod_id)
    
    # Verificar que el producto pertenece a la empresa actual
    if producto.empresa_id != session.get('user_id'):
        flash('No tienes permiso para editar este producto.', 'danger')
        return redirect(url_for('productos.ver_productos'))
    
    if request.method == 'POST':
        try:
            # Recoger SOLO los datos que vamos a modificar
            nombre = request.form.get("nombre", "").strip()
            precio_str = request.form.get("precio_venta", "$0").strip()
            marca = request.form.get("marca", "").strip()
            codigo_barras = request.form.get("codigo_barras_externo", "").strip()
            
            # Categoría
            categoria = request.form.get('categoria_existente', '').strip()
            categoria_normalizada = normalize_category(categoria)
            
            # Procesar precio
            precio_val = parse_money(precio_str)
            
            # Manejo de imagen
            nueva_foto = None
            
            # 1. Intentar usar el archivo subido
            if request.files.get('foto') and request.files['foto'].filename:
                file = request.files['foto']
                if allowed_file(file.filename):
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    
                    file.save(os.path.join(UPLOAD_FOLDER, filename))
                    nueva_foto = filename
            
            # 2. Si no hay archivo, usar displayed_image_url o ia_foto_filename
            if not nueva_foto:
                displayed_url = request.form.get("displayed_image_url", "").strip()
                ia_filename = request.form.get("ia_foto_filename", "").strip()
                
                if ia_filename and os.path.exists(os.path.join(UPLOAD_FOLDER, ia_filename)):
                    nueva_foto = ia_filename
                elif displayed_url:
                    if "/uploads/" in displayed_url:
                        nueva_foto = displayed_url.split("/uploads/")[-1]
                    elif displayed_url.startswith(("http://", "https://")):
                        try:
                            ext = "jpg"  # Default
                            filename = f"{uuid.uuid4().hex}.{ext}"
                            filepath = os.path.join(UPLOAD_FOLDER, filename)
                            
                            response = requests.get(displayed_url, timeout=10)
                            if response.status_code == 200:
                                if not os.path.exists(os.path.dirname(filepath)):
                                    os.makedirs(os.path.dirname(filepath))
                                
                                with open(filepath, "wb") as f:
                                    f.write(response.content)
                                
                                nueva_foto = filename
                        except Exception as e:
                            print(f"Error descargando imagen: {e}")
            
            url_imagen_truncada = truncar_url(request.form.get("displayed_image_url", "").strip(), 95)
            
            # Actualizar SOLO los campos que queremos modificar
            producto.nombre = nombre
            producto.precio_venta = precio_val
            producto.categoria = categoria_normalizada
            producto.categoria_color = get_category_color(categoria_normalizada)
            producto.codigo_barras_externo = codigo_barras
            producto.marca = marca
            
            # Actualizar foto solo si hay una nueva
            if nueva_foto:
                producto.foto = nueva_foto
                producto.url_imagen = url_imagen_truncada
            
            # Guardar en la base de datos
            db.session.commit()
            flash('Producto actualizado exitosamente', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el producto: {str(e)}', 'danger')
            print(f"ERROR al actualizar producto: {str(e)}")
        
        return redirect(url_for('productos.ver_productos'))
    
    # GET - Cargar página de editar producto
    categorias_db = (
        db.session.query(Producto.categoria, Producto.categoria_color)
        .filter(Producto.categoria.isnot(None))
        .filter(Producto.categoria != '')
        .distinct()
        .limit(50)
        .all()
    )
    categories_list = [
        {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
        for c in categorias_db
    ]
    
    # Verificar si la categoría del producto existe en la lista
    cat_encontrada = any(cat["nombre"] == producto.categoria for cat in categories_list)
    
    return render_template(
        'editar_producto.html', 
        producto=producto,
        categories=categories_list,
        cat_encontrada=cat_encontrada
    )

@productos_bp.route('/eliminar-producto/<int:prod_id>')
@login_requerido
def eliminar_producto(prod_id):
    """Elimina un producto del inventario y todos sus datos relacionados."""
    try:
        # Verificar que el producto existe y pertenece a la empresa actual
        producto = Producto.query.get_or_404(prod_id)
        
        if producto.empresa_id != session.get('user_id'):
            flash('No tienes permiso para eliminar este producto.', 'danger')
            return redirect(url_for('productos.ver_productos'))
        
        # 1. Eliminar las relaciones entre lotes y movimientos
        LoteMovimientoRelacion.query.filter(
            LoteMovimientoRelacion.lote_id.in_(
                db.session.query(LoteInventario.id).filter_by(producto_id=prod_id)
            )
        ).delete(synchronize_session=False)
        
        # 2. Eliminar todos los movimientos de inventario del producto
        MovimientoInventario.query.filter_by(producto_id=prod_id).delete()
        
        # 3. Eliminar todos los lotes de inventario del producto
        LoteInventario.query.filter_by(producto_id=prod_id).delete()
        
        # 4. Finalmente, eliminar el producto
        db.session.delete(producto)
        
        # Confirmar todos los cambios
        db.session.commit()
        flash('Producto y todos sus datos relacionados eliminados correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el producto: {str(e)}', 'danger')
        print(f"ERROR al eliminar producto y datos relacionados: {str(e)}")
    
    # Esta línea es crucial - asegura que siempre se devuelva una respuesta
    return redirect(url_for('productos.ver_productos'))

@productos_bp.route('/eliminar-todos-productos')
@login_requerido
def eliminar_todos_productos():
    """Elimina TODOS los productos del inventario de la empresa y todos sus datos relacionados."""
    try:
        # Verificar que el usuario está logueado
        empresa_id = session.get('user_id')
        if not empresa_id:
            flash('No tienes permiso para realizar esta acción.', 'danger')
            return redirect(url_for('productos.ver_productos'))
        
        # Obtener todos los productos de la empresa
        productos = Producto.query.filter_by(empresa_id=empresa_id).all()
        
        if not productos:
            flash('No hay productos para eliminar.', 'info')
            return redirect(url_for('productos.ver_productos'))
        
        productos_eliminados = 0
        
        # Eliminar cada producto y sus datos relacionados
        for producto in productos:
            try:
                # 1. Eliminar las relaciones entre lotes y movimientos de este producto
                LoteMovimientoRelacion.query.filter(
                    LoteMovimientoRelacion.lote_id.in_(
                        db.session.query(LoteInventario.id).filter_by(producto_id=producto.id)
                    )
                ).delete(synchronize_session=False)
                
                # 2. Eliminar todos los movimientos de inventario del producto
                MovimientoInventario.query.filter_by(producto_id=producto.id).delete()
                
                # 3. Eliminar todos los lotes de inventario del producto
                LoteInventario.query.filter_by(producto_id=producto.id).delete()
                
                # 4. Finalmente, eliminar el producto
                db.session.delete(producto)
                
                productos_eliminados += 1
                
            except Exception as e:
                print(f"ERROR al eliminar producto {producto.id}: {str(e)}")
                # Continuar con los demás productos en caso de error en uno específico
                continue
        
        # Confirmar todos los cambios
        db.session.commit()
        
        if productos_eliminados > 0:
            flash(f'Se eliminaron {productos_eliminados} productos y todos sus datos relacionados correctamente', 'success')
        else:
            flash('No se pudo eliminar ningún producto.', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar los productos: {str(e)}', 'danger')
        print(f"ERROR al eliminar todos los productos: {str(e)}")
    
    # Siempre redirigir a la página de productos
    return redirect(url_for('productos.ver_productos'))

@productos_bp.route('/pendientes_aprobacion')
@login_requerido
def pendientes_aprobacion():
    empresa_id = session['user_id']
    pendientes = Producto.query.filter_by(
        empresa_id=empresa_id, 
        is_approved=False
    ).all()
    return render_template('pendientes_aprobacion.html', pendientes=pendientes)

@productos_bp.route('/completar-datos/<int:prod_id>', methods=['GET','POST'])
@login_requerido
def completar_datos(prod_id):
    producto = Producto.query.get_or_404(prod_id)
    
    # Verificar que el producto pertenece a la empresa actual
    if producto.empresa_id != session.get('user_id'):
        flash('No tienes permiso para editar este producto.', 'danger')
        return redirect(url_for('productos.pendientes_aprobacion'))
    
    if request.method == 'POST':
        try:
            # Actualizar campos...
            producto.nombre = request.form.get('nombre', '')
            producto.stock = int(request.form.get('stock', 0))
            producto.costo = float(request.form.get('costo', 0))
            producto.precio_venta = float(request.form.get('precio_venta', 0))
            
            # Categoría
            cat_option = request.form.get('categoria_option')
            if cat_option == 'existente':
                categoria = request.form.get('categoria_existente', '')
            else:
                categoria = request.form.get('categoria_nueva', '')
                
            # MODIFICADO: Usar la nueva función para normalizar la categoría
            categoria_normalizada = normalize_category(categoria)
            producto.categoria = categoria_normalizada
            
            # MODIFICADO: Usar la nueva función para obtener el color de la categoría
            producto.categoria_color = get_category_color(categoria_normalizada)
            
            # Aprobar el producto
            producto.is_approved = True
            
            db.session.commit()
            flash('Producto aprobado y actualizado con éxito.')
            return redirect(url_for('productos.pendientes_aprobacion'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}')
    
    # Obtener lista de categorías
    categorias_db = (
        db.session.query(Producto.categoria, Producto.categoria_color)
        .filter(Producto.categoria.isnot(None))
        .filter(Producto.categoria != '')
        .distinct()
        .all()
    )
    categories_list = [
        {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
        for c in categorias_db
    ]
    
    # Verificar si la categoría del producto existe en la lista
    cat_encontrada = any(cat["nombre"] == producto.categoria for cat in categories_list)
    
    return render_template(
        'completar_datos.html',
        producto=producto,
        categories=categories_list,
        cat_encontrada=cat_encontrada
    )

@productos_bp.route('/producto-confirmacion/<int:producto_id>')
@login_requerido
def producto_confirmacion(producto_id):
    """Muestra la confirmación de un producto recién agregado."""
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    empresa_id = session.get('user_id')
    
    # Obtener el producto
    producto = Producto.query.get_or_404(producto_id)
    
    # Verificar que el producto pertenezca a la empresa del usuario actual
    if producto.empresa_id != empresa_id:
        flash('No tienes permiso para ver este producto.', 'danger')
        return redirect(url_for('productos.ver_productos'))
    
    # Determinar la página de origen para el botón "Volver"
    # Lo almacenamos en una cookie temporal
    pagina_origen = request.cookies.get('pagina_origen', 'agregar_producto')
    
    return render_template(
        'producto_confirmacion.html',
        producto=producto,
        pagina_origen=pagina_origen
    )

@productos_bp.route('/etiquetas-producto/<int:producto_id>', methods=['GET'])
@login_requerido
def etiquetas_producto(producto_id):
    """
    Vista para generar etiquetas para un producto específico.
    """
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    empresa_id = session.get('user_id')
    
    # Obtener el producto
    producto = Producto.query.filter_by(
        id=producto_id, 
        empresa_id=empresa_id
    ).first_or_404()
    
    # Calcular precio con descuento si tiene
    try:
        precio_final = producto.precio_venta or 0
        if hasattr(producto, 'tiene_descuento') and producto.tiene_descuento:
            if producto.tipo_descuento == 'percentage':
                precio_final = producto.precio_venta * (1 - producto.valor_descuento / 100)
            else:  # fixed amount
                precio_final = max(0, producto.precio_venta - producto.valor_descuento)
    except:
        precio_final = producto.precio_venta or 0
    
    # CRÍTICO: Verificar y asegurar que hay un código de barras válido
    if not producto.codigo_barras_externo or producto.codigo_barras_externo.strip() == "":
        # Intentar generar un código temporal
        try:
            # Prioridad 1: Usar código interno si existe
            if hasattr(producto, 'codigo_interno') and producto.codigo_interno:
                producto.codigo_barras_externo = producto.codigo_interno
            # Prioridad 2: Generar un código basado en el ID y timestamp
            else:
                from datetime import datetime
                import uuid
                # Crear código compatible con códigos de barras (solo números)
                codigo_temp = f"9{producto.id:08d}{int(datetime.now().timestamp()) % 10000:04d}"
                producto.codigo_barras_externo = codigo_temp
            
            # Guardar en base de datos
            db.session.commit()
            flash('Se ha generado un código de barras temporal para este producto.', 'warning')
        except Exception as e:
            flash(f'No se pudo generar un código de barras: {str(e)}', 'danger')
    
    # Formatos de etiquetas disponibles
    formatos_etiquetas = [
        {
            "id": "formato1",
            "nombre": "Estándar (50x30mm)",
            "descripcion": "Ideal para productos medianos y grandes",
            "dimensiones": "50mm x 30mm",
            "elementos": ["Nombre", "Precio", "Código de barras"],
            "icono": "fas fa-barcode"
        },
        {
            "id": "formato2",
            "nombre": "Compacta (40x25mm)",
            "descripcion": "Perfecta para productos pequeños",
            "dimensiones": "40mm x 25mm",
            "elementos": ["Nombre corto", "Precio", "Código QR"],
            "icono": "fas fa-qrcode"
        },
        {
            "id": "formato3",
            "nombre": "Detallada (60x40mm)",
            "descripcion": "Con información adicional del producto",
            "dimensiones": "60mm x 40mm",
            "elementos": ["Nombre", "Marca", "Precio", "Código", "Descripción"],
            "icono": "fas fa-tags"
        },
        {
            "id": "formato4",
            "nombre": "Mini (30x20mm)",
            "descripcion": "Solo lo esencial para espacios reducidos",
            "dimensiones": "30mm x 20mm",
            "elementos": ["Precio", "Código"],
            "icono": "fas fa-compress"
        }
    ]
    
    # Impresoras compatibles
    impresoras_compatibles = [
        {
            "marca": "Brother",
            "modelos": ["QL-800", "QL-810W", "QL-820NWB"],
            "tipo": "Etiquetas adhesivas"
        },
        {
            "marca": "DYMO",
            "modelos": ["LabelWriter 450", "LabelWriter 4XL"],
            "tipo": "Etiquetas térmicas"
        },
        {
            "marca": "Zebra",
            "modelos": ["GK420d", "ZD220", "ZD421"],
            "tipo": "Transferencia térmica"
        },
        {
            "marca": "TSC",
            "modelos": ["TDP-225", "TE200", "DA200"],
            "tipo": "Térmica directa"
        }
    ]
    
    return render_template(
        'etiquetas_producto.html',
        producto=producto,
        precio_final=precio_final,
        formatos_etiquetas=formatos_etiquetas,
        impresoras_compatibles=impresoras_compatibles
    )

@productos_bp.route('/actualizar-lotes-caducidad', methods=['GET'])
@login_requerido
def actualizar_lotes_caducidad():
    """
    Ruta para actualizar la lógica de próximo lote a caducar en productos existentes.
    Esta función recorre todos los productos con lotes y actualiza la información de caducidad.
    """
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    empresa_id = session.get('user_id')
    
    try:
        # Obtener todos los productos de la empresa
        productos = Producto.query.filter_by(empresa_id=empresa_id).all()
        
        productos_actualizados = 0
        lotes_procesados = 0
        
        for producto in productos:
            # Buscar lotes activos con stock positivo
            lotes_activos = LoteInventario.query.filter(
                LoteInventario.producto_id == producto.id,
                LoteInventario.esta_activo == True,
                LoteInventario.stock > 0
            ).all()
            
            if not lotes_activos:
                continue
                
            # Filtrar lotes con fecha de caducidad
            lotes_con_caducidad = [lote for lote in lotes_activos if lote.fecha_caducidad is not None]
            
            if not lotes_con_caducidad:
                continue
                
            # Ordenar por fecha de caducidad (más cercana primero)
            lotes_con_caducidad.sort(key=lambda x: x.fecha_caducidad)
            
            # Buscar el primer lote que no haya caducado aún
            hoy = date.today()
            lote_no_caducado = None
            
            for lote in lotes_con_caducidad:
                if lote.fecha_caducidad >= hoy:
                    lote_no_caducado = lote
                    break
            
            # Si no hay lote sin caducar, usar el que caducó más recientemente
            if not lote_no_caducado and lotes_con_caducidad:
                lote_no_caducado = lotes_con_caducidad[0]
            
            # Si encontramos un lote para mostrar, lo marcamos como "proximo_lote"
            if lote_no_caducado:
                # Opcional: Actualizar algún campo en el producto o lote si es necesario
                lotes_procesados += 1
            
            productos_actualizados += 1
                
        return f"""
        <h2>Actualización de Lotes Completada</h2>
        <p>Se procesaron {productos_actualizados} productos y {lotes_procesados} lotes con fechas de caducidad.</p>
        <p>Los cambios en la lógica de caducidad ahora están activos.</p>
        <p><a href="{url_for('productos.ver_productos')}">Volver a la lista de productos</a></p>
        """
        
    except Exception as e:
        return f"""
        <h2>Error al Actualizar Lotes</h2>
        <p>Se produjo un error: {str(e)}</p>
        <p><a href="{url_for('productos.ver_productos')}">Volver a la lista de productos</a></p>
        """

@productos_bp.route('/exportar-productos-excel')
@login_requerido
def exportar_productos_excel():
    """
    Exporta todos los productos de la empresa a un archivo Excel con formato profesional.
    """
    try:
        empresa_id = session['user_id']
        
        # Obtener todos los productos aprobados de la empresa
        productos_db = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True
        ).order_by(Producto.categoria, Producto.nombre).all()
        
        if not productos_db:
            flash('No hay productos para exportar', 'warning')
            return redirect(url_for('productos.ver_productos'))
        
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Definir encabezados
        headers = [
            'PRODUCTO', 'SKU', 'MARCA', 'CATEGORÍA', 'STOCK', 
            'PRECIO FINAL', 'ÚLTIMO COSTO', 'COSTO PROMEDIO', 
            'PRÓX LOTE CADUCA EN', 'FAVORITO', 'A LA VENTA'
        ]
        
        # Agregar encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # Estilo del encabezado: fondo azul oscuro con texto blanco
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Procesar cada producto
        for row_num, producto in enumerate(productos_db, 2):
            try:
                # Calcular costo promedio para este producto
                lotes_activos = LoteInventario.query.filter(
                    LoteInventario.producto_id == producto.id,
                    LoteInventario.esta_activo == True,
                    LoteInventario.stock > 0
                ).all()
                
                costo_promedio = producto.costo or 0
                if lotes_activos:
                    costo_total = sum((lote.costo_unitario or 0) * (lote.stock or 0) for lote in lotes_activos)
                    stock_total = sum(lote.stock or 0 for lote in lotes_activos)
                    if stock_total > 0:
                        costo_promedio = costo_total / stock_total
                
                # Calcular días hasta próxima caducidad
                dias_caducidad = "N/A"
                if lotes_activos:
                    from datetime import date
                    hoy = date.today()
                    lotes_con_caducidad = [l for l in lotes_activos if l.fecha_caducidad]
                    
                    if lotes_con_caducidad:
                        # Ordenar por fecha de caducidad
                        lotes_con_caducidad.sort(key=lambda x: x.fecha_caducidad)
                        
                        # Buscar el primer lote que no haya caducado
                        proximo_lote = None
                        for lote in lotes_con_caducidad:
                            if lote.fecha_caducidad >= hoy:
                                proximo_lote = lote
                                break
                        
                        # Si no hay lote sin caducar, usar el más recientemente caducado
                        if not proximo_lote and lotes_con_caducidad:
                            proximo_lote = lotes_con_caducidad[0]
                        
                        if proximo_lote:
                            if proximo_lote.fecha_caducidad < hoy:
                                dias_diff = (hoy - proximo_lote.fecha_caducidad).days
                                dias_caducidad = f"Caducado hace {dias_diff} días"
                            elif proximo_lote.fecha_caducidad == hoy:
                                dias_caducidad = "Caduca hoy"
                            else:
                                dias_diff = (proximo_lote.fecha_caducidad - hoy).days
                                dias_caducidad = f"{dias_diff} días"
                
                # Preparar datos de la fila - asegurar que todos sean valores básicos
                row_data = [
                    str(producto.nombre or "Sin nombre"),
                    str(producto.codigo_barras_externo or f"SKU-{producto.id:05d}"),
                    str(producto.marca or "No definida"),
                    str(producto.categoria or "Sin categoría"),
                    float(producto.stock or 0),
                    float(producto.precio_final or producto.precio_venta or 0),
                    float(producto.costo or 0),
                    float(costo_promedio),
                    str(dias_caducidad),
                    "Sí" if producto.es_favorito else "No",
                    "Sí" if producto.esta_a_la_venta else "No"
                ]
                
                # Agregar datos a la fila
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col, value=value)
                
            except Exception as e:
                print(f"Error procesando producto {producto.id}: {str(e)}")
                continue
        
        # Ajustar ancho de columnas
        column_widths = [25, 18, 18, 20, 12, 15, 15, 15, 20, 12, 12]
        for col, width in enumerate(column_widths, 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[column_letter].width = width
        
        # Aplicar filtros automáticos solo al rango de datos
        if len(productos_db) > 0:
            ws.auto_filter.ref = f"A1:K{len(productos_db) + 1}"
        
        # Crear buffer en memoria para el archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Preparar respuesta con el archivo
        from datetime import datetime
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"productos_inventario_{fecha_actual}.xlsx"
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        print(f"Error en exportar_productos_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al generar el archivo Excel: {str(e)}', 'danger')
        return redirect(url_for('productos.ver_productos'))

# ==============================
# FUNCIONES PARA EXPORTAR
# ==============================
__all__ = ['productos_bp', 'generar_codigo_unico', 'generar_codigo_a_granel', 'truncar_url']