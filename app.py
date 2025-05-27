import os
import random
import string
import unicodedata
import uuid
import time
import threading
import sys
import shutil
import urllib.parse
from io import BytesIO

from flask import Flask, request, render_template, session, redirect, url_for, jsonify, flash, send_file, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps

# 1) IMPORTA Flask-Migrate
from flask_migrate import Migrate

# NUEVO: Importar Flask-Compress y Flask-Caching
from flask_compress import Compress
from flask_caching import Cache

from models import db
# Importamos los modelos necesarios directamente
from models.models import Empresa, CodigoDisponible, CodigoAsignado, Producto, CatalogoGlobal
# Importamos los modelos de inventario
from models.modelos_inventario import MovimientoInventario, LoteInventario, LoteMovimientoRelacion

# (NUEVO) IMPORTA la función para leer Google Sheets
from sheets import leer_hoja  # Asegúrate de tener 'sheets.py' con leer_hoja()

# Importar funciones de utils
from utils import (
    process_image, async_download_image, parse_money, normalize_categoria_if_needed,
    optimize_db_session, cleanup_db_session, ensure_default_images,
    find_similar_product_image, find_similar_catalog_image, normalizar_categoria,
    obtener_o_generar_color_categoria
)

# NUEVO: Importar las funciones del nuevo archivo category_colors.py
from category_colors import get_category_color, normalize_category

# Importamos el blueprint de ajuste_stock
from ajuste_stock import init_app as init_ajuste_stock, crear_lote_registro, obtener_proximo_numero_lote

# Integración OPENAI (si la usas para otros fines)
import openai
openai.api_key = "sk-proj-KXZSGDJ6bGMjVUZXzGp1r3ZYfvUvpkVFbUVqPyWeJc1sxsEjeyodfaLEZOuq5Nc6RYV1f1JvyT3BlbkFJQo22FAJuvP6bF7Z4BQ3nsuEA"

# Integración SERPAPI
import requests
SERPAPI_API_KEY = "84d269bfa51876a1a092ace371d89f7dc2500d8c5a61b420c08d96e5351f5c79"

from sqlalchemy import or_, and_, func
from datetime import datetime, date, timedelta, time


# =========================================
# NUEVO: Conjunto de Categorías con Emojis
# =========================================
VALID_CATEGORIES = {
    "abarrotes secos",
    "enlatados y conservas",
    "botanas, dulces y snacks",
    "bebidas no alcohólicas",
    "bebidas alcohólicas",
    "panadería y repostería",
    "lácteos y huevos",
    "carnes frías y embutidos",
    "congelados y refrigerados",
    "frutas y verduras frescas",
    "productos de limpieza y hogar",
    "cuidado personal e higiene",
    "medicamentos de mostrador (otc)",
    "productos para bebés",
    "productos para mascotas",
    "artículos de papelería",
    "ferretería básica",
    "artesanías y manualidades",
    "productos a granel",
    "productos orgánicos",
    "productos gourmet",
    "suplementos alimenticios",
    "otros (miscelánea)"
}

# Configuración de subida de archivos
BASE_DIR = os.path.dirname(os.path.realpath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip', 'rar'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def download_image_with_headers(url, filepath):
    """
    Intenta descargar una imagen con diferentes cabeceras y agentes de usuario.
    Algunas páginas bloquean descargas sin un User-Agent adecuado.
    """
    headers_list = [
        # Firefox en Windows
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/115.0",
            "Accept": "image/jpeg, image/png, image/*;q=0.8",
            "Referer": url
        },
        # Chrome en Mac
        {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8"
        },
        # Safari en iPhone
        {
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1"
        }
    ]
    
    for headers in headers_list:
        try:
            print(f"DEBUG: Intentando descargar {url} con cabeceras: {headers}")
            response = requests.get(url, headers=headers, stream=True, timeout=10)
            if response.status_code == 200:
                if not os.path.exists(os.path.dirname(filepath)):
                    os.makedirs(os.path.dirname(filepath))
                
                with open(filepath, "wb") as f:
                    for chunk in response.iter_content(1024):
                        f.write(chunk)
                print(f"DEBUG: Descarga exitosa para {url}")
                return True
            print(f"DEBUG: Error al descargar {url}, código {response.status_code}")
        except Exception as e:
            print(f"DEBUG: Excepción al descargar {url}: {e}")
    
    return False

##############################################
# FUNCIONES AUXILIARES
##############################################
def generar_color_aleatorio() -> str:
    r = lambda: random.randint(0, 255)
    return f"#{r():02X}{r():02X}{r():02X}"

def buscar_imagen_google_images(nombre_producto, codigo_barras_externo=None):
    """
    Uso de SerpAPI (engine=google_images) para buscar imagen y guardarla localmente.
    """
    query = nombre_producto
    if codigo_barras_externo and not codigo_barras_externo.startswith("INT-"):
        query += f" {codigo_barras_externo}"

    params = {
        "engine": "google_images",
        "q": query,
        "api_key": SERPAPI_API_KEY,
        "ijn": "0",
        "num": "9"
    }
    print(f"Buscando imagen en Google Images para: {query}")
    print(f"URL SerpAPI: https://serpapi.com/search?{urllib.parse.urlencode(params)}")
    
    try:
        response = requests.get("https://serpapi.com/search", params=params)
        # Verificar el código de estado HTTP
        if response.status_code != 200:
            print(f"Error: SerpAPI devolvió código {response.status_code}")
            print(f"Respuesta: {response.text}")
            return None
            
        data = response.json()
        
        # Verificar si hay un error en la respuesta JSON
        if "error" in data:
            print(f"Error de SerpAPI: {data['error']}")
            return None
            
        # Verificar si hay resultados de imágenes
        images = data.get("images_results", [])
        if not images:
            print("No se encontraron imágenes para esta consulta")
            return None
            
        # Usar la primera imagen encontrada
        first_image = images[0]
        image_url = first_image.get("thumbnail") or first_image.get("original")
        
        if not image_url:
            print("URL de imagen no encontrada en la respuesta")
            return None
            
        # Descargar la imagen
        print(f"Intentando descargar imagen desde: {image_url}")
        ext = image_url.rsplit('.', 1)[-1].split('?')[0]
        if ext not in ALLOWED_EXTENSIONS:
            ext = "jpg"
        filename = secure_filename(f"{random.randint(100000,999999)}.{ext}")
        
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        # Intentar descargar con diferentes cabeceras
        if download_image_with_headers(image_url, filepath):
            print("Imagen descargada y guardada como:", filename)
            return filename
        else:
            print(f"No se pudo descargar la imagen desde {image_url}")
            return None
    except Exception as e:
        print(f"Error completo al buscar imagen con google_images: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

##############################################
# GPT (opcional) - sin cambios
##############################################
def buscar_titulos_serpapi(codigo_barras):
    """(Opcional) Búsqueda en Google con SerpAPI para extraer títulos de resultados orgánicos."""
    params = {
        "engine": "google",
        "q": codigo_barras,
        "api_key": SERPAPI_API_KEY,
        "hl": "es",
        "num": "10"
    }
    resp = requests.get("https://serpapi.com/search", params=params)
    data = resp.json()
    titles = []
    if "organic_results" in data:
        for item in data["organic_results"]:
            t = item.get("title")
            if t:
                titles.append(t)
    print(f"\n=== DEBUG SERPAPI TITLES PARA: {codigo_barras} ===")
    for t in titles:
        print(" -", t)
    print("=== FIN LISTADO TITULOS ===\n")
    return titles

def gpt_extraer_nombre_categoria(titulos):
    """
    (Opcional) Usa GPT para determinar nombre y categoría a partir de títulos de Google.
    """
    if not titulos:
        return {"nombre": "Producto IA", "categoria": "General"}
    prompt = f"""
Los siguientes títulos provienen de una búsqueda en Google sobre un código de barras:

{titulos}

Por favor, deduce un posible nombre de producto y una categoría.
Si no hay información clara, inventa algo genérico.
Devuélveme un JSON con claves "nombre" y "categoria".
Ejemplo: {{"nombre": "Coca Cola 600 ml", "categoria": "Bebidas"}}
"""
    try:
        print("\n=== Llamando a GPT con estos títulos: ===")
        for t in titulos:
            print(" -", t)
        print("=== FIN LISTADO TITULOS ===\n")
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=200,
            temperature=0.3
        )
        content = response.choices[0].message.content.strip()
        import json
        data = json.loads(content)
        nombre = data.get("nombre", "Producto IA")
        categoria = data.get("categoria", "General")
        return {"nombre": nombre, "categoria": categoria}
    except Exception as e:
        print("Error GPT:", e)
        return {"nombre": "Producto IA", "categoria": "General"}

def buscar_nombre_categoria_por_barcode(codigo_barras):
    titulos = buscar_titulos_serpapi(codigo_barras)
    return gpt_extraer_nombre_categoria(titulos)

# Función para truncar URLs largas de manera segura
def truncar_url(url, max_length=95):
    """
    Trunca una URL para que quepa en un campo de longitud limitada.
    Preserva el protocolo y dominio para URLs externas.
    """
    if not url or len(url) <= max_length:
        return url
    
    # Verificar si es una URL externa
    if url.startswith(('http://', 'https://')):
        try:
            # Dividir la URL en partes
            from urllib.parse import urlparse
            parsed_url = urlparse(url)
            base = f"{parsed_url.scheme}://{parsed_url.netloc}"
            path = parsed_url.path
            query = parsed_url.query
            
            # Si solo la base ya es demasiado larga
            if len(base) >= max_length - 5:
                return base[:max_length-5] + "..."
            
            # Calcular cuánto espacio queda para el path y query
            remaining = max_length - len(base) - 4  # -4 para "/..."
            
            # Extraer el nombre del archivo
            filename = path.split('/')[-1] if '/' in path else path
            
            # Si hay espacio para el nombre completo y no hay query
            if len(filename) <= remaining and not query:
                return f"{base}/.../{filename}"
            
            # Si hay query, incluirla parcialmente si hay espacio
            if query and len(filename) + len(query) + 1 <= remaining:  # +1 por el "?"
                return f"{base}/.../{filename}?{query}"
            
            # Si no hay espacio suficiente para todo, truncar adecuadamente
            if len(filename) <= remaining:
                return f"{base}/.../{filename}"
            
            # Si el nombre es demasiado largo, truncarlo
            return f"{base}/.../{filename[:remaining]}"
        except:
            # Si hay algún error, caer al método original pero preservando el dominio
            pass
    
    # Método original para URLs que no son externas o en caso de error
    try:
        # Intentar preservar al menos el dominio para URLs externas
        if url.startswith(('http://', 'https://')):
            from urllib.parse import urlparse
            parsed = urlparse(url)
            domain = parsed.netloc
            if domain and len(domain) + 8 <= max_length:  # 8 caracteres para "https://" y "..."
                return f"{parsed.scheme}://{domain}/..."
            
    except:
        pass
    
    # Si todo falla, usar el método original
    nombre_archivo = url.split('/')[-1]
    if len(nombre_archivo) > max_length:
        return nombre_archivo[-max_length:]
    return nombre_archivo

##############################################
# FUNCIÓN PARA GENERAR CÓDIGO ÚNICO PARA PRODUCTOS A GRANEL
##############################################
def generar_codigo_a_granel():
    """
    Genera un código único para productos a granel con formato GRA-XXXXXXXX.
    """
    prefix = "GRA-"
    caracteres = string.ascii_uppercase + string.digits
    codigo_aleatorio = ''.join(random.choice(caracteres) for _ in range(8))
    return prefix + codigo_aleatorio

##############################################
# FUNCIÓN PARA GENERAR CÓDIGO ÚNICO PARA PRODUCTOS SIN CÓDIGO DE BARRAS
##############################################
def generar_codigo_unico():
    """
    Genera un código único para productos sin código de barras con formato 1901XXXXXXXX.
    """
    # MODIFICADO: Ahora genera códigos empezando con "1901" seguido de 8 dígitos en lugar de "INT-"
    prefix = "1901"
    # MODIFICADO: Ahora usa solo dígitos para los 8 caracteres siguientes
    caracteres = string.digits
    codigo_aleatorio = ''.join(random.choice(caracteres) for _ in range(8))
    return prefix + codigo_aleatorio

##############################################
# FLASK APP
##############################################
app = Flask(__name__)
app.config['SECRET_KEY'] = "super_secreto"
BASE_DIR = os.path.dirname(os.path.realpath(__file__))

app.config['SQLALCHEMY_DATABASE_URI'] = (
    "postgresql://benjaminbecerriles:HammeredEnd872@localhost:5432/inventario_db"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

db.init_app(app)
migrate = Migrate(app, db)

# NUEVO: Configurar compresión GZIP
Compress(app)

# NUEVO: Configurar cache
cache = Cache(app, config={
    'CACHE_TYPE': 'simple',
    'CACHE_DEFAULT_TIMEOUT': 300
})

# Inicializar el módulo de ajuste de stock
init_ajuste_stock(app)

##############################
# Sistema de navegación "Volver" con rutas predefinidas
##############################
# Mapa de navegación: endpoint actual -> ruta destino del botón volver
MAPA_NAVEGACION = {
    # Grupo 1: Vuelven a dashboard/inventario
    'nuevo_producto': 'dashboard_inventario',
    'ajuste_stock': 'dashboard_inventario',
    'ajuste_stock.index': 'dashboard_inventario',
    'ajuste_stock.ajuste_stock': 'dashboard_inventario',  # NUEVA LÍNEA CORRECTA  
    'cambiar_precios': 'dashboard_inventario',
    'descuentos': 'dashboard_inventario',
    'ubicacion_productos': 'dashboard_inventario',
    'ver_productos': 'dashboard_inventario',
    
    # Grupo 2: Vuelven a nuevo-producto
    'agregar_producto': 'nuevo_producto',
    'agregar_sin_codigo': 'nuevo_producto',
    'agregar_a_granel': 'nuevo_producto',
    
    # Grupo 3: Vuelven a ajuste-inventario
    'historial_movimientos': 'ajuste_stock',
    'ajuste_stock.ajuste_entrada': 'ajuste_stock',
    'ajuste_stock.ajuste_salida': 'ajuste_stock',
    'ajuste_stock.ajuste_confirmacion': 'ajuste_stock',
    
    # Grupo 4: Vuelven a productos (CON LOS PREFIJOS CORRECTOS)
    'new_ajuste_stock.new_ajuste_entrada': 'ver_productos',
    'new_ajuste_stock.new_ajuste_salida': 'ver_productos',
    'new_ajuste_stock.new_ajuste_confirmacion': 'ver_productos',
    'cambiar_costos': 'ver_productos',
    'editar_producto': 'ver_productos',
    'etiquetas_producto': 'ver_productos',
}

@app.context_processor
def inject_navigation():
    """Inyecta la información de navegación predefinida en todos los templates"""
    current_endpoint = request.endpoint
    
    # DEBUG - Imprimir el endpoint actual
    print(f"DEBUG - Endpoint actual: {current_endpoint}")
    
    # Buscar si el endpoint actual tiene una ruta de retorno definida
    back_route = None
    show_button = False
    
    if current_endpoint in MAPA_NAVEGACION:
        back_route = url_for(MAPA_NAVEGACION[current_endpoint])
        show_button = True
        print(f"DEBUG - Ruta de retorno encontrada: {back_route}")
    else:
        print(f"DEBUG - No se encontró ruta de retorno para: {current_endpoint}")
    
    # No mostrar en login, registro o index
    if current_endpoint in ['login', 'registro', 'index']:
        show_button = False
    
    return {
        'previous_page': back_route,
        'show_back_button': show_button
    }

# NUEVO: Filtro Jinja2 para obtener el color de categoría
@app.template_filter('category_color')
def category_color_filter(category):
    return get_category_color(category)

# MODIFICADO: Sobrescribir la función existente para usar la nueva lógica
def obtener_o_generar_color_categoria(categoria):
    """
    Función sobrescrita que ahora utiliza get_category_color de category_colors.py
    """
    return get_category_color(categoria)

# MODIFICADO: Actualizar la función normalizar_categoria para usar la nueva
def normalizar_categoria(categoria):
    """
    Función sobrescrita que ahora utiliza normalize_category de category_colors.py
    """
    return normalize_category(categoria)

# Inicialización antes de la primera solicitud
# Inicialización en la creación de la aplicación
with app.app_context():
    # Asegurar que existan las imágenes predeterminadas
    ensure_default_images(BASE_DIR, UPLOAD_FOLDER)
    
    # Verificar si SerpAPI está disponible (opcional)
    try:
        requests.get('https://serpapi.com', timeout=2)
        app.config['SERPAPI_AVAILABLE'] = True
    except Exception as e:
        app.config['SERPAPI_AVAILABLE'] = False
        print(f"AVISO: SerpAPI no disponible, usando imágenes predeterminadas: {e}")

##############################
# Headers de seguridad y caché
##############################
@app.after_request
def add_security_headers(response):
    # Headers anti-cache para contenido dinámico
    if response.mimetype == 'text/html':
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private, max-age=0"
        response.headers["Expires"] = "0"
        response.headers["Pragma"] = "no-cache"
    
    # Headers de seguridad
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    
    # Cache agresivo para assets estáticos
    if response.mimetype in ['text/css', 'application/javascript', 'image/png', 'image/jpeg', 'image/webp']:
        response.headers['Cache-Control'] = 'public, max-age=31536000'
    
    return response

##############################
# Decoradores de Autenticación
##############################
def login_requerido(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrap

def admin_requerido(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        empresa_id = session.get('user_id')
        emp = Empresa.query.get(empresa_id)
        if not emp or not emp.is_admin:
            return "No tienes permisos de administrador."
        return f(*args, **kwargs)
    return wrap

##############################
# RUTAS BÁSICAS
##############################
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/registro', methods=['GET','POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form['nombre']
        email = request.form['email']
        password = request.form['password']
        codigo_ingresado = request.form.get('codigo', '')

        hashed_pw = generate_password_hash(password)
        nueva_empresa = Empresa(nombre=nombre, email=email, password=hashed_pw)
        db.session.add(nueva_empresa)
        db.session.commit()

        if codigo_ingresado:
            cod_disp = CodigoDisponible.query.filter_by(codigo=codigo_ingresado, esta_activo=True).first()
            if cod_disp:
                db.session.delete(cod_disp)
                nuevo_asig = CodigoAsignado(codigo=codigo_ingresado, esta_activo=True, empresa_id=nueva_empresa.id)
                db.session.add(nuevo_asig)
                db.session.commit()
        return redirect(url_for('login'))
    return render_template('registro.html')

@app.route('/ingresar-codigo', methods=['GET','POST'])
def ingresar_codigo():
    if not session.get('user_id'):
        return "Debes iniciar sesión primero para asignar tu código."
    emp = Empresa.query.get(session['user_id'])
    if not emp:
        return "No se encontró la empresa en la BD."

    if request.method == 'POST':
        codigo_ingresado = request.form['codigo']
        cod_disp = CodigoDisponible.query.filter_by(codigo=codigo_ingresado, esta_activo=True).first()
        if not cod_disp:
            return "Código inválido o inactivo."
        db.session.delete(cod_disp)
        nuevo_asig = CodigoAsignado(codigo=codigo_ingresado, esta_activo=True, empresa_id=emp.id)
        db.session.add(nuevo_asig)
        db.session.commit()
        return "¡Código asignado exitosamente! Ahora ya puedes usar el sistema."

    return render_template('ingresar_codigo.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        pw = request.form['password']
        emp = Empresa.query.filter_by(email=email).first()
        if not emp:
            return "Email no encontrado."
        if not check_password_hash(emp.password, pw):
            return "Contraseña incorrecta."

        session['logged_in'] = True
        session['user_id'] = emp.id
        session['user_name'] = emp.nombre

        if emp.is_admin:
            return redirect(url_for('admin_panel'))

        cod_asig = CodigoAsignado.query.filter_by(empresa_id=emp.id, esta_activo=True).first()
        if not cod_asig:
            return redirect(url_for('ingresar_codigo'))

        return redirect(url_for('dashboard_home'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

##############################
# NUEVO DASHBOARD
##############################
@app.route('/dashboard')
@login_requerido
def dashboard_home():
    """
    Vista del dashboard principal
    """
    # Obtener información del usuario y datos para el dashboard
    empresa_id = session['user_id']
    empresa = Empresa.query.get(empresa_id)
    
    # Obtener lista de productos para estadísticas
    productos = Producto.query.filter_by(empresa_id=empresa_id).all()
    total_productos = len(productos)
    
    return render_template(
        'dashboard_home.html',
        productos=productos,
        total_productos=total_productos
    )

@app.route('/ubicacion-productos', methods=['GET'])
@login_requerido
def ubicacion_productos():
    """
    Vista para gestionar la ubicación física de los productos.
    Permite asignar y ver dónde se encuentran los productos físicamente.
    """
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # Obtener información del usuario
    empresa_id = session.get('user_id')
    
    # Parámetros de filtrado
    categoria_filtro = request.args.get('categoria')
    marca_filtro = request.args.get('marca')
    global_view = request.args.get('global') == '1'
    
    # Consulta base
    productos_query = Producto.query.filter_by(
        empresa_id=empresa_id,
        is_approved=True
    )
    
    # Aplicar filtros si existen
    if categoria_filtro:
        productos_query = productos_query.filter_by(categoria=categoria_filtro)
    elif marca_filtro:
        productos_query = productos_query.filter_by(marca=marca_filtro)
    
    # Para vista de detalle no usamos paginación
    if categoria_filtro or marca_filtro or global_view:
        productos = productos_query.all()
        page = 1
        total_pages = 1
    else:
        # Parámetros de paginación para vista general
        page = request.args.get('page', 1, type=int)
        per_page = 100  # Ampliado para mostrar más productos
        productos_paginados = productos_query.paginate(page=page, per_page=per_page, error_out=False)
        productos = productos_paginados.items
        total_pages = productos_paginados.pages
    
    # Obtener categorías únicas para filtros (corregido para evitar duplicados)
    categorias_db = (
        db.session.query(Producto.categoria, Producto.categoria_color)
        .filter_by(empresa_id=empresa_id)
        .filter(Producto.categoria.isnot(None))
        .filter(Producto.categoria != '')
        .distinct()
        .all()
    )
    
    # Usar un conjunto para garantizar que no hay duplicados
    categorias_set = set()
    categorias = []
    for cat in categorias_db:
        if cat[0] not in categorias_set:
            categorias_set.add(cat[0])
            categorias.append({
                "nombre": cat[0], 
                "color": cat[1] if cat[1] else "#6B7280"
            })
    
    # Obtener marcas únicas (también evitando duplicados)
    marcas_db = (
        db.session.query(Producto.marca)
        .filter_by(empresa_id=empresa_id)
        .filter(Producto.marca.isnot(None))
        .filter(Producto.marca != '')
        .distinct()
        .all()
    )
    
    marcas = list(set([marca[0] for marca in marcas_db if marca[0]]))
    
    # Obtener ubicaciones únicas existentes para filtros
    ubicaciones_db = (
        db.session.query(Producto.ubicacion)
        .filter_by(empresa_id=empresa_id)
        .filter(Producto.ubicacion.isnot(None))
        .filter(Producto.ubicacion != '')
        .distinct()
        .all()
    )
    
    ubicaciones = list(set([ub[0] for ub in ubicaciones_db if ub[0]]))
    
    # Convertir a diccionarios para JSON en JavaScript
    productos_dict = []
    for p in productos:
        productos_dict.append({
            'id': p.id,
            'nombre': p.nombre,
            'stock': p.stock if hasattr(p, 'stock') else 0,
            'precio_venta': p.precio_venta if hasattr(p, 'precio_venta') else None,
            'categoria': p.categoria,
            'categoria_color': p.categoria_color,
            'foto': p.foto,
            'codigo_barras_externo': p.codigo_barras_externo,
            'marca': p.marca if hasattr(p, 'marca') else None,
            'ubicacion': p.ubicacion if hasattr(p, 'ubicacion') else None,
            'es_favorito': p.es_favorito if hasattr(p, 'es_favorito') else False,
            'esta_a_la_venta': p.esta_a_la_venta if hasattr(p, 'esta_a_la_venta') else True
        })
    
    return render_template(
        'ubicacion_productos.html',
        productos=productos_dict,
        page=page,
        total_pages=total_pages,
        categorias=categorias,
        marcas=marcas,
        ubicaciones=ubicaciones,
        filtro_categoria=categoria_filtro,
        filtro_marca=marca_filtro,
        vista_global=global_view
    )

@app.route('/ubicacion-detalle/<string:tipo>/<path:valor>', methods=['GET'])
@login_requerido
def ubicacion_detalle(tipo, valor):
    """
    Vista para mostrar detalles de ubicaciones por tipo (global, categoría, marca, individual)
    """
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # Obtener información del usuario
    empresa_id = session.get('user_id')
    
    # Inicializar variables
    productos = []
    titulo = "Ubicaciones"
    ubicacion_valor = ""
    
    # Consulta base
    productos_query = Producto.query.filter_by(
        empresa_id=empresa_id,
        is_approved=True
    )
    
    # Filtrar según el tipo
    if tipo == 'global':
        # Si es 'all', mostrar todos con ubicación
        if valor == 'all':
            productos_query = productos_query.filter(Producto.ubicacion.isnot(None))
            titulo = "Ubicación Global"
            
            # Obtener la ubicación más común como "global"
            ubicaciones = (
                db.session.query(Producto.ubicacion, db.func.count(Producto.id).label('total'))
                .filter(Producto.empresa_id == empresa_id)
                .filter(Producto.ubicacion.isnot(None))
                .group_by(Producto.ubicacion)
                .order_by(db.text('total DESC'))
                .first()
            )
            
            if ubicaciones:
                ubicacion_valor = ubicaciones[0]
        
    elif tipo == 'categoria':
        # Verificar si es 'all' para mostrar todas las categorías
        if valor == 'all':
            # Seleccionar productos con ubicación agrupados por categoría
            categorias_con_ubicacion = (
                db.session.query(Producto.categoria)
                .filter(Producto.empresa_id == empresa_id)
                .filter(Producto.ubicacion.isnot(None))
                .group_by(Producto.categoria)
                .all()
            )
            
            # Obtener todos los productos que tienen una ubicación y pertenecen a alguna de estas categorías
            productos_query = productos_query.filter(
                Producto.categoria.in_([cat[0] for cat in categorias_con_ubicacion if cat[0]])
            )
            
            titulo = "Ubicaciones por Categoría"
        else:
            # Filtrar por la categoría específica
            productos_query = productos_query.filter_by(categoria=valor)
            
            # Obtener la ubicación más común para esta categoría
            ubicaciones = (
                db.session.query(Producto.ubicacion, db.func.count(Producto.id).label('total'))
                .filter(Producto.empresa_id == empresa_id)
                .filter(Producto.categoria == valor)
                .filter(Producto.ubicacion.isnot(None))
                .group_by(Producto.ubicacion)
                .order_by(db.text('total DESC'))
                .first()
            )
            
            if ubicaciones:
                ubicacion_valor = ubicaciones[0]
                
            titulo = f"Categoría: {valor}"
    
    elif tipo == 'marca':
        # Verificar si es 'all' para mostrar todas las marcas
        if valor == 'all':
            # Seleccionar productos con ubicación agrupados por marca
            marcas_con_ubicacion = (
                db.session.query(Producto.marca)
                .filter(Producto.empresa_id == empresa_id)
                .filter(Producto.ubicacion.isnot(None))
                .group_by(Producto.marca)
                .all()
            )
            
            # Obtener todos los productos que tienen una ubicación y pertenecen a alguna de estas marcas
            productos_query = productos_query.filter(
                Producto.marca.in_([marca[0] for marca in marcas_con_ubicacion if marca[0]])
            )
            
            titulo = "Ubicaciones por Marca"
        else:
            # Filtrar por la marca específica
            productos_query = productos_query.filter_by(marca=valor)
            
            # Obtener la ubicación más común para esta marca
            ubicaciones = (
                db.session.query(Producto.ubicacion, db.func.count(Producto.id).label('total'))
                .filter(Producto.empresa_id == empresa_id)
                .filter(Producto.marca == valor)
                .filter(Producto.ubicacion.isnot(None))
                .group_by(Producto.ubicacion)
                .order_by(db.text('total DESC'))
                .first()
            )
            
            if ubicaciones:
                ubicacion_valor = ubicaciones[0]
                
            titulo = f"Marca: {valor}"
    
    elif tipo == 'individual':
        # Si es 'all', filtrar productos que tienen ubicación pero no están categorizados
        if valor == 'all':
            # Obtener productos con ubicación que no pertenecen a una categoría o marca dominante
            # Esto es aproximado - una implementación más precisa requeriría análisis adicional
            productos_query = productos_query.filter(Producto.ubicacion.isnot(None))
            titulo = "Ubicaciones Individuales"
        else:
            # Acá sería por producto individual (usando ID)
            try:
                prod_id = int(valor)
                productos_query = productos_query.filter_by(id=prod_id)
                titulo = "Producto Individual"
            except:
                # Si el valor no es un ID válido, mostrar todos
                pass
    
    # Obtener productos finales
    productos_db = productos_query.all()
    
    # Convertir a diccionarios para JSON
    productos = []
    for p in productos_db:
        productos.append({
            'id': p.id,
            'nombre': p.nombre,
            'stock': p.stock if hasattr(p, 'stock') else 0,
            'precio_venta': p.precio_venta if hasattr(p, 'precio_venta') else None,
            'categoria': p.categoria,
            'categoria_color': p.categoria_color,
            'foto': p.foto,
            'codigo_barras_externo': p.codigo_barras_externo,
            'marca': p.marca if hasattr(p, 'marca') else None,
            'ubicacion': p.ubicacion if hasattr(p, 'ubicacion') else None,
            'es_favorito': p.es_favorito if hasattr(p, 'es_favorito') else False,
            'esta_a_la_venta': p.esta_a_la_venta if hasattr(p, 'esta_a_la_venta') else True
        })
    
    return render_template(
        'ubicacion_detalle.html',
        productos=productos,
        titulo=titulo,
        tipo=tipo,
        valor=valor,
        ubicacion_valor=ubicacion_valor
    )

@app.route('/dashboard/inventario')
@login_requerido
def dashboard_inventario():
    """
    Vista del dashboard de inventario con valoración híbrida (lotes o producto directo)
    """
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # Obtener información del usuario
    empresa_id = session.get('user_id')
    
    # Obtener lista de productos para mostrar (solo los aprobados)
    productos_query = Producto.query.filter_by(
        empresa_id=empresa_id,
        is_approved=True
    ).order_by(Producto.id.desc())
    
    # Ejecutar query y obtener resultados
    productos_raw = productos_query.all()
    
    # Crear una lista simplificada con solo los datos necesarios
    productos = []
    total_unidades = 0
    valor_inventario = 0
    
    print("\n\n==== DIAGNÓSTICO DE PRODUCTOS Y LOTES ====")
    
    for p in productos_raw:
        try:
            p_id = p.id
            p_nombre = p.nombre
            
            # Obtener todos los lotes activos con stock positivo para este producto
            lotes_activos = LoteInventario.query.filter_by(
                producto_id=p_id,
                esta_activo=True
            ).filter(LoteInventario.stock > 0).all()
            
            # Verificar stock y costo directos del producto
            try:
                p_stock_directo = float(p.stock) if p.stock is not None else 0
                p_costo_directo = float(p.costo) if p.costo is not None else 0
            except (TypeError, ValueError):
                p_stock_directo = 0
                p_costo_directo = 0
            
            print(f"Producto [{p_id}]: {p_nombre}")
            print(f"  - Stock directo: {p_stock_directo}, Costo directo: {p_costo_directo}")
            print(f"  - Lotes activos encontrados: {len(lotes_activos)}")
            
            # CASO 1: Si hay lotes activos, usar sus valores
            if lotes_activos:
                # Calcular stock total y valor total para este producto basado en lotes
                p_stock_total = 0
                p_valor_total = 0
                
                # Procesar cada lote del producto
                for lote in lotes_activos:
                    try:
                        # Convertir a float
                        lote_stock = float(lote.stock) if lote.stock is not None else 0
                        lote_costo = float(lote.costo_unitario) if lote.costo_unitario is not None else 0
                        
                        # Calcular valor de este lote
                        lote_valor = lote_stock * lote_costo
                        
                        # Sumar al total del producto
                        p_stock_total += lote_stock
                        p_valor_total += lote_valor
                        
                        print(f"    > Lote {lote.numero_lote}: stock={lote_stock}, costo={lote_costo}, valor={lote_valor}")
                    except Exception as e:
                        print(f"    > ERROR procesando lote {lote.id}: {e}")
                
                # Verificar si el stock del producto coincide con la suma de lotes
                if abs(p_stock_directo - p_stock_total) > 0.001:  # Pequeña tolerancia para errores de redondeo
                    print(f"  ⚠️ ADVERTENCIA: Stock en producto ({p_stock_directo}) NO coincide con suma de lotes ({p_stock_total})")
                
                # Usar los valores calculados de lotes
                p_stock_final = p_stock_total
                p_valor_final = p_valor_total
                print(f"  ✅ Usando valores de lotes: stock={p_stock_final}, valor={p_valor_final}")
            
            # CASO 2: Si no hay lotes activos pero el producto tiene stock, usar valores directos
            elif p_stock_directo > 0:
                p_stock_final = p_stock_directo
                p_valor_final = p_stock_directo * p_costo_directo
                print(f"  ⚠️ No hay lotes activos, usando valores directos: stock={p_stock_final}, valor={p_valor_final}")
            
            # CASO 3: Producto sin stock
            else:
                p_stock_final = 0
                p_valor_final = 0
                print(f"  ℹ️ Producto sin stock")
            
            # Acumular totales globales
            total_unidades += p_stock_final
            valor_inventario += p_valor_final
            
            # Intentar obtener precio_venta para el diccionario
            try:
                p_precio = float(p.precio_venta) if p.precio_venta is not None else 0
            except (TypeError, ValueError):
                p_precio = 0
            
            # Agregar a la lista simplificada
            productos.append({
                'id': p_id,
                'nombre': p_nombre,
                'stock': p_stock_final,  # Usamos el stock calculado
                'precio_venta': p_precio,
                'categoria': p.categoria,
                'categoria_color': p.categoria_color,
                'foto': p.foto,
                'codigo_barras_externo': p.codigo_barras_externo,
                'marca': p.marca or '',
                'es_favorito': p.es_favorito,
                'esta_a_la_venta': p.esta_a_la_venta
            })
            
        except Exception as e:
            print(f"ERROR procesando producto: {e}")
            import traceback
            traceback.print_exc()
    
    # Calcular estadísticas
    total_productos = len(productos)
    
    # Productos con poco stock (por agotarse)
    productos_por_agotarse = sum(1 for p in productos if p['stock'] > 0 and p['stock'] <= 5)
    
    # ========== NUEVOS DATOS PARA EL DASHBOARD PROFESIONAL ==========
    
    # Calcular crecimiento mensual (comparar con mes anterior)
    from datetime import datetime, timedelta
    fecha_mes_anterior = datetime.now() - timedelta(days=30)
    
    try:
        # Contar movimientos del último mes
        movimientos_mes = MovimientoInventario.query.filter(
            MovimientoInventario.empresa_id == empresa_id,
            MovimientoInventario.fecha >= fecha_mes_anterior
        ).count()
        
        # Calcular crecimiento (simulado si no tienes datos históricos)
        if movimientos_mes > 0:
            crecimiento = 15.3  # Porcentaje positivo si hay movimiento
        else:
            crecimiento = -5.2  # Negativo si no hay actividad
    except:
        crecimiento = 0.0
    
    # Contar pedidos pendientes (órdenes de compra activas)
    try:
        pedidos_pendientes = OrdenCompra.query.filter_by(
            empresa_id=empresa_id,
            estado='pendiente'
        ).count() if 'OrdenCompra' in globals() else 0
    except:
        pedidos_pendientes = 0
    
    # Contar alertas activas
    alertas_activas = 0
    
    # Alerta 1: Productos por agotarse
    if productos_por_agotarse > 0:
        alertas_activas += 1
    
    # Alerta 2: Productos sin stock pero con ventas recientes
    productos_sin_stock = sum(1 for p in productos if p['stock'] == 0)
    if productos_sin_stock > 0:
        alertas_activas += 1
    
    # Alerta 3: Productos con caducidad próxima (si manejas caducidades)
    try:
        from datetime import datetime, timedelta
        fecha_limite = datetime.now() + timedelta(days=30)
        
        productos_por_caducar = LoteInventario.query.filter(
            LoteInventario.producto_id.in_([p['id'] for p in productos]),
            LoteInventario.fecha_caducidad <= fecha_limite,
            LoteInventario.stock > 0,
            LoteInventario.esta_activo == True
        ).count()
        
        if productos_por_caducar > 0:
            alertas_activas += 1
    except:
        pass
    
    print(f"\nRESUMEN GLOBAL:")
    print(f"TOTAL UNIDADES: {total_unidades}")
    print(f"VALOR INVENTARIO: {valor_inventario}")
    print(f"TOTAL PRODUCTOS: {total_productos}")
    print(f"PRODUCTOS POR AGOTARSE: {productos_por_agotarse}")
    print(f"CRECIMIENTO MENSUAL: {crecimiento}%")
    print(f"PEDIDOS PENDIENTES: {pedidos_pendientes}")
    print(f"ALERTAS ACTIVAS: {alertas_activas}")
    print("==== FIN DIAGNÓSTICO ====\n\n")
    
    # Redondear valores para mostrarlo correctamente
    valor_inventario_redondeado = int(round(valor_inventario))
    total_unidades_redondeado = int(round(total_unidades))
    
    # Renderizar template con todos los datos necesarios
    return render_template(
        'dashboard_inventario.html',
        productos=productos,
        total_productos=total_productos,
        total_unidades=total_unidades_redondeado,
        valor_inventario=valor_inventario_redondeado,
        productos_por_agotarse=productos_por_agotarse,
        # Nuevos datos para el dashboard profesional
        crecimiento=crecimiento,
        pedidos_pendientes=pedidos_pendientes,
        alertas_activas=alertas_activas,
        productos_agotarse=productos_por_agotarse  # Para el sidebar
    )

# ========== RUTAS AUXILIARES DEL DASHBOARD ==========

@app.route('/api/dashboard/stats')
@login_requerido
def api_dashboard_stats():
    """
    API endpoint para obtener estadísticas en tiempo real del dashboard
    """
    empresa_id = session.get('user_id')
    
    try:
        # Estadísticas rápidas
        total_productos = Producto.query.filter_by(empresa_id=empresa_id, is_approved=True).count()
        productos_activos = Producto.query.filter_by(empresa_id=empresa_id, is_approved=True, esta_a_la_venta=True).count()
        
        # Movimientos recientes (últimos 7 días)
        fecha_limite = datetime.now() - timedelta(days=7)
        movimientos_recientes = MovimientoInventario.query.filter(
            MovimientoInventario.empresa_id == empresa_id,
            MovimientoInventario.fecha >= fecha_limite
        ).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total_productos': total_productos,
                'productos_activos': productos_activos,
                'movimientos_recientes': movimientos_recientes,
                'ultima_actualizacion': datetime.now().strftime('%H:%M:%S')
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/product/<int:product_id>')
@login_requerido
def api_product_detail(product_id):
    """
    API endpoint para obtener detalles de un producto (usado por el modal del dashboard)
    """
    empresa_id = session.get('user_id')
    
    producto = Producto.query.filter_by(
        id=product_id,
        empresa_id=empresa_id
    ).first()
    
    if not producto:
        return jsonify({'success': False, 'error': 'Producto no encontrado'}), 404
    
    # Obtener lotes activos
    lotes = LoteInventario.query.filter_by(
        producto_id=product_id,
        esta_activo=True
    ).filter(LoteInventario.stock > 0).all()
    
    lotes_data = []
    for lote in lotes:
        lotes_data.append({
            'numero_lote': lote.numero_lote,
            'stock': float(lote.stock),
            'costo': float(lote.costo_unitario) if lote.costo_unitario else 0,
            'fecha_entrada': lote.fecha_entrada.strftime('%Y-%m-%d') if lote.fecha_entrada else None,
            'fecha_caducidad': lote.fecha_caducidad.strftime('%Y-%m-%d') if lote.fecha_caducidad else None
        })
    
    return jsonify({
        'success': True,
        'producto': {
            'id': producto.id,
            'nombre': producto.nombre,
            'codigo_barras': producto.codigo_barras_externo,
            'categoria': producto.categoria,
            'marca': producto.marca,
            'stock': float(producto.stock) if producto.stock else 0,
            'precio_venta': float(producto.precio_venta) if producto.precio_venta else 0,
            'costo': float(producto.costo) if producto.costo else 0,
            'ubicacion': producto.ubicacion,
            'descripcion': producto.descripcion,
            'foto': producto.foto,
            'lotes': lotes_data
        }
    })


@app.route('/api/product/<int:product_id>/quick-update', methods=['POST'])
@login_requerido
def api_quick_update_product(product_id):
    """
    API endpoint para actualización rápida de productos desde el dashboard
    """
    empresa_id = session.get('user_id')
    
    producto = Producto.query.filter_by(
        id=product_id,
        empresa_id=empresa_id
    ).first()
    
    if not producto:
        return jsonify({'success': False, 'error': 'Producto no encontrado'}), 404
    
    try:
        data = request.get_json()
        
        # Actualizar solo los campos permitidos
        if 'precio_venta' in data:
            producto.precio_venta = float(data['precio_venta'])
        
        if 'ubicacion' in data:
            producto.ubicacion = data['ubicacion']
        
        if 'esta_a_la_venta' in data:
            producto.esta_a_la_venta = bool(data['esta_a_la_venta'])
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Producto actualizado correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/cambiar-precios')
@login_requerido
def cambiar_precios():
    """
    Vista para cambiar precios de productos organizados por categorías.
    Permite búsqueda y modificación rápida de precios.
    """
    # Obtener información del usuario y datos para el dashboard
    empresa_id = session['user_id']
    
    # Obtener lista de productos aprobados
    productos_db = Producto.query.filter_by(
        empresa_id=empresa_id,
        is_approved=True
    ).order_by(Producto.categoria).all()
    
    # Convertir a diccionarios con todos los campos
    productos = []
    for p in productos_db:
        producto_dict = {
            'id': p.id,
            'nombre': p.nombre,
            'stock': p.stock or 0,
            'precio_venta': p.precio_venta or 0,
            'categoria': p.categoria or '',
            'categoria_color': p.categoria_color or '#6B7280',
            'foto': p.foto or 'default_product.jpg',
            'codigo_barras_externo': p.codigo_barras_externo or '',
            'marca': p.marca or '',
            'es_favorito': p.es_favorito or False,
            'esta_a_la_venta': p.esta_a_la_venta or True,
            'unidad': getattr(p, 'unidad', 'pieza'),
            'tiene_descuento': getattr(p, 'tiene_descuento', False),
            'tipo_descuento': getattr(p, 'tipo_descuento', None),
            'valor_descuento': getattr(p, 'valor_descuento', 0.0),
            # NUEVOS CAMPOS DE RASTREO
            'origen_descuento': getattr(p, 'origen_descuento', None),
            'descuento_grupo_id': getattr(p, 'descuento_grupo_id', None),
            'fecha_aplicacion_descuento': getattr(p, 'fecha_aplicacion_descuento', None)
        }
        productos.append(producto_dict)
    
    # Obtener categorías únicas y sus colores
    categorias = []
    categorias_set = set()
    
    for producto in productos_db:
        if producto.categoria and producto.categoria not in categorias_set:
            categorias_set.add(producto.categoria)
            categorias.append({
                'nombre': producto.categoria,
                'color': producto.categoria_color or '#6B7280' # Color por defecto si no tiene
            })
    
    # Ordenar categorías alfabéticamente
    categorias.sort(key=lambda x: x['nombre'])
    
    return render_template(
        'cambiar_precios.html',
        productos=productos,
        categorias=categorias
    )

@app.route('/cambiar-costos')
@login_requerido
def cambiar_costos_general():
    """
    Vista general para cambiar precios de productos
    """
    # Redirigir a la lista de productos para seleccionar uno
    flash('Seleccione un producto para ver y gestionar sus costos', 'info')
    return redirect(url_for('ver_productos'))

@app.route('/cambiar-costos/<int:producto_id>')
@login_requerido
def cambiar_costos(producto_id):
    """
    Vista para ver y cambiar los costos de un producto específico.
    Muestra el costo actual y el historial de lotes.
    """
    # Obtener información del usuario
    empresa_id = session.get('user_id')
    
    # Obtener el producto específico
    producto = Producto.query.filter_by(
        id=producto_id,
        empresa_id=empresa_id,
        is_approved=True
    ).first_or_404()
    
    # Calcular costo promedio
    lotes_activos = LoteInventario.query.filter(
        LoteInventario.producto_id == producto_id,
        LoteInventario.esta_activo == True,
        LoteInventario.stock > 0
    ).all()
    
    # Calcular costo promedio ponderado por cantidad en stock
    if lotes_activos:
        costo_total = 0
        stock_total = 0
        
        for lote in lotes_activos:
            costo_total += lote.costo_unitario * lote.stock
            stock_total += lote.stock
        
        if stock_total > 0:
            producto.costo_promedio = costo_total / stock_total
        else:
            producto.costo_promedio = producto.costo
    else:
        producto.costo_promedio = producto.costo
    
    # Obtener todos los lotes del producto ordenados por fecha (más reciente primero)
    lotes = LoteInventario.query.filter_by(
        producto_id=producto_id
    ).order_by(LoteInventario.fecha_entrada.desc()).all()
    
    return render_template(
        'cambiar_costos.html',
        producto=producto,
        lotes=lotes
    )

@app.route('/costo-confirmacion/<int:producto_id>')
@login_requerido
def costo_confirmacion(producto_id):
    """Muestra la confirmación después de actualizar el costo de un producto."""
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    empresa_id = session.get('user_id')
    
    # Obtener el producto
    producto = Producto.query.get_or_404(producto_id)
    
    # Verificar que el producto pertenezca a la empresa del usuario actual
    if producto.empresa_id != empresa_id:
        flash('No tienes permiso para ver este producto.', 'danger')
        return redirect(url_for('ver_productos'))
    
    # Obtener el ID del movimiento de la URL
    movimiento_id = request.args.get('movimiento_id')
    
    # Obtener el movimiento
    movimiento = MovimientoInventario.query.get_or_404(movimiento_id)
    
    # Obtener el costo anterior de la URL 
    costo_anterior = float(request.args.get('costo_anterior', 0))
    costo_nuevo = float(movimiento.costo_unitario)
    
    # Calcular costo promedio
    lotes_activos = LoteInventario.query.filter(
        LoteInventario.producto_id == producto_id,
        LoteInventario.esta_activo == True,
        LoteInventario.stock > 0
    ).all()
    
    costo_promedio = costo_nuevo  # Valor por defecto
    
    if lotes_activos:
        costo_total = 0
        stock_total = 0
        
        for lote in lotes_activos:
            costo_total += lote.costo_unitario * lote.stock
            stock_total += lote.stock
        
        if stock_total > 0:
            costo_promedio = costo_total / stock_total
    
    return render_template(
        'cambiar_costo_confirmacion.html',
        producto=producto,
        movimiento=movimiento,
        costo_anterior=costo_anterior,
        costo_nuevo=costo_nuevo,
        costo_promedio=costo_promedio
    )

@app.route('/costo-confirmacion/<int:producto_id>')
@login_requerido
def mostrar_costo_confirmacion(producto_id):  # Nombre cambiado para evitar conflictos
    """Muestra la confirmación después de actualizar el costo de un producto."""
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    empresa_id = session.get('user_id')
    
    # Obtener el producto
    producto = Producto.query.get_or_404(producto_id)
    
    # Verificar que el producto pertenezca a la empresa del usuario actual
    if producto.empresa_id != empresa_id:
        flash('No tienes permiso para ver este producto.', 'danger')
        return redirect(url_for('ver_productos'))
    
    # Obtener el ID del movimiento de la URL
    movimiento_id = request.args.get('movimiento_id')
    
    # Obtener el movimiento
    movimiento = MovimientoInventario.query.get_or_404(movimiento_id)
    
    # Obtener el costo anterior de la URL 
    costo_anterior = float(request.args.get('costo_anterior', 0))
    costo_nuevo = float(movimiento.costo_unitario)
    
    # Calcular costo promedio
    lotes_activos = LoteInventario.query.filter(
        LoteInventario.producto_id == producto_id,
        LoteInventario.esta_activo == True,
        LoteInventario.stock > 0
    ).all()
    
    costo_promedio = costo_nuevo  # Valor por defecto
    
    if lotes_activos:
        costo_total = 0
        stock_total = 0
        
        for lote in lotes_activos:
            costo_total += lote.costo_unitario * lote.stock
            stock_total += lote.stock
        
        if stock_total > 0:
            costo_promedio = costo_total / stock_total
    
    return render_template(
        'cambiar_costo_confirmacion.html',
        producto=producto,
        movimiento=movimiento,
        costo_anterior=costo_anterior,
        costo_nuevo=costo_nuevo,
        costo_promedio=costo_promedio
    )

##############################
# ADMIN
##############################
@app.route('/admin')
@admin_requerido
def admin_panel():
    return render_template('admin.html')

@app.route('/admin/empresas')
@admin_requerido
def admin_empresas():
    empresas = Empresa.query.all()
    return render_template('admin_empresas.html', empresas=empresas)

@app.route('/admin/codigos-disponibles')
@admin_requerido
def admin_disponibles():
    codigos = CodigoDisponible.query.all()
    return render_template('admin_disponibles.html', codigos=codigos)

def generar_codigo():
    letras = string.ascii_uppercase + string.digits
    return ''.join(random.choice(letras) for _ in range(8))

@app.route('/admin/generar-disponible')
@admin_requerido
def generar_disponible():
    nuevo = CodigoDisponible(codigo=generar_codigo(), esta_activo=True)
    db.session.add(nuevo)
    db.session.commit()
    return redirect(url_for('admin_disponibles'))

@app.route('/admin/eliminar-disponible/<int:cod_id>')
@admin_requerido
def eliminar_disponible(cod_id):
    cod = CodigoDisponible.query.get_or_404(cod_id)
    db.session.delete(cod)
    db.session.commit()
    return redirect(url_for('admin_disponibles'))

@app.route('/admin/codigos-asignados')
@admin_requerido
def admin_asignados():
    codigos = CodigoAsignado.query.all()
    return render_template('admin_asignados.html', codigos=codigos)

@app.route('/admin/toggle-asignado/<int:cod_id>')
@admin_requerido
def toggle_asignado(cod_id):
    cod = CodigoAsignado.query.get_or_404(cod_id)
    cod.esta_activo = not cod.esta_activo
    db.session.commit()
    return redirect(url_for('admin_asignados'))

##############################
# CRUD DE PRODUCTOS
##############################
@app.route('/productos', methods=['GET'])
@login_requerido
def ver_productos():
    empresa_id = session['user_id']
    filtro_aprobacion = request.args.get('filtroAprobacion', 'aprobados')
    termino_busqueda = request.args.get('q', '').strip()

    # Actualizamos automáticamente los lotes sin fecha de caducidad
    try:
        # Obtener productos con caducidad activada
        productos_caducidad = Producto.query.filter(
            Producto.empresa_id == empresa_id,
            Producto.has_caducidad == True,
            Producto.metodo_caducidad.isnot(None)
        ).all()
        
        fixed_count = 0
        for producto in productos_caducidad:
            # Obtener lotes activos sin fecha de caducidad
            lotes = LoteInventario.query.filter(
                LoteInventario.producto_id == producto.id,
                LoteInventario.esta_activo == True,
                LoteInventario.stock > 0,
                LoteInventario.fecha_caducidad.is_(None)
            ).all()
            
            if not lotes:
                continue
                
            # Calcular fecha de caducidad basada en metodo_caducidad
            caducidad_lapso = producto.metodo_caducidad
            fecha_actual = datetime.now()
            fecha_caducidad = None
            
            if caducidad_lapso == '1 día':
                fecha_caducidad = fecha_actual + timedelta(days=1)
            elif caducidad_lapso == '3 días':
                fecha_caducidad = fecha_actual + timedelta(days=3)
            elif caducidad_lapso == '1 semana':
                fecha_caducidad = fecha_actual + timedelta(days=7)
            elif caducidad_lapso == '2 semanas':
                fecha_caducidad = fecha_actual + timedelta(days=14)
            elif caducidad_lapso == '1 mes':
                fecha_caducidad = fecha_actual + timedelta(days=30)
            elif caducidad_lapso == '3 meses':
                fecha_caducidad = fecha_actual + timedelta(days=90)
            elif caducidad_lapso == '6 meses':
                fecha_caducidad = fecha_actual + timedelta(days=180)
            elif caducidad_lapso == '1 año':
                fecha_caducidad = fecha_actual + timedelta(days=365)
            elif caducidad_lapso == '2 años':
                fecha_caducidad = fecha_actual + timedelta(days=730)
            elif caducidad_lapso == '3 años':
                fecha_caducidad = fecha_actual + timedelta(days=1460)  # 4 años (considerados como +3 años)
            
            # Convertir a date
            if fecha_caducidad:
                fecha_caducidad = fecha_caducidad.date()
                
                # Actualizar los lotes
                for lote in lotes:
                    lote.fecha_caducidad = fecha_caducidad
                    fixed_count += 1
                    
        # Guardar cambios si se hicieron modificaciones
        if fixed_count > 0:
            db.session.commit()
    except Exception as e:
        db.session.rollback()

    query = Producto.query.filter_by(empresa_id=empresa_id)
    if filtro_aprobacion == 'aprobados':
        query = query.filter(Producto.is_approved == True)
    elif filtro_aprobacion == 'pendientes':
        query = query.filter(Producto.is_approved == False)
    elif filtro_aprobacion == 'todos':
        pass
    else:
        query = query.filter(Producto.is_approved == True)

    if termino_busqueda:
        query = query.filter(
            or_(
                Producto.nombre.ilike(f"%{termino_busqueda}%"),
                Producto.categoria.ilike(f"%{termino_busqueda}%"),
                Producto.codigo_barras_externo.ilike(f"%{termino_busqueda}%")
            )
        )
    productos = query.all()
    total_productos = Producto.query.filter_by(empresa_id=empresa_id).count()

    # Calcular costo promedio y días hasta caducidad para cada producto
    for producto in productos:
        # Inicializar los atributos temporales
        producto.costo_promedio = producto.costo  # Valor predeterminado
        producto.proximo_lote_dias = None  # Valor predeterminado
        
        try:
            # Buscar lotes activos con stock positivo
            lotes_activos = LoteInventario.query.filter(
                LoteInventario.producto_id == producto.id,
                LoteInventario.esta_activo == True,
                LoteInventario.stock > 0
            ).all()
            
            if lotes_activos:
                # Cálculo del costo promedio ponderado por cantidad en stock
                costo_total = 0
                stock_total = 0
                
                for lote in lotes_activos:
                    costo_total += lote.costo_unitario * lote.stock
                    stock_total += lote.stock
                
                if stock_total > 0:
                    producto.costo_promedio = costo_total / stock_total
                
                # Buscar el próximo lote a caducar entre los que tienen fecha
                lotes_con_caducidad = []
                for lote in lotes_activos:
                    if lote.fecha_caducidad is not None:
                        lotes_con_caducidad.append(lote)
                
                if lotes_con_caducidad:
                    # Ordenar por fecha de caducidad (más cercana primero)
                    lotes_con_caducidad.sort(key=lambda x: x.fecha_caducidad)
                    
                    # MODIFICACIÓN: Buscar el primer lote que no haya caducado aún
                    hoy = date.today()
                    proximo_lote = None
                    lote_caducado = None
                    
                    # Primero intentamos encontrar un lote que no haya caducado
                    for lote in lotes_con_caducidad:
                        if lote.fecha_caducidad >= hoy:
                            proximo_lote = lote
                            break
                        elif lote_caducado is None or lote.fecha_caducidad > lote_caducado.fecha_caducidad:
                            lote_caducado = lote  # Guardamos el lote caducado más reciente
                    
                    # Si no encontramos ningún lote sin caducar, usamos el más recientemente caducado
                    if proximo_lote is None and lote_caducado is not None:
                        proximo_lote = lote_caducado
                    elif proximo_lote is None and len(lotes_con_caducidad) > 0:
                        proximo_lote = lotes_con_caducidad[0]  # Último recurso
                    
                    # Si encontramos un lote para mostrar, calculamos los días
                    if proximo_lote:
                        # Calcular días hasta caducidad directamente
                        if proximo_lote.fecha_caducidad < hoy:
                            dias_calculados = -1 * (hoy - proximo_lote.fecha_caducidad).days
                        else:
                            dias_calculados = (proximo_lote.fecha_caducidad - hoy).days
                        
                        # Asignar el valor a la propiedad del producto
                        producto.proximo_lote_dias = dias_calculados

        except Exception as e:
            # No interrumpir el proceso por un error en un producto
            continue
    
    # Crear diccionarios con todos los campos necesarios
    productos_dict = []
    for p in productos:
        # Calcular días hasta caducidad directamente
        dias_calculados = None
        hoy = date.today()
        
        lotes_activos = LoteInventario.query.filter(
            LoteInventario.producto_id == p.id,
            LoteInventario.esta_activo == True,
            LoteInventario.stock > 0
        ).all()
        
        # Buscar lotes con fecha de caducidad
        proximo_lote = None
        lote_caducado = None

        # Primero buscamos un lote que NO haya caducado
        for lote in lotes_activos:
            if lote.fecha_caducidad is not None:
                # Primero buscamos lotes que no hayan caducado
                if lote.fecha_caducidad >= hoy:
                    if proximo_lote is None or lote.fecha_caducidad < proximo_lote.fecha_caducidad:
                        proximo_lote = lote
                # También guardamos el lote caducado más reciente como respaldo
                elif lote_caducado is None or lote.fecha_caducidad > lote_caducado.fecha_caducidad:
                    lote_caducado = lote

        # Si no encontramos lote sin caducar, usamos el caducado más reciente
        if proximo_lote is None and lote_caducado is not None:
            proximo_lote = lote_caducado
        
        # Calcular días si hay lote con fecha
        if proximo_lote and proximo_lote.fecha_caducidad:
            if proximo_lote.fecha_caducidad < hoy:
                dias_calculados = -1 * (hoy - proximo_lote.fecha_caducidad).days
            else:
                dias_calculados = (proximo_lote.fecha_caducidad - hoy).days
        
        # Crear diccionario con datos seguros - INCLUYENDO NUEVOS CAMPOS
        p_dict = {
            'id': p.id,
            'nombre': p.nombre,
            'stock': p.stock,
            'costo': p.costo,
            'precio_venta': p.precio_venta,
            'precio_final': p.precio_final,  # ✅ Ahora usa directamente el campo de la BD
            'categoria': p.categoria,
            'categoria_color': p.categoria_color,
            'foto': p.foto,
            'codigo_barras_externo': p.codigo_barras_externo,
            'marca': p.marca,
            'es_favorito': p.es_favorito,
            'esta_a_la_venta': p.esta_a_la_venta,
            'unidad': getattr(p, 'unidad', None),
            'costo_promedio': getattr(p, 'costo_promedio', p.costo),
            'proximo_lote_dias': dias_calculados,  # Usar valor calculado directamente
            'tiene_descuento': getattr(p, 'tiene_descuento', False),
            'tipo_descuento': getattr(p, 'tipo_descuento', None),
            'valor_descuento': getattr(p, 'valor_descuento', 0.0),
            # NUEVOS CAMPOS DE RASTREO
            'origen_descuento': getattr(p, 'origen_descuento', None),
            'descuento_grupo_id': getattr(p, 'descuento_grupo_id', None),
            'fecha_aplicacion_descuento': getattr(p, 'fecha_aplicacion_descuento', None)
        }
        
        productos_dict.append(p_dict)

    return render_template(
        'productos.html',
        productos=productos_dict,
        filtro_aprobacion=filtro_aprobacion,
        termino_busqueda=termino_busqueda,
        total_productos=total_productos
    )

@app.route('/exportar-productos-excel')
@login_requerido
def exportar_productos_excel():
    """
    Exporta todos los productos de la empresa a un archivo Excel con formato profesional.
    """
    try:
        empresa_id = session['user_id']
        
        # Obtener todos los productos aprobados de la empresa
        productos_db = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True
        ).order_by(Producto.categoria, Producto.nombre).all()
        
        if not productos_db:
            flash('No hay productos para exportar', 'warning')
            return redirect(url_for('ver_productos'))
        
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Definir encabezados
        headers = [
            'PRODUCTO', 'SKU', 'MARCA', 'CATEGORÍA', 'STOCK', 
            'PRECIO FINAL', 'ÚLTIMO COSTO', 'COSTO PROMEDIO', 
            'PRÓX LOTE CADUCA EN', 'FAVORITO', 'A LA VENTA'
        ]
        
        # Agregar encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # Estilo del encabezado: fondo azul oscuro con texto blanco
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Procesar cada producto
        for row_num, producto in enumerate(productos_db, 2):
            try:
                # Calcular costo promedio para este producto
                lotes_activos = LoteInventario.query.filter(
                    LoteInventario.producto_id == producto.id,
                    LoteInventario.esta_activo == True,
                    LoteInventario.stock > 0
                ).all()
                
                costo_promedio = producto.costo or 0
                if lotes_activos:
                    costo_total = sum((lote.costo_unitario or 0) * (lote.stock or 0) for lote in lotes_activos)
                    stock_total = sum(lote.stock or 0 for lote in lotes_activos)
                    if stock_total > 0:
                        costo_promedio = costo_total / stock_total
                
                # Calcular días hasta próxima caducidad
                dias_caducidad = "N/A"
                if lotes_activos:
                    from datetime import date
                    hoy = date.today()
                    lotes_con_caducidad = [l for l in lotes_activos if l.fecha_caducidad]
                    
                    if lotes_con_caducidad:
                        # Ordenar por fecha de caducidad
                        lotes_con_caducidad.sort(key=lambda x: x.fecha_caducidad)
                        
                        # Buscar el primer lote que no haya caducado
                        proximo_lote = None
                        for lote in lotes_con_caducidad:
                            if lote.fecha_caducidad >= hoy:
                                proximo_lote = lote
                                break
                        
                        # Si no hay lote sin caducar, usar el más recientemente caducado
                        if not proximo_lote and lotes_con_caducidad:
                            proximo_lote = lotes_con_caducidad[0]
                        
                        if proximo_lote:
                            if proximo_lote.fecha_caducidad < hoy:
                                dias_diff = (hoy - proximo_lote.fecha_caducidad).days
                                dias_caducidad = f"Caducado hace {dias_diff} días"
                            elif proximo_lote.fecha_caducidad == hoy:
                                dias_caducidad = "Caduca hoy"
                            else:
                                dias_diff = (proximo_lote.fecha_caducidad - hoy).days
                                dias_caducidad = f"{dias_diff} días"
                
                # Preparar datos de la fila - asegurar que todos sean valores básicos
                row_data = [
                    str(producto.nombre or "Sin nombre"),
                    str(producto.codigo_barras_externo or f"SKU-{producto.id:05d}"),
                    str(producto.marca or "No definida"),
                    str(producto.categoria or "Sin categoría"),
                    float(producto.stock or 0),
                    float(producto.precio_final or producto.precio_venta or 0),
                    float(producto.costo or 0),
                    float(costo_promedio),
                    str(dias_caducidad),
                    "Sí" if producto.es_favorito else "No",
                    "Sí" if producto.esta_a_la_venta else "No"
                ]
                
                # Agregar datos a la fila
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col, value=value)
                
            except Exception as e:
                print(f"Error procesando producto {producto.id}: {str(e)}")
                continue
        
        # Ajustar ancho de columnas
        column_widths = [25, 18, 18, 20, 12, 15, 15, 15, 20, 12, 12]
        for col, width in enumerate(column_widths, 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[column_letter].width = width
        
        # Aplicar filtros automáticos solo al rango de datos
        if len(productos_db) > 0:
            ws.auto_filter.ref = f"A1:K{len(productos_db) + 1}"
        
        # Crear buffer en memoria para el archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Preparar respuesta con el archivo
        from datetime import datetime
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"productos_inventario_{fecha_actual}.xlsx"
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        print(f"Error en exportar_productos_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al generar el archivo Excel: {str(e)}', 'danger')
        return redirect(url_for('ver_productos'))

@app.route('/actualizar-lotes-caducidad', methods=['GET'])
@login_requerido
def actualizar_lotes_caducidad():
    """
    Ruta para actualizar la lógica de próximo lote a caducar en productos existentes.
    Esta función recorre todos los productos con lotes y actualiza la información de caducidad.
    """
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    empresa_id = session.get('user_id')
    
    try:
        # Obtener todos los productos de la empresa
        productos = Producto.query.filter_by(empresa_id=empresa_id).all()
        
        productos_actualizados = 0
        lotes_procesados = 0
        
        for producto in productos:
            # Buscar lotes activos con stock positivo
            lotes_activos = LoteInventario.query.filter(
                LoteInventario.producto_id == producto.id,
                LoteInventario.esta_activo == True,
                LoteInventario.stock > 0
            ).all()
            
            if not lotes_activos:
                continue
                
            # Filtrar lotes con fecha de caducidad
            lotes_con_caducidad = [lote for lote in lotes_activos if lote.fecha_caducidad is not None]
            
            if not lotes_con_caducidad:
                continue
                
            # Ordenar por fecha de caducidad (más cercana primero)
            lotes_con_caducidad.sort(key=lambda x: x.fecha_caducidad)
            
            # Buscar el primer lote que no haya caducado aún
            hoy = date.today()
            lote_no_caducado = None
            
            for lote in lotes_con_caducidad:
                if lote.fecha_caducidad >= hoy:
                    lote_no_caducado = lote
                    break
            
            # Si no hay lote sin caducar, usar el que caducó más recientemente
            if not lote_no_caducado and lotes_con_caducidad:
                lote_no_caducado = lotes_con_caducidad[0]
            
            # Si encontramos un lote para mostrar, lo marcamos como "proximo_lote"
            if lote_no_caducado:
                # Opcional: Actualizar algún campo en el producto o lote si es necesario
                lotes_procesados += 1
            
            productos_actualizados += 1
                
        return f"""
        <h2>Actualización de Lotes Completada</h2>
        <p>Se procesaron {productos_actualizados} productos y {lotes_procesados} lotes con fechas de caducidad.</p>
        <p>Los cambios en la lógica de caducidad ahora están activos.</p>
        <p><a href="{url_for('ver_productos')}">Volver a la lista de productos</a></p>
        """
        
    except Exception as e:
        return f"""
        <h2>Error al Actualizar Lotes</h2>
        <p>Se produjo un error: {str(e)}</p>
        <p><a href="{url_for('ver_productos')}">Volver a la lista de productos</a></p>
        """

@app.route('/api/toggle_cost_type', methods=['POST'])
@login_requerido
def toggle_cost_type():
    try:
        data = request.get_json()
        cost_type = data.get('type', 'last')  # 'last' o 'average'
        
        # Guardar la preferencia en la sesión (opcional)
        session['cost_display_type'] = cost_type
        
        return jsonify({
            "success": True,
            "message": f"Tipo de costo cambiado a: {cost_type}",
            "type": cost_type
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error al cambiar tipo de costo: {str(e)}"
        }), 500

# NUEVO: Endpoint para cambiar el estado de favorito
@app.route('/api/toggle_favorite/<int:product_id>', methods=['POST'])
@login_requerido
def toggle_favorite(product_id):
    try:
        # Obtener el producto y verificar que pertenezca a la empresa del usuario
        producto = Producto.query.get_or_404(product_id)
        
        if producto.empresa_id != session.get('user_id'):
            return jsonify({"success": False, "message": "No tienes permiso para modificar este producto"}), 403
        
        # Obtener el nuevo estado desde la solicitud JSON
        data = request.get_json()
        nuevo_estado = data.get('status') == '1'
        
        # Actualizar el estado
        producto.es_favorito = nuevo_estado
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "message": "Estado de favorito actualizado correctamente", 
            "estado": "1" if nuevo_estado else "0"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"Error al actualizar: {str(e)}"}), 500

# NUEVO: Endpoint para cambiar el estado de visibilidad
@app.route('/api/toggle_visibility/<int:product_id>', methods=['POST'])
@login_requerido
def toggle_visibility(product_id):
    try:
        # Obtener el producto y verificar que pertenezca a la empresa del usuario
        producto = Producto.query.get_or_404(product_id)
        
        if producto.empresa_id != session.get('user_id'):
            return jsonify({"success": False, "message": "No tienes permiso para modificar este producto"}), 403
        
        # Obtener el nuevo estado desde la solicitud JSON
        data = request.get_json()
        nuevo_estado = data.get('status') == '1'
        
        # Actualizar el estado
        producto.esta_a_la_venta = nuevo_estado
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "message": "Visibilidad actualizada correctamente", 
            "estado": "1" if nuevo_estado else "0"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"Error al actualizar: {str(e)}"}), 500

# NUEVO: Endpoint para actualizar el precio
@app.route('/api/update_price/<int:product_id>', methods=['POST'])
@login_requerido
def update_price(product_id):
    try:
        producto = Producto.query.get_or_404(product_id)
        
        if producto.empresa_id != session.get('user_id'):
            return jsonify({"success": False, "message": "No tienes permiso para modificar este producto"}), 403
        
        data = request.get_json()
        try:
            nuevo_precio = float(data.get('price', 0))
            if nuevo_precio < 0:
                return jsonify({"success": False, "message": "El precio no puede ser negativo"}), 400
        except ValueError:
            return jsonify({"success": False, "message": "El precio debe ser un número válido"}), 400
        
        precio_anterior = producto.precio_venta
        producto.precio_venta = nuevo_precio
        
        # CALCULAR PRECIO FINAL DINÁMICAMENTE
        precio_final = nuevo_precio
        tiene_descuento = False
        tipo_descuento = None
        valor_descuento = 0.0
        
        # Verificar descuento activo
        if (hasattr(producto, 'tiene_descuento') and producto.tiene_descuento and 
            hasattr(producto, 'valor_descuento') and producto.valor_descuento > 0):
            
            tiene_descuento = True
            tipo_descuento = getattr(producto, 'tipo_descuento', None)
            valor_descuento = getattr(producto, 'valor_descuento', 0.0)
            
            if tipo_descuento == 'percentage':
                precio_final = nuevo_precio * (1 - valor_descuento / 100)
            elif tipo_descuento == 'fixed':
                precio_final = max(0, nuevo_precio - valor_descuento)
        
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "message": "Precio actualizado correctamente", 
            "precio": nuevo_precio,
            "precio_base": nuevo_precio,
            "precio_final": precio_final,
            "precio_anterior": precio_anterior,
            "tiene_descuento": tiene_descuento,
            "tipo_descuento": tipo_descuento,
            "valor_descuento": valor_descuento
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False, 
            "message": f"Error al actualizar: {str(e)}",
            "precio": 0,
            "precio_base": 0,
            "precio_final": 0,
            "precio_anterior": 0,
            "tiene_descuento": False,
            "tipo_descuento": None,
            "valor_descuento": 0
        }), 500

@app.route('/api/actualizar-ubicacion/<int:product_id>', methods=['POST'])
@login_requerido
def actualizar_ubicacion(product_id):
    """
    API para actualizar la ubicación de un producto específico.
    """
    try:
        # Obtener el producto y verificar que pertenezca a la empresa del usuario
        producto = Producto.query.get_or_404(product_id)
        
        if producto.empresa_id != session.get('user_id'):
            return jsonify({"success": False, "message": "No tienes permiso para modificar este producto"}), 403
        
        # Obtener la nueva ubicación desde el request
        data = request.get_json()
        nueva_ubicacion = data.get('ubicacion', '').strip()
        
        print(f"DEBUG: Actualizando ubicación de producto {product_id} a '{nueva_ubicacion}'")
        print(f"DEBUG: Estado previo: tipo={producto.ubicacion_tipo}, grupo={producto.ubicacion_grupo}")
        
        # Limpiar ubicación anterior
        producto.ubicacion_tipo = None
        producto.ubicacion_grupo = None
        
        # Actualizar la ubicación
        producto.ubicacion = nueva_ubicacion
        
        # Si la ubicación es vacía, quitar el tipo y grupo
        if not nueva_ubicacion:
            producto.ubicacion_tipo = None
            producto.ubicacion_grupo = None
            print(f"DEBUG: Removiendo ubicación del producto {product_id}")
        else:
            # Por defecto, si se actualiza directamente un producto, es individual
            producto.ubicacion_tipo = 'individual'
            producto.ubicacion_grupo = None
            print(f"DEBUG: Estableciendo producto {product_id} como ubicación individual")
        
        db.session.commit()
        print(f"DEBUG: Ubicación actualizada exitosamente para producto {product_id}")
        
        return jsonify({
            "success": True, 
            "message": "Ubicación actualizada correctamente",
            "ubicacion": nueva_ubicacion,
            "ubicacion_tipo": producto.ubicacion_tipo,
            "ubicacion_grupo": producto.ubicacion_grupo
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"ERROR al actualizar ubicación de producto {product_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Error al actualizar: {str(e)}"}), 500

@app.route('/api/get_active_locations', methods=['GET'])
@login_requerido
def get_active_locations():
    """Obtiene todos los tipos de ubicaciones activas para facilitar visualización"""
    try:
        empresa_id = session.get('user_id')
        if not empresa_id:
            return jsonify({"success": False, "message": "Usuario no autenticado"}), 401
        
        # Estructura para almacenar los resultados
        active_locations = {
            "global": False,
            "categorias": [],
            "marcas": [],
            "individuales": 0,
            "location_values": {},
            "conteos": {}  # NUEVO: Diccionario para almacenar conteos
        }
        
        # Verificar si hay productos con ubicación
        hay_productos = Producto.query.filter(
            Producto.empresa_id == empresa_id,
            Producto.is_approved == True,
            Producto.ubicacion.isnot(None),
            Producto.ubicacion != ''
        ).first() is not None
        
        if not hay_productos:
            return jsonify({"success": True, "active_locations": active_locations})
        
        # PASO 1: Buscar ubicación global (tipo = 'global')
        global_products = Producto.query.filter(
            Producto.empresa_id == empresa_id,
            Producto.is_approved == True,
            Producto.ubicacion_tipo == 'global',
            Producto.ubicacion.isnot(None),
            Producto.ubicacion != ''
        ).first()
        
        if global_products:
            active_locations["global"] = True
            active_locations["location_values"]["global"] = global_products.ubicacion
            
            # Contar productos con ubicación global para el frontend
            count_global = Producto.query.filter(
                Producto.empresa_id == empresa_id,
                Producto.is_approved == True,
                Producto.ubicacion_tipo == 'global'
            ).count()
            
            # NUEVO: Guardar el conteo en el diccionario
            active_locations["conteos"]["global"] = count_global
            
            print(f"DEBUG: Encontrada ubicación global con {count_global} productos")
            
            # ELIMINADO: Ya no retornamos temprano, continuamos procesando otras ubicaciones
        
        # PASO 2: Buscar ubicaciones por categoría
        categorias_con_ubicacion = db.session.query(
            Producto.ubicacion_grupo, Producto.ubicacion
        ).filter(
            Producto.empresa_id == empresa_id,
            Producto.is_approved == True,
            Producto.ubicacion_tipo == 'categoria',
            Producto.ubicacion_grupo.isnot(None),
            Producto.ubicacion.isnot(None),
            Producto.ubicacion != ''
        ).distinct().all()
        
        # NUEVO: Inicializar conteos de categorías
        active_locations["conteos"]["categorias"] = {}
        
        # Procesar categorías
        for categoria_data in categorias_con_ubicacion:
            categoria = categoria_data[0]
            ubicacion = categoria_data[1]
            
            if categoria not in active_locations["categorias"]:
                active_locations["categorias"].append(categoria)
                active_locations["location_values"][categoria] = ubicacion
                
                # NUEVO: Contar productos de esta categoría
                count_categoria = Producto.query.filter(
                    Producto.empresa_id == empresa_id,
                    Producto.is_approved == True,
                    Producto.ubicacion_tipo == 'categoria',
                    Producto.ubicacion_grupo == categoria
                ).count()
                
                # NUEVO: Guardar conteo por categoría
                active_locations["conteos"]["categorias"][categoria] = count_categoria
        
        # Depuración
        print(f"DEBUG: Encontradas {len(active_locations['categorias'])} categorías con ubicación")
        
        # PASO 3: Buscar ubicaciones por marca
        marcas_con_ubicacion = db.session.query(
            Producto.ubicacion_grupo, Producto.ubicacion
        ).filter(
            Producto.empresa_id == empresa_id,
            Producto.is_approved == True,
            Producto.ubicacion_tipo == 'marca',
            Producto.ubicacion_grupo.isnot(None),
            Producto.ubicacion.isnot(None),
            Producto.ubicacion != ''
        ).distinct().all()
        
        # NUEVO: Inicializar conteos de marcas
        active_locations["conteos"]["marcas"] = {}
        
        # Procesar marcas
        for marca_data in marcas_con_ubicacion:
            marca = marca_data[0]
            ubicacion = marca_data[1]
            
            if marca not in active_locations["marcas"]:
                active_locations["marcas"].append(marca)
                active_locations["location_values"][marca] = ubicacion
                
                # NUEVO: Contar productos de esta marca
                count_marca = Producto.query.filter(
                    Producto.empresa_id == empresa_id,
                    Producto.is_approved == True,
                    Producto.ubicacion_tipo == 'marca',
                    Producto.ubicacion_grupo == marca
                ).count()
                
                # NUEVO: Guardar conteo por marca
                active_locations["conteos"]["marcas"][marca] = count_marca
        
        # Depuración
        print(f"DEBUG: Encontradas {len(active_locations['marcas'])} marcas con ubicación")
        
        # PASO 4: Contar productos individuales - SOLO contar los que realmente
        # tienen ubicacion_tipo='individual'
        productos_individuales = Producto.query.filter(
            Producto.empresa_id == empresa_id,
            Producto.is_approved == True,
            Producto.ubicacion_tipo == 'individual',
            Producto.ubicacion.isnot(None),
            Producto.ubicacion != ''
        ).count()
        
        active_locations["individuales"] = productos_individuales
        active_locations["conteos"]["individuales"] = productos_individuales
        
        # Depuración
        print(f"DEBUG: Encontrados {productos_individuales} productos con ubicación individual")
        
        return jsonify({
            "success": True,
            "active_locations": active_locations
        })
        
    except Exception as e:
        print(f"Error al obtener ubicaciones activas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/ubicacion/migrar-datos-existentes', methods=['GET'])
@login_requerido
def migrar_datos_ubicacion():
    """
    Ruta para migrar datos existentes de ubicaciones al nuevo esquema con tipo y grupo.
    Esta ruta debe ejecutarse una sola vez después de aplicar la migración de base de datos.
    """
    try:
        empresa_id = session.get('user_id')
        if not empresa_id:
            return "Usuario no autenticado", 401
        
        # 1. Obtener todos los productos con ubicación
        productos_con_ubicacion = Producto.query.filter(
            Producto.empresa_id == empresa_id,
            Producto.is_approved == True,
            Producto.ubicacion.isnot(None),
            Producto.ubicacion != ''
        ).all()
        
        if not productos_con_ubicacion:
            return "No hay productos con ubicación para migrar."
        
        # Contadores para el informe
        migrados_global = 0
        migrados_categoria = 0
        migrados_marca = 0
        migrados_individual = 0
        
        # Estructura para análisis
        total_productos = len(productos_con_ubicacion)
        ubicaciones_conteo = {}
        
        for producto in productos_con_ubicacion:
            if producto.ubicacion not in ubicaciones_conteo:
                ubicaciones_conteo[producto.ubicacion] = 0
            ubicaciones_conteo[producto.ubicacion] += 1
        
        # 2. Verificar ubicación global
        global_ubicacion = None
        for ubicacion, count in ubicaciones_conteo.items():
            if count / total_productos >= 0.8:
                global_ubicacion = ubicacion
                break
        
        # Si hay ubicación global, migrar todos los productos con esa ubicación
        if global_ubicacion:
            for producto in productos_con_ubicacion:
                if producto.ubicacion == global_ubicacion:
                    producto.ubicacion_tipo = 'global'
                    producto.ubicacion_grupo = None
                    migrados_global += 1
            
            # Guardar cambios y retornar resultado
            db.session.commit()
            return f"Migrados {migrados_global} productos a ubicación global."
        
        # Si no hay global, continuar con categorías y marcas
        # Crear estructuras para análisis
        productos_por_categoria = {}
        productos_por_marca = {}
        productos_asignados = set()
        
        # Agrupar por categoría y marca
        for producto in productos_con_ubicacion:
            # Por categoría
            if producto.categoria:
                if producto.categoria not in productos_por_categoria:
                    productos_por_categoria[producto.categoria] = []
                productos_por_categoria[producto.categoria].append(producto)
            
            # Por marca
            if producto.marca:
                if producto.marca not in productos_por_marca:
                    productos_por_marca[producto.marca] = []
                productos_por_marca[producto.marca].append(producto)
        
        # 3. Procesar categorías
        for categoria, productos in productos_por_categoria.items():
            # Contar ubicaciones por categoría
            ubicaciones = {}
            for producto in productos:
                if producto.ubicacion not in ubicaciones:
                    ubicaciones[producto.ubicacion] = 0
                ubicaciones[producto.ubicacion] += 1
            
            # Verificar ubicación dominante
            ubicacion_dominante = None
            total_productos_categoria = len(productos)
            
            for ubicacion, count in ubicaciones.items():
                if count / total_productos_categoria >= 0.8:
                    ubicacion_dominante = ubicacion
                    break
            
            # Si hay ubicación dominante, migrar
            if ubicacion_dominante:
                for producto in productos:
                    if producto.ubicacion == ubicacion_dominante:
                        producto.ubicacion_tipo = 'categoria'
                        producto.ubicacion_grupo = categoria
                        migrados_categoria += 1
                        productos_asignados.add(producto.id)
        
        # 4. Procesar marcas
        for marca, productos in productos_por_marca.items():
            # Filtrar productos ya asignados
            productos_no_asignados = [p for p in productos if p.id not in productos_asignados]
            
            if not productos_no_asignados:
                continue
            
            # Contar ubicaciones por marca
            ubicaciones = {}
            for producto in productos_no_asignados:
                if producto.ubicacion not in ubicaciones:
                    ubicaciones[producto.ubicacion] = 0
                ubicaciones[producto.ubicacion] += 1
            
            # Verificar ubicación dominante
            ubicacion_dominante = None
            total_productos_marca = len(productos_no_asignados)
            
            for ubicacion, count in ubicaciones.items():
                if count / total_productos_marca >= 0.8:
                    ubicacion_dominante = ubicacion
                    break
            
            # Si hay ubicación dominante, migrar
            if ubicacion_dominante:
                for producto in productos_no_asignados:
                    if producto.ubicacion == ubicacion_dominante:
                        producto.ubicacion_tipo = 'marca'
                        producto.ubicacion_grupo = marca
                        migrados_marca += 1
                        productos_asignados.add(producto.id)
        
        # 5. Productos individuales (los que quedan)
        for producto in productos_con_ubicacion:
            if producto.id not in productos_asignados:
                producto.ubicacion_tipo = 'individual'
                producto.ubicacion_grupo = None
                migrados_individual += 1
        
        # Guardar cambios
        db.session.commit()
        
        # Construir mensaje de resultado
        mensaje = f"""
        <h2>Migración de ubicaciones completada</h2>
        <ul>
            <li>Productos en ubicación global: {migrados_global}</li>
            <li>Productos en ubicación por categoría: {migrados_categoria}</li>
            <li>Productos en ubicación por marca: {migrados_marca}</li>
            <li>Productos en ubicación individual: {migrados_individual}</li>
            <li>Total de productos migrados: {migrados_global + migrados_categoria + migrados_marca + migrados_individual}</li>
        </ul>
        <p><a href="/ubicacion-productos">Volver a Ubicaciones</a></p>
        """
        
        return mensaje
        
    except Exception as e:
        db.session.rollback()
        return f"""
        <h2>Error en la migración de datos</h2>
        <p>Se produjo un error: {str(e)}</p>
        <p><a href="/ubicacion-productos">Volver a Ubicaciones</a></p>
        """

@app.route('/api/actualizar-ubicacion-masiva', methods=['POST'])
@login_requerido
def actualizar_ubicacion_masiva():
    """
    API para actualizar la ubicación de múltiples productos a la vez.
    Puede ser por categoría, marca, o lista de IDs.
    """
    try:
        # Obtener datos del request
        data = request.get_json()
        print(f"DEBUG: Datos recibidos en actualizar-ubicacion-masiva: {data}")
        
        ubicacion = data.get('ubicacion', '').strip()
        producto_ids = data.get('producto_ids', [])
        filtro_categoria = data.get('categoria')
        filtro_marca = data.get('marca')
        es_global = data.get('global', False)
        es_remover = data.get('remove', False)
        
        empresa_id = session.get('user_id')
        
        # Validar datos
        if not ubicacion and not es_remover:
            return jsonify({"success": False, "message": "La ubicación no puede estar vacía al asignar"}), 400
        
        productos_actualizados = 0
        
        # Si es una operación de remover, configuramos los valores
        if es_remover:
            ubicacion = ''
            ubicacion_tipo = None
            ubicacion_grupo = None
        
        # Caso 1: Actualizar todos (global)
        if es_global:
            # PRIMERO: Limpiar todas las ubicaciones existentes para evitar duplicación
            if not es_remover:
                Producto.query.filter_by(
                    empresa_id=empresa_id,
                    is_approved=True
                ).update({
                    "ubicacion_tipo": None,
                    "ubicacion_grupo": None
                }, synchronize_session="fetch")
                
                print(f"DEBUG: Se limpiaron ubicaciones previas antes de aplicar global")
            
            # Actualizar todos los productos con la nueva ubicación global
            data_update = {
                "ubicacion": ubicacion
            }
            
            # Si no es remover, establecer tipo de ubicación
            if not es_remover:
                data_update["ubicacion_tipo"] = "global"
                data_update["ubicacion_grupo"] = None
            else:
                data_update["ubicacion_tipo"] = None
                data_update["ubicacion_grupo"] = None
            
            count = Producto.query.filter_by(
                empresa_id=empresa_id,
                is_approved=True
            ).update(data_update, synchronize_session="fetch")
            
            productos_actualizados = count
            print(f"DEBUG: Actualización global completada. Productos actualizados: {count}")
        
        # Caso 2: Actualizar por IDs específicos (individual)
        elif producto_ids and len(producto_ids) > 0:
            # PRIMERO: Limpiar ubicaciones previas de estos productos
            if not es_remover:
                Producto.query.filter(
                    Producto.id.in_(producto_ids),
                    Producto.empresa_id == empresa_id
                ).update({
                    "ubicacion_tipo": None,
                    "ubicacion_grupo": None
                }, synchronize_session="fetch")
                
                print(f"DEBUG: Se limpiaron ubicaciones previas para {len(producto_ids)} productos individuales")
            
            # Filtrar para asegurar que solo se actualicen productos de la empresa actual
            data_update = {
                "ubicacion": ubicacion
            }
            
            # Si no es remover, establecer tipo individual
            if not es_remover:
                data_update["ubicacion_tipo"] = "individual"
                data_update["ubicacion_grupo"] = None
            else:
                data_update["ubicacion_tipo"] = None
                data_update["ubicacion_grupo"] = None
            
            count = Producto.query.filter(
                Producto.id.in_(producto_ids),
                Producto.empresa_id == empresa_id
            ).update(data_update, synchronize_session="fetch")
            
            productos_actualizados = count
            print(f"DEBUG: Actualización por IDs completada. Productos actualizados: {count}")
        
        # Caso 3: Actualizar por categoría
        elif filtro_categoria:
            # PRIMERO: Limpiar ubicaciones previas de estos productos
            if not es_remover:
                Producto.query.filter_by(
                    empresa_id=empresa_id,
                    categoria=filtro_categoria
                ).update({
                    "ubicacion_tipo": None,
                    "ubicacion_grupo": None
                }, synchronize_session="fetch")
                
                print(f"DEBUG: Se limpiaron ubicaciones previas para productos de categoría {filtro_categoria}")
            
            data_update = {
                "ubicacion": ubicacion
            }
            
            # Si no es remover, establecer tipo categoría
            if not es_remover:
                data_update["ubicacion_tipo"] = "categoria"
                data_update["ubicacion_grupo"] = filtro_categoria
            else:
                data_update["ubicacion_tipo"] = None
                data_update["ubicacion_grupo"] = None
            
            count = Producto.query.filter_by(
                empresa_id=empresa_id,
                categoria=filtro_categoria
            ).update(data_update, synchronize_session="fetch")
            
            productos_actualizados = count
            print(f"DEBUG: Actualización por categoría completada. Productos actualizados: {count}")
        
        # Caso 4: Actualizar por marca
        elif filtro_marca:
            # PRIMERO: Limpiar ubicaciones previas de estos productos
            if not es_remover:
                Producto.query.filter_by(
                    empresa_id=empresa_id,
                    marca=filtro_marca
                ).update({
                    "ubicacion_tipo": None,
                    "ubicacion_grupo": None
                }, synchronize_session="fetch")
                
                print(f"DEBUG: Se limpiaron ubicaciones previas para productos de marca {filtro_marca}")
            
            data_update = {
                "ubicacion": ubicacion
            }
            
            # Si no es remover, establecer tipo marca
            if not es_remover:
                data_update["ubicacion_tipo"] = "marca"
                data_update["ubicacion_grupo"] = filtro_marca
            else:
                data_update["ubicacion_tipo"] = None
                data_update["ubicacion_grupo"] = None
            
            count = Producto.query.filter_by(
                empresa_id=empresa_id,
                marca=filtro_marca
            ).update(data_update, synchronize_session="fetch")
            
            productos_actualizados = count
            print(f"DEBUG: Actualización por marca completada. Productos actualizados: {count}")
        
        # Guardar cambios
        if productos_actualizados > 0:
            db.session.commit()
            
            return jsonify({
                "success": True,
                "message": f"Se actualizaron {productos_actualizados} productos",
                "count": productos_actualizados
            })
        else:
            return jsonify({
                "success": False,
                "message": "No se encontraron productos para actualizar"
            }), 404
            
    except Exception as e:
        db.session.rollback()
        print(f"ERROR en actualizar-ubicacion-masiva: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500

@app.route('/api/lotes/<int:producto_id>')
@login_requerido
def api_lotes(producto_id):
    """
    API para obtener los lotes de un producto.
    """
    try:
        # Verificar que el producto pertenece a la empresa
        empresa_id = session.get('user_id')
        producto = Producto.query.filter_by(
            id=producto_id,
            empresa_id=empresa_id
        ).first()
        
        if not producto:
            return jsonify({
                "success": False,
                "message": "Producto no encontrado o sin permisos"
            }), 404
        
        # Obtener todos los lotes del producto
        lotes = LoteInventario.query.filter_by(
            producto_id=producto_id
        ).order_by(LoteInventario.fecha_entrada.desc()).all()
        
        # Convertir lotes a diccionarios
        lotes_dict = []
        for lote in lotes:
            lotes_dict.append({
                "id": lote.id,
                "numero_lote": lote.numero_lote,
                "costo_unitario": float(lote.costo_unitario),
                "stock": float(lote.stock),
                "fecha_entrada": lote.fecha_entrada.isoformat(),
                "fecha_caducidad": lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None,
                "esta_activo": lote.esta_activo
            })
        
        return jsonify({
            "success": True,
            "producto_nombre": producto.nombre,
            "lotes": lotes_dict
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@app.route('/api/actualizar-costo', methods=['POST'])
@login_requerido
def api_actualizar_costo():
    """
    API para actualizar el costo del último lote activo, manteniendo trazabilidad.
    """
    # Importar módulos necesarios al inicio de la función
    from datetime import datetime
    
    try:
        # Obtener datos del request
        data = request.get_json()
        producto_id = data.get('producto_id')
        costo = float(data.get('costo', 0))
        afectar_precio = data.get('afectar_precio', False)
        
        # Validaciones básicas
        if not producto_id or costo < 0:
            return jsonify({
                "success": False,
                "message": "Datos inválidos"
            }), 400
        
        # Verificar que el producto pertenece a la empresa
        empresa_id = session.get('user_id')
        producto = Producto.query.filter_by(
            id=producto_id,
            empresa_id=empresa_id
        ).first()
        
        if not producto:
            return jsonify({
                "success": False,
                "message": "Producto no encontrado o sin permisos"
            }), 404
        
        # Buscar el último lote activo con stock > 0
        ultimo_lote = LoteInventario.query.filter(
            LoteInventario.producto_id == producto_id,
            LoteInventario.esta_activo == True,
            LoteInventario.stock > 0
        ).order_by(LoteInventario.fecha_entrada.desc()).first()
        
        if not ultimo_lote:
            return jsonify({
                "success": False,
                "message": "No se encontró un lote activo para actualizar"
            }), 400
        
        # Guardar valores antiguos para registro
        costo_anterior_lote = ultimo_lote.costo_unitario
        costo_anterior_producto = producto.costo
        
        # Actualizar el costo del lote 
        ultimo_lote.costo_unitario = costo
        
        # Actualizar el costo base del producto
        producto.costo = costo
        
        # Calcular proporciones antes de la actualización si se solicita
        if afectar_precio and costo_anterior_producto > 0:
            proporcion = producto.precio_venta / costo_anterior_producto
            # Actualizar el precio proporcionalmente
            nuevo_precio = costo * proporcion
            producto.precio_venta = nuevo_precio
        
        # Registrar el cambio en el historial (Movimiento de Inventario) como CORRECCIÓN
        from models.modelos_inventario import MovimientoInventario
        nuevo_movimiento = MovimientoInventario(
            producto_id=producto_id,
            tipo_movimiento='ENTRADA',
            cantidad=0,  # No cambiamos stock, solo el costo
            motivo='corrección de costo',
            fecha_movimiento=datetime.now(),
            costo_unitario=costo,
            numero_lote=ultimo_lote.numero_lote,
            notas=f"Corrección de costo en lote '{ultimo_lote.numero_lote}': de ${costo_anterior_lote} a ${costo}",
            usuario_id=empresa_id
        )
        db.session.add(nuevo_movimiento)
        
        # Guardar cambios
        db.session.commit()
        
        # Recalcular costo promedio
        lotes_activos = LoteInventario.query.filter(
            LoteInventario.producto_id == producto_id,
            LoteInventario.esta_activo == True,
            LoteInventario.stock > 0
        ).all()
        
        costo_promedio = costo  # Valor por defecto
        
        if lotes_activos:
            costo_total = 0
            stock_total = 0
            
            for lote in lotes_activos:
                costo_total += lote.costo_unitario * lote.stock
                stock_total += lote.stock
            
            if stock_total > 0:
                costo_promedio = costo_total / stock_total
        
        return jsonify({
            "success": True,
            "message": f"Costo corregido en lote {ultimo_lote.numero_lote}",
            "costo": costo,
            "costo_promedio": costo_promedio,
            "precio_actualizado": afectar_precio,
            "lote_actualizado": ultimo_lote.numero_lote,
            "movimiento_id": nuevo_movimiento.id,
            "costo_anterior": costo_anterior_lote
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@app.route('/api/reset-product-locations', methods=['POST'])
@login_requerido
def reset_product_locations():
    """
    API para resetear todas las ubicaciones de productos para la empresa actual.
    Útil cuando hay inconsistencias y se quiere empezar de cero.
    """
    try:
        empresa_id = session.get('user_id')
        if not empresa_id:
            return jsonify({"success": False, "message": "Usuario no autenticado"}), 401
        
        # Actualizar todos los productos para quitar ubicaciones
        data_update = {
            "ubicacion": '',
            "ubicacion_tipo": None,
            "ubicacion_grupo": None
        }
        
        count = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True
        ).update(data_update, synchronize_session="fetch")
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"Se han reseteado las ubicaciones de {count} productos",
            "count": count
        })
    
    except Exception as e:
        db.session.rollback()
        print(f"ERROR al resetear ubicaciones: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500

@app.route('/agregar-producto', methods=['GET','POST'])
@login_requerido
def agregar_producto():
    if request.method == 'POST':
        try:
            # Recoger campos del formulario
            codigo_barras_externo = request.form.get("codigo_barras_externo", "").strip()
            nombre = request.form.get("nombre", "").strip()
            stock_str = request.form.get("stock", "0").strip() 
            costo_str = request.form.get("costo", "$0").strip()
            precio_str = request.form.get("precio_venta", "$0").strip()
            marca = request.form.get("marca", "").strip()
            
            # NUEVO: Verificar duplicados si hay código de barras
            if codigo_barras_externo:
                # Buscar productos con el mismo código (solo en los productos de esta empresa)
                producto_existente = Producto.query.filter_by(
                    empresa_id=session['user_id'],
                    codigo_barras_externo=codigo_barras_externo
                ).first()
                
                if producto_existente:
                    # Si existe, devolver mensaje de error
                    flash(f'Error: Ya existe un producto con el código de barras {codigo_barras_externo}', 'danger')
                    return redirect(url_for('agregar_producto'))
            
            # Categoría
            cat_option = request.form.get('categoria_option', 'existente')
            if cat_option == 'existente':
                categoria = request.form.get('categoria_existente', '').strip()
            else:
                categoria = request.form.get('categoria_nueva', '').strip()
            
            # MODIFICADO: Usar la nueva función de normalización
            categoria_normalizada = normalize_category(categoria)
            
            # Favorito y a la venta
            es_favorito_bool = (request.form.get("es_favorito", "0") == "1")
            esta_a_la_venta_bool = (request.form.get("esta_a_la_venta", "1") == "1")
            
            # Caducidad
            has_caducidad = (request.form.get("toggle_caducidad_estado", "DESACTIVADO") == "ACTIVADO")
            caducidad_lapso = request.form.get("caducidad_lapso", None) if has_caducidad else None
            
            # Parse numéricos
            try:
                stock_int = int(stock_str)
            except:
                stock_int = 0
            
            costo_val = parse_money(costo_str)
            precio_val = parse_money(precio_str)
            
            # ===== MANEJO SIMPLIFICADO DE IMÁGENES =====
            foto_final = None
            
            # 1. Primero intentar usar el archivo subido
            if request.files.get('foto') and request.files['foto'].filename:
                file = request.files['foto']
                if allowed_file(file.filename):
                    # Generar nombre único
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    
                    # Guardar archivo
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    
                    file.save(os.path.join(UPLOAD_FOLDER, filename))
                    foto_final = filename
            
            # 2. Si no hay archivo, usar displayed_image_url o ia_foto_filename
            if not foto_final:
                displayed_url = request.form.get("displayed_image_url", "").strip()
                ia_filename = request.form.get("ia_foto_filename", "").strip()
                
                if ia_filename and os.path.exists(os.path.join(UPLOAD_FOLDER, ia_filename)):
                    # Usar el archivo IA que ya existe
                    foto_final = ia_filename
                elif displayed_url:
                    # Es una URL - intentar descargar o usar nombre de archivo
                    if "/uploads/" in displayed_url:
                        # Es local, extraer nombre
                        foto_final = displayed_url.split("/uploads/")[-1]
                    elif displayed_url.startswith(("http://", "https://")):
                        # Es externa, descargar
                        try:
                            ext = "jpg"  # Default
                            filename = f"{uuid.uuid4().hex}.{ext}"
                            filepath = os.path.join(UPLOAD_FOLDER, filename)
                            
                            # Descarga síncrona para asegurar que se complete
                            response = requests.get(displayed_url, timeout=10)
                            if response.status_code == 200:
                                if not os.path.exists(os.path.dirname(filepath)):
                                    os.makedirs(os.path.dirname(filepath))
                                
                                with open(filepath, "wb") as f:
                                    f.write(response.content)
                                
                                foto_final = filename
                        except Exception as e:
                            print(f"Error descargando imagen: {e}")
            
            # 3. Si todo falla, usar imagen predeterminada
            if not foto_final:
                foto_final = "default_product.jpg"
            
            # Usar la nueva función para truncar la URL para asegurar que quepa en la columna
            url_imagen_truncada = truncar_url(request.form.get("displayed_image_url", "").strip(), 95)
                
            # Crear el producto
            nuevo = Producto(
                nombre=nombre,
                stock=stock_int,
                costo=costo_val, 
                precio_venta=precio_val,
                categoria=categoria_normalizada,
                # MODIFICADO: Usar la nueva función de color de categoría
                categoria_color=get_category_color(categoria_normalizada),
                foto=foto_final,
                url_imagen=url_imagen_truncada,  # Usando la URL truncada
                is_approved=True,
                empresa_id=session['user_id'],
                codigo_barras_externo=codigo_barras_externo,
                marca=marca,
                es_favorito=es_favorito_bool,
                esta_a_la_venta=esta_a_la_venta_bool,
                has_caducidad=has_caducidad,
                metodo_caducidad=caducidad_lapso
            )
            
            # Guardar en la base de datos
            db.session.add(nuevo)
            db.session.commit()
            
            # ===== MODIFICADO: Crear siempre un lote y movimiento inicial si hay stock positivo =====
            if nuevo.stock > 0:
                try:
                    # Determinar la fecha de caducidad según el método
                    fecha_caducidad = None
                    if nuevo.has_caducidad and nuevo.metodo_caducidad:
                        fecha_actual = datetime.now()

                        if nuevo.metodo_caducidad == '1 día':
                            fecha_caducidad = fecha_actual + timedelta(days=1)
                        elif nuevo.metodo_caducidad == '3 días':
                            fecha_caducidad = fecha_actual + timedelta(days=3)
                        elif nuevo.metodo_caducidad == '1 semana':
                            fecha_caducidad = fecha_actual + timedelta(days=7)
                        elif nuevo.metodo_caducidad == '2 semanas':
                            fecha_caducidad = fecha_actual + timedelta(days=14)
                        elif nuevo.metodo_caducidad == '1 mes':
                            fecha_caducidad = fecha_actual + timedelta(days=30)
                        elif nuevo.metodo_caducidad == '3 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=90)
                        elif nuevo.metodo_caducidad == '6 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=180)
                        elif nuevo.metodo_caducidad == '1 año':
                            fecha_caducidad = fecha_actual + timedelta(days=365)
                        elif nuevo.metodo_caducidad == '2 años':
                            fecha_caducidad = fecha_actual + timedelta(days=730)
                        elif nuevo.metodo_caducidad == '3 años' or nuevo.metodo_caducidad == 'más de 3 años':
                            fecha_caducidad = fecha_actual + timedelta(days=1460)

                        # Convertir a date
                        if fecha_caducidad:
                            fecha_caducidad = fecha_caducidad.date()
                    
                    # Crear el lote inicial y el movimiento inicial
                    # Usamos la función crear_lote_registro existente
                    movimiento, lote = crear_lote_registro(
                        producto=nuevo,
                        cantidad=nuevo.stock,
                        costo=nuevo.costo,
                        fecha_caducidad=fecha_caducidad,
                        usuario_id=session['user_id']
                    )
                    
                    db.session.commit()
                    print(f"Lote inicial y movimiento creados para producto {nuevo.id}")

                except Exception as e:
                    db.session.rollback()
                    print(f"Error al crear lote inicial y movimiento: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    # Continuamos aunque haya error en la creación del lote
            
            flash('Producto guardado exitosamente', 'success')
            
            # MODIFICADO: Redireccionar a la página de confirmación
            response = make_response(redirect(url_for('producto_confirmacion', producto_id=nuevo.id)))
            response.set_cookie('pagina_origen', 'agregar_producto', max_age=300)  # Cookie válida por 5 minutos
            return response
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar el producto: {str(e)}', 'danger')
            print(f"ERROR al guardar producto: {str(e)}")
        
        return redirect(url_for('ver_productos'))
        
    else:
        # GET - Cargar página de agregar producto
        categorias_db = (
            db.session.query(Producto.categoria, Producto.categoria_color)
            .filter(Producto.categoria.isnot(None))
            .filter(Producto.categoria != '')
            .distinct()
            .limit(50)  # Limitar cantidad para que sea más rápido
            .all()
        )
        categories_list = [
            {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
            for c in categorias_db
        ]
        return render_template('agregar_producto.html', categories=categories_list)

@app.route('/agregar-sin-codigo', methods=['GET','POST'])
@login_requerido
def agregar_sin_codigo():
    if request.method == 'POST':
        try:
            # Recoger campos del formulario
            nombre = request.form.get("nombre", "").strip()
            stock_str = request.form.get("stock", "0").strip() 
            costo_str = request.form.get("costo", "$0").strip()
            precio_str = request.form.get("precio_venta", "$0").strip()
            marca = request.form.get("marca", "").strip()
            
            # Obtener el código de barras externo del formulario
            codigo_barras_externo = request.form.get("codigo_barras_externo", "").strip()
            
            # Verificar si el código está presente, si no, generar uno nuevo
            if not codigo_barras_externo:
                codigo_barras_externo = generar_codigo_unico()
                
            print(f"DEBUG: Usando código de barras: {codigo_barras_externo}")
            
            # Categoría
            cat_option = request.form.get('categoria_option', 'existente')
            if cat_option == 'existente':
                categoria = request.form.get('categoria_existente', '').strip()
            else:
                categoria = request.form.get('categoria_nueva', '').strip()
                
            # MODIFICADO: Usar la nueva función de normalización
            categoria_normalizada = normalize_category(categoria)
            
            # Favorito y a la venta
            es_favorito_bool = (request.form.get("es_favorito", "0") == "1")
            esta_a_la_venta_bool = (request.form.get("esta_a_la_venta", "1") == "1")
            
            # Caducidad
            has_caducidad = (request.form.get("toggle_caducidad_estado", "DESACTIVADO") == "ACTIVADO")
            caducidad_lapso = request.form.get("caducidad_lapso", None) if has_caducidad else None
            
            # Parse numéricos
            try:
                stock_int = int(stock_str)
            except:
                stock_int = 0
            
            costo_val = parse_money(costo_str)
            precio_val = parse_money(precio_str)
            
            # Manejo de imagen
            foto_final = process_image(request, UPLOAD_FOLDER, BASE_DIR)
            if not foto_final:
                foto_final = "default_product.jpg"
            
            # Usar la función para truncar la URL para asegurar que quepa en la columna
            url_imagen_truncada = truncar_url(request.form.get("displayed_image_url", "").strip(), 95)
                
            # Crear el producto con solo los campos válidos
            nuevo = Producto(
                nombre=nombre,
                stock=stock_int,
                costo=costo_val, 
                precio_venta=precio_val,
                categoria=categoria_normalizada,
                # MODIFICADO: Usar la nueva función de color de categoría
                categoria_color=get_category_color(categoria_normalizada),
                foto=foto_final,
                url_imagen=url_imagen_truncada,
                is_approved=True,
                empresa_id=session['user_id'],
                codigo_barras_externo=codigo_barras_externo,
                marca=marca,
                es_favorito=es_favorito_bool,
                esta_a_la_venta=esta_a_la_venta_bool,
                has_caducidad=has_caducidad,
                metodo_caducidad=caducidad_lapso
                # Eliminados los campos que no existen en el modelo
                # tipo_medida, unidad_medida, fabricacion, origen
            )
            
            # Guardar en la base de datos
            db.session.add(nuevo)
            db.session.commit()
            
            # ===== MODIFICADO: Crear siempre un lote y movimiento inicial si hay stock positivo =====
            if nuevo.stock > 0:
                try:
                    # Determinar la fecha de caducidad según el método
                    fecha_caducidad = None
                    if nuevo.has_caducidad and nuevo.metodo_caducidad:
                        fecha_actual = datetime.now()

                        if nuevo.metodo_caducidad == '1 día':
                            fecha_caducidad = fecha_actual + timedelta(days=1)
                        elif nuevo.metodo_caducidad == '3 días':
                            fecha_caducidad = fecha_actual + timedelta(days=3)
                        elif nuevo.metodo_caducidad == '1 semana':
                            fecha_caducidad = fecha_actual + timedelta(days=7)
                        elif nuevo.metodo_caducidad == '2 semanas':
                            fecha_caducidad = fecha_actual + timedelta(days=14)
                        elif nuevo.metodo_caducidad == '1 mes':
                            fecha_caducidad = fecha_actual + timedelta(days=30)
                        elif nuevo.metodo_caducidad == '3 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=90)
                        elif nuevo.metodo_caducidad == '6 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=180)
                        elif nuevo.metodo_caducidad == '1 año':
                            fecha_caducidad = fecha_actual + timedelta(days=365)
                        elif nuevo.metodo_caducidad == '2 años':
                            fecha_caducidad = fecha_actual + timedelta(days=730)
                        elif nuevo.metodo_caducidad == '3 años' or nuevo.metodo_caducidad == 'más de 3 años':
                            fecha_caducidad = fecha_actual + timedelta(days=1460)

                        # Convertir a date
                        if fecha_caducidad:
                            fecha_caducidad = fecha_caducidad.date()
                    
                    # Crear el lote inicial y el movimiento inicial
                    movimiento, lote = crear_lote_registro(
                        producto=nuevo,
                        cantidad=nuevo.stock,
                        costo=nuevo.costo,
                        fecha_caducidad=fecha_caducidad,
                        usuario_id=session['user_id']
                    )
                    
                    db.session.commit()
                    print(f"Lote inicial y movimiento creados para producto {nuevo.id}")

                except Exception as e:
                    db.session.rollback()
                    print(f"Error al crear lote inicial y movimiento: {str(e)}")
                    import traceback
                    traceback.print_exc()
                        
            flash('Producto sin código de barras guardado exitosamente', 'success')
            
            # MODIFICADO: Redireccionar a la página de confirmación
            response = make_response(redirect(url_for('producto_confirmacion', producto_id=nuevo.id)))
            response.set_cookie('pagina_origen', 'agregar_sin_codigo', max_age=300)  # Cookie válida por 5 minutos
            return response
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar el producto: {str(e)}', 'danger')
            print(f"ERROR al guardar producto sin código: {str(e)}")
        
        return redirect(url_for('ver_productos'))
        
    else:
        # GET - Cargar página de agregar producto sin código
        categorias_db = (
            db.session.query(Producto.categoria, Producto.categoria_color)
            .filter(Producto.categoria.isnot(None))
            .filter(Producto.categoria != '')
            .distinct()
            .limit(50)
            .all()
        )
        categories_list = [
            {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
            for c in categorias_db
        ]
        return render_template('agregar_sin_codigo.html', categories=categories_list)

@app.route('/agregar-a-granel', methods=['GET', 'POST'])
@login_requerido
def agregar_a_granel():
    if request.method == 'POST':
        try:
            # Recoger campos del formulario
            nombre = request.form.get("nombre", "").strip()
            stock_str = request.form.get("stock", "0").strip() 
            costo_str = request.form.get("costo", "$0").strip()
            precio_str = request.form.get("precio_venta", "$0").strip()
            marca = request.form.get("marca", "").strip()
            
            # Obtener unidad de medida (ya no usamos tipo_medida ni origen directamente)
            unidad_medida = request.form.get("unidad_medida", "").strip()
            
            # Generar código de barras único para productos a granel
            codigo_barras_externo = request.form.get("codigo_barras_externo", "").strip()
            if not codigo_barras_externo:
                codigo_barras_externo = generar_codigo_a_granel()
            
            # Categoría
            cat_option = request.form.get('categoria_option', 'existente')
            if cat_option == 'existente':
                categoria = request.form.get('categoria_existente', '').strip()
            else:
                categoria = request.form.get('categoria_nueva', '').strip()
                
            # MODIFICADO: Usar la nueva función de normalización
            categoria_normalizada = normalize_category(categoria)
            
            # Favorito y a la venta
            es_favorito_bool = (request.form.get("es_favorito", "0") == "1")
            esta_a_la_venta_bool = (request.form.get("esta_a_la_venta", "1") == "1")
            
            # Caducidad
            has_caducidad = (request.form.get("toggle_caducidad_estado", "DESACTIVADO") == "ACTIVADO")
            caducidad_lapso = request.form.get("caducidad_lapso", None) if has_caducidad else None
            
            # Parse numéricos
            try:
                stock_int = float(stock_str)  # Cambiado a float para soportar decimales
            except:
                stock_int = 0
            
            costo_val = parse_money(costo_str)
            precio_val = parse_money(precio_str)
            
            # ===== MANEJO SIMPLIFICADO DE IMÁGENES =====
            foto_final = None
            
            # 1. Primero intentar usar el archivo subido
            if request.files.get('foto') and request.files['foto'].filename:
                file = request.files['foto']
                if allowed_file(file.filename):
                    # Generar nombre único
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    
                    # Guardar archivo
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    
                    file.save(os.path.join(UPLOAD_FOLDER, filename))
                    foto_final = filename
            
            # 2. Si no hay archivo, usar displayed_image_url o ia_foto_filename
            if not foto_final:
                displayed_url = request.form.get("displayed_image_url", "").strip()
                ia_filename = request.form.get("ia_foto_filename", "").strip()
                
                if ia_filename and os.path.exists(os.path.join(UPLOAD_FOLDER, ia_filename)):
                    # Usar el archivo IA que ya existe
                    foto_final = ia_filename
                elif displayed_url:
                    # Es una URL - intentar descargar o usar nombre de archivo
                    if "/uploads/" in displayed_url:
                        # Es local, extraer nombre
                        foto_final = displayed_url.split("/uploads/")[-1]
                    elif displayed_url.startswith(("http://", "https://")):
                        # Es externa, descargar
                        try:
                            ext = "jpg"  # Default
                            filename = f"{uuid.uuid4().hex}.{ext}"
                            filepath = os.path.join(UPLOAD_FOLDER, filename)
                            
                            # Descarga síncrona para asegurar que se complete
                            response = requests.get(displayed_url, timeout=10)
                            if response.status_code == 200:
                                if not os.path.exists(os.path.dirname(filepath)):
                                    os.makedirs(os.path.dirname(filepath))
                                
                                with open(filepath, "wb") as f:
                                    f.write(response.content)
                                
                                foto_final = filename
                        except Exception as e:
                            print(f"Error descargando imagen: {e}")
            
            # 3. Si todo falla, usar imagen predeterminada
            if not foto_final:
                foto_final = "default_product.jpg"
            
            # Usar la nueva función para truncar la URL para asegurar que quepa en la columna
            url_imagen_truncada = truncar_url(request.form.get("displayed_image_url", "").strip(), 95)
            
            # Crear el producto - Solo usar campos que existen en el modelo
            nuevo = Producto(
                nombre=nombre,
                stock=stock_int,
                costo=costo_val, 
                precio_venta=precio_val,
                categoria=categoria_normalizada,
                # MODIFICADO: Usar la nueva función de color de categoría
                categoria_color=get_category_color(categoria_normalizada),
                foto=foto_final,
                url_imagen=url_imagen_truncada,
                is_approved=True,
                empresa_id=session['user_id'],
                codigo_barras_externo=codigo_barras_externo,
                marca=marca,
                es_favorito=es_favorito_bool,
                esta_a_la_venta=esta_a_la_venta_bool,
                has_caducidad=has_caducidad,
                metodo_caducidad=caducidad_lapso,
                unidad=unidad_medida  # Usar unidad correctamente
                # Ya no usamos "origen" que no existe en el modelo
            )
            
            # Guardar en la base de datos
            db.session.add(nuevo)
            db.session.commit()
            
            # ===== MODIFICADO: Crear siempre un lote y movimiento inicial si hay stock positivo =====
            if nuevo.stock > 0:
                try:
                    # Determinar la fecha de caducidad según el método
                    fecha_caducidad = None
                    if nuevo.has_caducidad and nuevo.metodo_caducidad:
                        fecha_actual = datetime.now()

                        if nuevo.metodo_caducidad == '1 día':
                            fecha_caducidad = fecha_actual + timedelta(days=1)
                        elif nuevo.metodo_caducidad == '3 días':
                            fecha_caducidad = fecha_actual + timedelta(days=3)
                        elif nuevo.metodo_caducidad == '1 semana':
                            fecha_caducidad = fecha_actual + timedelta(days=7)
                        elif nuevo.metodo_caducidad == '2 semanas':
                            fecha_caducidad = fecha_actual + timedelta(days=14)
                        elif nuevo.metodo_caducidad == '1 mes':
                            fecha_caducidad = fecha_actual + timedelta(days=30)
                        elif nuevo.metodo_caducidad == '3 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=90)
                        elif nuevo.metodo_caducidad == '6 meses':
                            fecha_caducidad = fecha_actual + timedelta(days=180)
                        elif nuevo.metodo_caducidad == '1 año':
                            fecha_caducidad = fecha_actual + timedelta(days=365)
                        elif nuevo.metodo_caducidad == '2 años':
                            fecha_caducidad = fecha_actual + timedelta(days=730)
                        elif nuevo.metodo_caducidad == '3 años' or nuevo.metodo_caducidad == 'más de 3 años':
                            fecha_caducidad = fecha_actual + timedelta(days=1460)

                        # Convertir a date
                        if fecha_caducidad:
                            fecha_caducidad = fecha_caducidad.date()
                    
                    # Crear el lote inicial y el movimiento inicial
                    movimiento, lote = crear_lote_registro(
                        producto=nuevo,
                        cantidad=nuevo.stock,
                        costo=nuevo.costo,
                        fecha_caducidad=fecha_caducidad,
                        usuario_id=session['user_id']
                    )
                    
                    db.session.commit()
                    print(f"Lote inicial y movimiento creados para producto a granel {nuevo.id}")

                except Exception as e:
                    db.session.rollback()
                    print(f"Error al crear lote inicial y movimiento: {str(e)}")
                    import traceback
                    traceback.print_exc()
                        
            flash('Producto a granel guardado exitosamente', 'success')
            
            # MODIFICADO: Redireccionar a la página de confirmación
            response = make_response(redirect(url_for('producto_confirmacion', producto_id=nuevo.id)))
            response.set_cookie('pagina_origen', 'agregar_a_granel', max_age=300)  # Cookie válida por 5 minutos
            return response
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar el producto a granel: {str(e)}', 'danger')
            print(f"ERROR al guardar producto a granel: {str(e)}")
        
        return redirect(url_for('ver_productos'))
        
    else:
        # GET - Cargar página de agregar producto a granel
        categorias_db = (
            db.session.query(Producto.categoria, Producto.categoria_color)
            .filter(Producto.categoria.isnot(None))
            .filter(Producto.categoria != '')
            .distinct()
            .limit(50)
            .all()
        )
        categories_list = [
            {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
            for c in categorias_db
        ]
        return render_template('agregar_a_granel.html', categories=categories_list)

@app.route('/editar-producto/<int:prod_id>', methods=['GET','POST'])
@login_requerido
def editar_producto(prod_id):
    """Edita un producto existente."""
    producto = Producto.query.get_or_404(prod_id)
    
    # Verificar que el producto pertenece a la empresa actual
    if producto.empresa_id != session.get('user_id'):
        flash('No tienes permiso para editar este producto.', 'danger')
        return redirect(url_for('ver_productos'))
    
    if request.method == 'POST':
        try:
            # Recoger SOLO los datos que vamos a modificar
            nombre = request.form.get("nombre", "").strip()
            precio_str = request.form.get("precio_venta", "$0").strip()
            marca = request.form.get("marca", "").strip()
            codigo_barras = request.form.get("codigo_barras_externo", "").strip()
            
            # Categoría
            categoria = request.form.get('categoria_existente', '').strip()
            categoria_normalizada = normalize_category(categoria)
            
            # Procesar precio
            precio_val = parse_money(precio_str)
            
            # Manejo de imagen
            nueva_foto = None
            
            # 1. Intentar usar el archivo subido
            if request.files.get('foto') and request.files['foto'].filename:
                file = request.files['foto']
                if allowed_file(file.filename):
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    
                    file.save(os.path.join(UPLOAD_FOLDER, filename))
                    nueva_foto = filename
            
            # 2. Si no hay archivo, usar displayed_image_url o ia_foto_filename
            if not nueva_foto:
                displayed_url = request.form.get("displayed_image_url", "").strip()
                ia_filename = request.form.get("ia_foto_filename", "").strip()
                
                if ia_filename and os.path.exists(os.path.join(UPLOAD_FOLDER, ia_filename)):
                    nueva_foto = ia_filename
                elif displayed_url:
                    if "/uploads/" in displayed_url:
                        nueva_foto = displayed_url.split("/uploads/")[-1]
                    elif displayed_url.startswith(("http://", "https://")):
                        try:
                            ext = "jpg"  # Default
                            filename = f"{uuid.uuid4().hex}.{ext}"
                            filepath = os.path.join(UPLOAD_FOLDER, filename)
                            
                            response = requests.get(displayed_url, timeout=10)
                            if response.status_code == 200:
                                if not os.path.exists(os.path.dirname(filepath)):
                                    os.makedirs(os.path.dirname(filepath))
                                
                                with open(filepath, "wb") as f:
                                    f.write(response.content)
                                
                                nueva_foto = filename
                        except Exception as e:
                            print(f"Error descargando imagen: {e}")
            
            url_imagen_truncada = truncar_url(request.form.get("displayed_image_url", "").strip(), 95)
            
            # Actualizar SOLO los campos que queremos modificar
            producto.nombre = nombre
            producto.precio_venta = precio_val
            producto.categoria = categoria_normalizada
            producto.categoria_color = get_category_color(categoria_normalizada)
            producto.codigo_barras_externo = codigo_barras
            producto.marca = marca
            
            # Actualizar foto solo si hay una nueva
            if nueva_foto:
                producto.foto = nueva_foto
                producto.url_imagen = url_imagen_truncada
            
            # Guardar en la base de datos
            db.session.commit()
            flash('Producto actualizado exitosamente', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el producto: {str(e)}', 'danger')
            print(f"ERROR al actualizar producto: {str(e)}")
        
        return redirect(url_for('ver_productos'))
    
    # GET - Cargar página de editar producto
    categorias_db = (
        db.session.query(Producto.categoria, Producto.categoria_color)
        .filter(Producto.categoria.isnot(None))
        .filter(Producto.categoria != '')
        .distinct()
        .limit(50)
        .all()
    )
    categories_list = [
        {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
        for c in categorias_db
    ]
    
    # Verificar si la categoría del producto existe en la lista
    cat_encontrada = any(cat["nombre"] == producto.categoria for cat in categories_list)
    
    return render_template(
        'editar_producto.html', 
        producto=producto,
        categories=categories_list,
        cat_encontrada=cat_encontrada
    )

@app.route('/eliminar-producto/<int:prod_id>')
@login_requerido
def eliminar_producto(prod_id):
    """Elimina un producto del inventario y todos sus datos relacionados."""
    try:
        # Verificar que el producto existe y pertenece a la empresa actual
        producto = Producto.query.get_or_404(prod_id)
        
        if producto.empresa_id != session.get('user_id'):
            flash('No tienes permiso para eliminar este producto.', 'danger')
            return redirect(url_for('ver_productos'))
        
        # 1. Eliminar las relaciones entre lotes y movimientos
        LoteMovimientoRelacion.query.filter(
            LoteMovimientoRelacion.lote_id.in_(
                db.session.query(LoteInventario.id).filter_by(producto_id=prod_id)
            )
        ).delete(synchronize_session=False)
        
        # 2. Eliminar todos los movimientos de inventario del producto
        MovimientoInventario.query.filter_by(producto_id=prod_id).delete()
        
        # 3. Eliminar todos los lotes de inventario del producto
        LoteInventario.query.filter_by(producto_id=prod_id).delete()
        
        # 4. Finalmente, eliminar el producto
        db.session.delete(producto)
        
        # Confirmar todos los cambios
        db.session.commit()
        flash('Producto y todos sus datos relacionados eliminados correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el producto: {str(e)}', 'danger')
        print(f"ERROR al eliminar producto y datos relacionados: {str(e)}")
    
    # Esta línea es crucial - asegura que siempre se devuelva una respuesta
    return redirect(url_for('ver_productos'))

@app.route('/eliminar-todos-productos')
@login_requerido
def eliminar_todos_productos():
    """Elimina TODOS los productos del inventario de la empresa y todos sus datos relacionados."""
    try:
        # Verificar que el usuario está logueado
        empresa_id = session.get('user_id')
        if not empresa_id:
            flash('No tienes permiso para realizar esta acción.', 'danger')
            return redirect(url_for('ver_productos'))
        
        # Obtener todos los productos de la empresa
        productos = Producto.query.filter_by(empresa_id=empresa_id).all()
        
        if not productos:
            flash('No hay productos para eliminar.', 'info')
            return redirect(url_for('ver_productos'))
        
        productos_eliminados = 0
        
        # Eliminar cada producto y sus datos relacionados
        for producto in productos:
            try:
                # 1. Eliminar las relaciones entre lotes y movimientos de este producto
                LoteMovimientoRelacion.query.filter(
                    LoteMovimientoRelacion.lote_id.in_(
                        db.session.query(LoteInventario.id).filter_by(producto_id=producto.id)
                    )
                ).delete(synchronize_session=False)
                
                # 2. Eliminar todos los movimientos de inventario del producto
                MovimientoInventario.query.filter_by(producto_id=producto.id).delete()
                
                # 3. Eliminar todos los lotes de inventario del producto
                LoteInventario.query.filter_by(producto_id=producto.id).delete()
                
                # 4. Finalmente, eliminar el producto
                db.session.delete(producto)
                
                productos_eliminados += 1
                
            except Exception as e:
                print(f"ERROR al eliminar producto {producto.id}: {str(e)}")
                # Continuar con los demás productos en caso de error en uno específico
                continue
        
        # Confirmar todos los cambios
        db.session.commit()
        
        if productos_eliminados > 0:
            flash(f'Se eliminaron {productos_eliminados} productos y todos sus datos relacionados correctamente', 'success')
        else:
            flash('No se pudo eliminar ningún producto.', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar los productos: {str(e)}', 'danger')
        print(f"ERROR al eliminar todos los productos: {str(e)}")
    
    # Siempre redirigir a la página de productos
    return redirect(url_for('ver_productos'))

@app.route('/historial-movimientos')
@login_requerido
def historial_movimientos():
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # Obtener ID de empresa del usuario actual
    empresa_id = session.get('user_id')
    
    # Obtener parámetros de filtro
    tipo = request.args.get('tipo', 'todos')
    periodo = int(request.args.get('periodo', 30))
    busqueda = request.args.get('q', '')
    page = int(request.args.get('page', 1))
    per_page = 30  # Movimientos por página
    
    # Calcular fecha límite según el período seleccionado
    fecha_limite = datetime.now() - timedelta(days=periodo)
    
    # Consulta base CON FILTRO POR EMPRESA_ID
    # Modificación aquí: Excluir movimientos de corrección de costo
    query = MovimientoInventario.query.join(Producto).filter(
        MovimientoInventario.fecha_movimiento >= fecha_limite,
        Producto.empresa_id == empresa_id,  # Filtro importante por empresa_id
        MovimientoInventario.motivo != 'corrección de costo'  # NUEVA LÍNEA: filtrar correcciones de costo
    )
    
    # Aplicar filtro de tipo
    if tipo == 'entrada':
        query = query.filter(MovimientoInventario.tipo_movimiento == 'ENTRADA')
    elif tipo == 'salida':
        query = query.filter(MovimientoInventario.tipo_movimiento == 'SALIDA')
    
    # Aplicar búsqueda
    if busqueda:
        query = query.filter(
            or_(
                Producto.nombre.ilike(f'%{busqueda}%'),
                Producto.codigo_barras_externo.ilike(f'%{busqueda}%')
            )
        )
    
    # Ordenar por fecha descendente (más reciente primero)
    query = query.order_by(MovimientoInventario.fecha_movimiento.desc())
    
    # Contar total de resultados
    total_movimientos = query.count()
    
    # Aplicar paginación
    movimientos = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # Determinar si hay página siguiente
    has_next = page * per_page < total_movimientos
    
    return render_template(
        'historial_movimientos.html',
        movimientos=movimientos.items,
        current_page=page,
        has_next=has_next,
        total_movimientos=total_movimientos
    )

@app.route('/generar-movimientos-iniciales-debug')
@login_requerido
def generar_movimientos_iniciales_debug():
    """
    Versión depurada para generar movimientos iniciales para productos sin movimiento inicial.
    """
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # Obtener ID de empresa del usuario actual
    empresa_id = session.get('user_id')
    
    # Iniciar registro HTML para visualizar resultados
    resultado_html = "<h2>Diagnóstico de Movimientos Iniciales</h2>"
    resultado_html += "<style>.success{color:green;} .error{color:red;} .warning{color:orange;}</style>"
    resultado_html += "<ul>"
    
    try:
        # Obtener todos los productos aprobados con stock positivo
        productos = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True
        ).filter(Producto.stock > 0).all()
        
        resultado_html += f"<li>Encontrados {len(productos)} productos con stock positivo</li>"
        
        productos_procesados = 0
        movimientos_creados = 0
        lotes_creados = 0
        productos_con_errores = 0
        
        for producto in productos:
            try:
                productos_procesados += 1
                resultado_html += f"<li>Procesando producto ID {producto.id}: {producto.nombre} (Stock: {producto.stock})</li>"
                
                # 1. Verificar si ya existe movimiento inicial
                movimientos = MovimientoInventario.query.filter_by(
                    producto_id=producto.id
                ).all()
                
                movimiento_inicial_existe = False
                for m in movimientos:
                    if m.motivo == 'Registro inicial' or (m.motivo and 'registro' in m.motivo.lower()):
                        movimiento_inicial_existe = True
                        resultado_html += f"<li class='warning'>- Ya tiene movimiento inicial (ID: {m.id})</li>"
                        break
                
                # 2. Verificar si ya existe lote de registro
                lotes = LoteInventario.query.filter_by(
                    producto_id=producto.id
                ).all()
                
                lote_registro_existe = False
                for l in lotes:
                    if l.numero_lote == "Lote de Registro":
                        lote_registro_existe = True
                        resultado_html += f"<li class='warning'>- Ya tiene Lote de Registro (ID: {l.id}, Stock: {l.stock})</li>"
                        break
                
                # 3. Crear movimiento y lote si no existen
                if not movimiento_inicial_existe or not lote_registro_existe:
                    # Determinar fecha_caducidad si aplica
                    fecha_caducidad = None
                    if hasattr(producto, 'has_caducidad') and producto.has_caducidad and producto.metodo_caducidad:
                        try:
                            fecha_actual = datetime.now()
                            
                            if producto.metodo_caducidad == '1 día':
                                fecha_caducidad = fecha_actual + timedelta(days=1)
                            elif producto.metodo_caducidad == '3 días':
                                fecha_caducidad = fecha_actual + timedelta(days=3)
                            elif producto.metodo_caducidad == '1 semana':
                                fecha_caducidad = fecha_actual + timedelta(days=7)
                            elif producto.metodo_caducidad == '2 semanas':
                                fecha_caducidad = fecha_actual + timedelta(days=14)
                            elif producto.metodo_caducidad == '1 mes':
                                fecha_caducidad = fecha_actual + timedelta(days=30)
                            elif producto.metodo_caducidad == '3 meses':
                                fecha_caducidad = fecha_actual + timedelta(days=90)
                            elif producto.metodo_caducidad == '6 meses':
                                fecha_caducidad = fecha_actual + timedelta(days=180)
                            elif producto.metodo_caducidad == '1 año':
                                fecha_caducidad = fecha_actual + timedelta(days=365)
                            elif producto.metodo_caducidad == '2 años':
                                fecha_caducidad = fecha_actual + timedelta(days=730)
                            elif producto.metodo_caducidad == '3 años':
                                fecha_caducidad = fecha_actual + timedelta(days=1460)
                            
                            # Convertir a date
                            if fecha_caducidad:
                                fecha_caducidad = fecha_caducidad.date()
                        except Exception as e:
                            resultado_html += f"<li class='error'>- Error al calcular fecha de caducidad: {str(e)}</li>"
                    
                    # Crear movimiento si no existe
                    if not movimiento_inicial_existe:
                        try:
                            costo = producto.costo if hasattr(producto, 'costo') and producto.costo is not None else 0
                            nuevo_movimiento = MovimientoInventario(
                                producto_id=producto.id,
                                tipo_movimiento='ENTRADA',
                                cantidad=float(producto.stock),
                                motivo='Registro inicial',
                                fecha_movimiento=datetime.now(),
                                costo_unitario=float(costo),
                                numero_lote="Lote de Registro",
                                fecha_caducidad=fecha_caducidad,
                                usuario_id=empresa_id
                            )
                            db.session.add(nuevo_movimiento)
                            db.session.flush()  # Para obtener el ID asignado
                            
                            movimientos_creados += 1
                            resultado_html += f"<li class='success'>- Creado movimiento inicial ID: {nuevo_movimiento.id}</li>"
                        except Exception as e:
                            resultado_html += f"<li class='error'>- Error al crear movimiento: {str(e)}</li>"
                            import traceback
                            resultado_html += f"<li class='error'>- Trace: {traceback.format_exc()}</li>"
                    
                    # Crear lote si no existe
                    if not lote_registro_existe:
                        try:
                            costo = producto.costo if hasattr(producto, 'costo') and producto.costo is not None else 0
                            nuevo_lote = LoteInventario(
                                producto_id=producto.id,
                                numero_lote="Lote de Registro",
                                costo_unitario=float(costo),
                                stock=float(producto.stock),
                                fecha_entrada=datetime.now(),
                                fecha_caducidad=fecha_caducidad,
                                esta_activo=True
                            )
                            db.session.add(nuevo_lote)
                            db.session.flush()  # Para obtener el ID asignado
                            
                            lotes_creados += 1
                            resultado_html += f"<li class='success'>- Creado lote de registro ID: {nuevo_lote.id}</li>"
                        except Exception as e:
                            resultado_html += f"<li class='error'>- Error al crear lote: {str(e)}</li>"
                            import traceback
                            resultado_html += f"<li class='error'>- Trace: {traceback.format_exc()}</li>"
                
                # Actualizar todo el producto cada vez (commit por producto)
                db.session.commit()
                resultado_html += "<li class='success'>- Commit exitoso para este producto</li>"
                
            except Exception as e:
                db.session.rollback()
                productos_con_errores += 1
                resultado_html += f"<li class='error'>Error general para producto {producto.id}: {str(e)}</li>"
                import traceback
                resultado_html += f"<li class='error'>Trace: {traceback.format_exc()}</li>"
        
        # Resumen final
        resultado_html += "</ul>"
        resultado_html += f"""
        <h3>Resumen del Proceso</h3>
        <ul>
            <li>Productos procesados: {productos_procesados}</li>
            <li>Movimientos iniciales creados: {movimientos_creados}</li>
            <li>Lotes de registro creados: {lotes_creados}</li>
            <li>Productos con errores: {productos_con_errores}</li>
        </ul>
        
        <p>
            <a href="{url_for('historial_movimientos')}?periodo=365" 
               style="display:inline-block; padding:10px 20px; background-color:#e52e2e; 
                      color:white; text-decoration:none; border-radius:5px;">
                Ver Historial de Movimientos (Último año)
            </a>
        </p>
        """
        
        return resultado_html
        
    except Exception as e:
        db.session.rollback()
        import traceback
        return f"""
        <h2>Error General en el Proceso</h2>
        <p style="color:red;">Se produjo un error: {str(e)}</p>
        <pre>{traceback.format_exc()}</pre>
        <p>
            <a href="{url_for('dashboard_home')}" style="color:#0066cc;">Volver al Dashboard</a>
        </p>
        """

##############################
# ESCÁNER / CAPTURA RÁPIDA
##############################
@app.route('/inventario-escaner', methods=['GET','POST'])
@login_requerido
def inventario_escaner():
    return render_template('inventario_escaner.html')

@app.route('/pendientes_aprobacion')
@login_requerido
def pendientes_aprobacion():
    empresa_id = session['user_id']
    pendientes = Producto.query.filter_by(
        empresa_id=empresa_id, 
        is_approved=False
    ).all()
    return render_template('pendientes_aprobacion.html', pendientes=pendientes)

@app.route('/completar-datos/<int:prod_id>', methods=['GET','POST'])
@login_requerido
def completar_datos(prod_id):
    producto = Producto.query.get_or_404(prod_id)
    
    # Verificar que el producto pertenece a la empresa actual
    if producto.empresa_id != session.get('user_id'):
        flash('No tienes permiso para editar este producto.', 'danger')
        return redirect(url_for('pendientes_aprobacion'))
    
    if request.method == 'POST':
        try:
            # Actualizar campos...
            producto.nombre = request.form.get('nombre', '')
            producto.stock = int(request.form.get('stock', 0))
            producto.costo = float(request.form.get('costo', 0))
            producto.precio_venta = float(request.form.get('precio_venta', 0))
            
            # Categoría
            cat_option = request.form.get('categoria_option')
            if cat_option == 'existente':
                categoria = request.form.get('categoria_existente', '')
            else:
                categoria = request.form.get('categoria_nueva', '')
                
            # MODIFICADO: Usar la nueva función para normalizar la categoría
            categoria_normalizada = normalize_category(categoria)
            producto.categoria = categoria_normalizada
            
            # MODIFICADO: Usar la nueva función para obtener el color de la categoría
            producto.categoria_color = get_category_color(categoria_normalizada)
            
            # Aprobar el producto
            producto.is_approved = True
            
            db.session.commit()
            flash('Producto aprobado y actualizado con éxito.')
            return redirect(url_for('pendientes_aprobacion'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}')
    
    # Obtener lista de categorías
    categorias_db = (
        db.session.query(Producto.categoria, Producto.categoria_color)
        .filter(Producto.categoria.isnot(None))
        .filter(Producto.categoria != '')
        .distinct()
        .all()
    )
    categories_list = [
        {"nombre": c[0], "color": c[1] if c[1] else "#000000"}
        for c in categorias_db
    ]
    
    # Verificar si la categoría del producto existe en la lista
    cat_encontrada = any(cat["nombre"] == producto.categoria for cat in categories_list)
    
    return render_template(
        'completar_datos.html',
        producto=producto,
        categories=categories_list,
        cat_encontrada=cat_encontrada
    )

@app.route('/nuevo-producto')
@login_requerido
def nuevo_producto():
    return render_template('nuevo_producto.html')

@app.route('/producto-confirmacion/<int:producto_id>')
@login_requerido
def producto_confirmacion(producto_id):
    """Muestra la confirmación de un producto recién agregado."""
    # Verificar si el usuario está logueado
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    empresa_id = session.get('user_id')
    
    # Obtener el producto
    producto = Producto.query.get_or_404(producto_id)
    
    # Verificar que el producto pertenezca a la empresa del usuario actual
    if producto.empresa_id != empresa_id:
        flash('No tienes permiso para ver este producto.', 'danger')
        return redirect(url_for('ver_productos'))
    
    # Determinar la página de origen para el botón "Volver"
    # Lo almacenamos en una cookie temporal
    pagina_origen = request.cookies.get('pagina_origen', 'agregar_producto')
    
    return render_template(
        'producto_confirmacion.html',
        producto=producto,
        pagina_origen=pagina_origen
    )

@app.route('/etiquetas-producto/<int:producto_id>', methods=['GET'])
@login_requerido
def etiquetas_producto(producto_id):
    """
    Vista para generar etiquetas para un producto específico.
    """
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    empresa_id = session.get('user_id')
    
    # Obtener el producto
    producto = Producto.query.filter_by(
        id=producto_id, 
        empresa_id=empresa_id
    ).first_or_404()
    
    # Calcular precio con descuento si tiene
    try:
        precio_final = producto.precio_venta or 0
        if hasattr(producto, 'tiene_descuento') and producto.tiene_descuento:
            if producto.tipo_descuento == 'percentage':
                precio_final = producto.precio_venta * (1 - producto.valor_descuento / 100)
            else:  # fixed amount
                precio_final = max(0, producto.precio_venta - producto.valor_descuento)
    except:
        precio_final = producto.precio_venta or 0
    
    # CRÍTICO: Verificar y asegurar que hay un código de barras válido
    if not producto.codigo_barras_externo or producto.codigo_barras_externo.strip() == "":
        # Intentar generar un código temporal
        try:
            # Prioridad 1: Usar código interno si existe
            if hasattr(producto, 'codigo_interno') and producto.codigo_interno:
                producto.codigo_barras_externo = producto.codigo_interno
            # Prioridad 2: Generar un código basado en el ID y timestamp
            else:
                from datetime import datetime
                import uuid
                # Crear código compatible con códigos de barras (solo números)
                codigo_temp = f"9{producto.id:08d}{int(datetime.now().timestamp()) % 10000:04d}"
                producto.codigo_barras_externo = codigo_temp
            
            # Guardar en base de datos
            db.session.commit()
            flash('Se ha generado un código de barras temporal para este producto.', 'warning')
        except Exception as e:
            flash(f'No se pudo generar un código de barras: {str(e)}', 'danger')
    
    # Resto del código sin cambios...
    formatos_etiquetas = [...]
    impresoras_compatibles = [...]
    
    return render_template(
        'etiquetas_producto.html',
        producto=producto,
        precio_final=precio_final,
        formatos_etiquetas=formatos_etiquetas,
        impresoras_compatibles=impresoras_compatibles
    )

##############################
# DESCUENTOS
##############################
@app.route('/api/apply_discount/<int:product_id>', methods=['POST'])
@login_requerido
def apply_discount(product_id):
    """Aplica un descuento a un producto específico con rastreo mejorado."""
    try:
        producto = Producto.query.get_or_404(product_id)
        
        # Verificar que el producto pertenezca a la empresa del usuario
        if producto.empresa_id != session.get('user_id'):
            return jsonify({"success": False, "message": "Sin permisos"}), 403
        
        data = request.get_json()
        discount_type = data.get('type')  # 'percentage' o 'fixed'
        discount_value = float(data.get('value', 0))
        discount_origin = data.get('origin', 'individual')
        discount_group_id = data.get('group_id', None)
        
        # Validar datos
        if discount_type not in ['percentage', 'fixed'] or discount_value <= 0:
            return jsonify({"success": False, "message": "Datos de descuento inválidos"}), 400
        
        # Validar porcentaje
        if discount_type == 'percentage' and discount_value > 100:
            return jsonify({"success": False, "message": "El porcentaje no puede ser mayor a 100"}), 400
        
        # Usar el método del modelo que actualiza precio_final correctamente
        precio_final = producto.aplicar_descuento(
            valor=discount_value,
            tipo=discount_type,
            origen=discount_origin,
            grupo_id=discount_group_id
        )
        
        # Si es descuento global, aplicar a todos los productos
        if discount_origin == 'global':
            try:
                empresa_id = session.get('user_id')
                # Solo actualizar otros productos, no el actual que ya fue actualizado
                otros_productos = Producto.query.filter(
                    Producto.empresa_id == empresa_id,
                    Producto.id != product_id
                ).all()
                
                for otro_producto in otros_productos:
                    # Usar el método del modelo para cada producto
                    otro_producto.aplicar_descuento(
                        valor=discount_value,
                        tipo=discount_type,
                        origen='global',
                        grupo_id=None
                    )
            except Exception as e:
                print(f"Error al aplicar descuento global: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "precio_base": producto.precio_venta,
            "precio_final": precio_final,
            "tipo_descuento": discount_type,
            "valor_descuento": discount_value,
            "origen_descuento": discount_origin,
            "descuento_grupo_id": discount_group_id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/remove_discount/<int:product_id>', methods=['POST'])
@login_requerido
def remove_discount(product_id):
    """Remueve el descuento de un producto específico."""
    try:
        producto = Producto.query.get_or_404(product_id)
        
        # Verificar que el producto pertenezca a la empresa del usuario
        if producto.empresa_id != session.get('user_id'):
            return jsonify({"success": False, "message": "Sin permisos"}), 403
        
        # Usar el método del modelo que actualiza precio_final correctamente
        precio_final_actualizado = producto.quitar_descuento()
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "precio_base": producto.precio_venta,
            "precio_final": precio_final_actualizado
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/get_global_discount_status', methods=['GET'])
@login_requerido
def get_global_discount_status():
    """Obtiene el estado del descuento global"""
    try:
        empresa_id = session.get('user_id')
        if not empresa_id:
            return jsonify({"success": False, "message": "Usuario no autenticado"}), 401
        
        # Contar productos con descuento global
        productos_con_global = Producto.query.filter_by(
            empresa_id=empresa_id,
            origen_descuento='global',
            tiene_descuento=True
        ).count()
        
        # Contar total de productos
        total_productos = Producto.query.filter_by(
            empresa_id=empresa_id
        ).count()
        
        # Si hay al menos un producto con descuento global, consideramos que está activo
        is_global_active = productos_con_global > 0
        
        # Obtener valor del descuento global (tomando el primero que encontremos)
        valor_descuento = 0
        tipo_descuento = None
        primer_producto_global = Producto.query.filter_by(
            empresa_id=empresa_id,
            origen_descuento='global',
            tiene_descuento=True
        ).first()
        
        if primer_producto_global:
            valor_descuento = primer_producto_global.valor_descuento
            tipo_descuento = primer_producto_global.tipo_descuento
        
        return jsonify({
            "success": True,
            "is_global_active": is_global_active,
            "productos_con_global": productos_con_global,
            "total_productos": total_productos,
            "valor_descuento": valor_descuento,
            "tipo_descuento": tipo_descuento
        })
        
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/get_active_discounts', methods=['GET'])
@login_requerido
def get_active_discounts():
    """Obtiene todos los tipos de descuentos activos para facilitar visualización"""
    try:
        empresa_id = session.get('user_id')
        if not empresa_id:
            return jsonify({"success": False, "message": "Usuario no autenticado"}), 401
        
        # Estructura para almacenar los resultados
        active_discounts = {
            "global": False,
            "categorias": [],
            "marcas": [],
            "individuales": 0
        }
        
        # Verificar descuento global (basta 1 producto)
        active_discounts["global"] = Producto.query.filter_by(
            empresa_id=empresa_id,
            origen_descuento='global',
            tiene_descuento=True
        ).limit(1).count() > 0
        
        # Obtener categorías con descuentos
        categorias_con_descuento = db.session.query(Producto.descuento_grupo_id).filter(
            Producto.empresa_id == empresa_id,
            Producto.origen_descuento == 'categoria',
            Producto.tiene_descuento == True
        ).distinct().all()
        
        active_discounts["categorias"] = [cat[0] for cat in categorias_con_descuento if cat[0]]
        
        # Obtener marcas con descuentos
        marcas_con_descuento = db.session.query(Producto.descuento_grupo_id).filter(
            Producto.empresa_id == empresa_id,
            Producto.origen_descuento == 'marca',
            Producto.tiene_descuento == True
        ).distinct().all()
        
        active_discounts["marcas"] = [marca[0] for marca in marcas_con_descuento if marca[0]]
        
        # Contar productos con descuentos individuales
        active_discounts["individuales"] = Producto.query.filter_by(
            empresa_id=empresa_id,
            origen_descuento='individual',
            tiene_descuento=True
        ).count()
        
        return jsonify({
            "success": True,
            "active_discounts": active_discounts
        })
        
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/descuentos/detalle/<tipo>/<grupo_id>')
@login_requerido
def descuentos_detalle(tipo, grupo_id):
    """Muestra los detalles de productos con descuento por tipo y grupo"""
    empresa_id = session['user_id']
    
    # Decodificar grupo_id (puede contener espacios)
    import urllib.parse
    grupo_id = urllib.parse.unquote(grupo_id)
    
    # Filtrar productos según el tipo de descuento
    if tipo == 'global':
        productos = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True,
            tiene_descuento=True,
            origen_descuento='global'
        ).all()
        titulo = "Descuento Global"
    elif tipo == 'categoria':
        productos = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True,
            tiene_descuento=True,
            origen_descuento='categoria',
            descuento_grupo_id=grupo_id
        ).all()
        titulo = f"Categoría: {grupo_id}"
    elif tipo == 'marca':
        productos = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True,
            tiene_descuento=True,
            origen_descuento='marca',
            descuento_grupo_id=grupo_id
        ).all()
        titulo = f"Marca: {grupo_id}"
    else:
        # Por defecto, mostrar productos individuales
        productos = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True,
            tiene_descuento=True,
            origen_descuento='individual'
        ).all()
        titulo = "Productos Individuales"
    
    return render_template(
        'descuentos_detalle.html',
        productos=productos,
        titulo=titulo,
        tipo=tipo,
        grupo_id=grupo_id
    )

@app.route('/descuentos')
@login_requerido
def descuentos():
    """Vista para aplicar descuentos a productos."""
    try:
        # Obtener información del usuario
        empresa_id = session.get('user_id')
        
        if not empresa_id:
            flash('Error: Usuario no identificado', 'error')
            return redirect(url_for('login'))
        
        # Obtener lista de productos aprobados
        productos_db = Producto.query.filter_by(
            empresa_id=empresa_id,
            is_approved=True
        ).order_by(Producto.categoria).all()
        
        # Convertir objetos Producto a diccionarios simples - INCLUYENDO NUEVOS CAMPOS
        productos = []
        for p in productos_db:
            try:
                producto_dict = {
                    'id': p.id,
                    'nombre': p.nombre or '',
                    'stock': p.stock or 0,
                    'precio_venta': p.precio_venta or 0,
                    'categoria': p.categoria or '',
                    'categoria_color': p.categoria_color or '#6B7280',
                    'foto': p.foto or 'default_product.jpg',
                    'codigo_barras_externo': p.codigo_barras_externo or '',
                    'marca': p.marca or '',
                    'es_favorito': p.es_favorito or False,
                    'esta_a_la_venta': p.esta_a_la_venta or True,
                    'tiene_descuento': getattr(p, 'tiene_descuento', False),
                    'tipo_descuento': getattr(p, 'tipo_descuento', None),
                    'valor_descuento': float(getattr(p, 'valor_descuento', 0.0)),
                    # NUEVOS CAMPOS DE RASTREO
                    'origen_descuento': getattr(p, 'origen_descuento', None),
                    'descuento_grupo_id': getattr(p, 'descuento_grupo_id', None),
                    'fecha_aplicacion_descuento': getattr(p, 'fecha_aplicacion_descuento', None)
                }
                productos.append(producto_dict)
            except Exception as e:
                print(f"Error procesando producto {p.id}: {str(e)}")
                continue
        
        # Obtener categorías únicas y sus colores
        categorias = []
        categorias_set = set()
        
        for producto in productos_db:
            if producto.categoria and producto.categoria not in categorias_set:
                categorias_set.add(producto.categoria)
                categorias.append({
                    'nombre': producto.categoria,
                    'color': producto.categoria_color or '#6B7280'  # Color por defecto si no tiene
                })
        
        # Ordenar categorías alfabéticamente
        categorias.sort(key=lambda x: x['nombre'])
        
        # Renderizar template con datos simplificados
        return render_template(
            'descuentos.html',
            productos=productos,
            categorias=categorias
        )
        
    except Exception as e:
        print(f"Error en ruta descuentos: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # En caso de error, mostrar mensaje amigable
        flash('Error al cargar la página de descuentos', 'error')
        return redirect(url_for('dashboard_inventario'))

@app.route('/api/search_products', methods=['GET'])
@login_requerido
def api_search_products():
    """
    Busca productos por nombre, marca o código para el selector de descuentos.
    """
    query = request.args.get('q', '').strip()
    empresa_id = session['user_id']
    
    if len(query) < 2:
        return jsonify({"results": []})
    
    # Buscar productos que coincidan
    productos = Producto.query.filter(
        Producto.empresa_id == empresa_id,
        Producto.is_approved == True,
        or_(
            Producto.nombre.ilike(f"%{query}%"),
            Producto.marca.ilike(f"%{query}%"),
            Producto.codigo_barras_externo.ilike(f"%{query}%")
        )
    ).limit(10).all()
    
    # Convertir a diccionarios
    results = []
    for p in productos:
        results.append({
            'id': p.id,
            'nombre': p.nombre,
            'marca': p.marca or 'Sin marca',
            'codigo_barras_externo': p.codigo_barras_externo or 'N/A',
            'foto': p.foto or 'default_product.jpg',
            'tiene_descuento': getattr(p, 'tiene_descuento', False),
            'valor_descuento': getattr(p, 'valor_descuento', 0.0)
        })
    
    return jsonify({"results": results})
    
@app.route('/api/fix-existing-discounts', methods=['POST'])
@login_requerido
def fix_existing_discounts():
    """Corrige los descuentos existentes que no tienen origen_descuento o descuento_grupo_id."""
    try:
        empresa_id = session.get('user_id')
        if not empresa_id:
            return jsonify({"success": False, "message": "Usuario no autenticado"}), 401
        
        # Obtener todos los productos con descuento pero sin origen
        productos_descuento_sin_origen = Producto.query.filter_by(
            empresa_id=empresa_id,
            tiene_descuento=True,
            origen_descuento=None
        ).all()
        
        fixed_count = 0
        
        # Asumir que son descuentos individuales
        for producto in productos_descuento_sin_origen:
            producto.origen_descuento = 'individual'
            producto.fecha_aplicacion_descuento = datetime.now()
            fixed_count += 1
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"Se han corregido {fixed_count} productos con descuentos",
            "count": fixed_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/corregir-descuentos', methods=['GET'])
@login_requerido
def corregir_descuentos_vista():
    """Página para corregir descuentos manualmente."""
    try:
        empresa_id = session.get('user_id')
        
        # Corregir productos con tiene_descuento=True pero sin origen_descuento
        productos_para_corregir = Producto.query.filter_by(
            empresa_id=empresa_id,
            tiene_descuento=True
        ).filter(
            (Producto.origen_descuento.is_(None)) | 
            (Producto.origen_descuento == '')
        ).all()
        
        productos_corregidos = 0
        
        for producto in productos_para_corregir:
            producto.origen_descuento = 'individual'
            productos_corregidos += 1
        
        # Obtener productos con descuentos globales
        productos_descuento_global = Producto.query.filter_by(
            empresa_id=empresa_id,
            tiene_descuento=True,
            origen_descuento='global'
        ).all()
        
        # Corregir productos con descuentos de categoría
        productos_categoria = db.session.query(Producto).filter(
            Producto.empresa_id == empresa_id,
            Producto.tiene_descuento == True,
            Producto.origen_descuento == 'categoria',
            Producto.descuento_grupo_id.is_(None)
        ).all()
        
        for producto in productos_categoria:
            producto.descuento_grupo_id = producto.categoria
        
        # Corregir productos con descuentos de marca
        productos_marca = db.session.query(Producto).filter(
            Producto.empresa_id == empresa_id,
            Producto.tiene_descuento == True,
            Producto.origen_descuento == 'marca',
            Producto.descuento_grupo_id.is_(None)
        ).all()
        
        for producto in productos_marca:
            producto.descuento_grupo_id = producto.marca
        
        db.session.commit()
        
        mensaje = f"""
        <html>
        <head>
            <title>Corrección de Descuentos</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 20px; }}
                .result {{ margin: 20px 0; padding: 15px; border: 1px solid #ddd; border-radius: 5px; }}
                .success {{ color: green; }}
                a {{ display: inline-block; margin-top: 20px; padding: 10px 15px; background: #e52e2e; color: white; text-decoration: none; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <h1>Corrección de Descuentos</h1>
            <div class="result">
                <p class="success">✅ Se han corregido {productos_corregidos} productos con descuentos sin origen.</p>
                <p>Productos con descuento global: {len(productos_descuento_global)}</p>
                <p>Productos con descuento por categoría: {len(productos_categoria)}</p>
                <p>Productos con descuento por marca: {len(productos_marca)}</p>
            </div>
            <a href="/descuentos">Volver a Descuentos</a>
        </body>
        </html>
        """
        
        return mensaje
    
    except Exception as e:
        db.session.rollback()
        return f"""
        <h1>Error al corregir descuentos</h1>
        <p>Se produjo un error: {str(e)}</p>
        <p><a href="/descuentos">Volver a Descuentos</a></p>
        """

##############################
# REABASTECER
##############################
@app.route('/reabastecer', methods=['GET'])
@login_requerido
def reabastecer_listado():
    # Obtener productos aprobados de la empresa
    empresa_id = session['user_id']
    productos = Producto.query.filter_by(empresa_id=empresa_id, is_approved=True).all()
    return render_template('reabastecer_listado.html', productos=productos)

@app.route('/reabastecer-producto/<int:prod_id>', methods=['GET','POST'])
@login_requerido
def reabastecer_producto(prod_id):
    producto = Producto.query.get_or_404(prod_id)
    
    # Verificar que el producto pertenece a la empresa actual
    if producto.empresa_id != session.get('user_id'):
        return "No tienes permiso para reabastecer este producto."
    
    if request.method == 'POST':
        try:
            stock_to_add = int(request.form.get('stock_to_add', 0))
            producto.stock += stock_to_add
            db.session.commit()
            flash(f'Se han añadido {stock_to_add} unidades al stock de {producto.nombre}.', 'success')
            return redirect(url_for('reabastecer_listado'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al reabastecer: {str(e)}', 'danger')
    
    return render_template('reabastecer_producto.html', producto=producto)

##############################
# AUTOCOMPLETE CON CatalogoGlobal
##############################
@app.route('/api/autocomplete', methods=['GET'])
@login_requerido
def api_autocomplete():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({"results": []})

    # Consulta optimizada con límite y selección específica de campos
    referencias = (
        CatalogoGlobal.query
        .with_entities(
            CatalogoGlobal.id,
            CatalogoGlobal.codigo_barras,
            CatalogoGlobal.nombre,
            CatalogoGlobal.marca,
            CatalogoGlobal.url_imagen,
            CatalogoGlobal.categoria
        )
        .filter(CatalogoGlobal.codigo_barras.ilike(f"%{q}%"))
        .limit(10)
        .all()
    )

    results = []
    for ref in referencias:
        results.append({
            "id": ref.id,
            "codigo_barras": ref.codigo_barras,
            "nombre": ref.nombre,
            "marca": ref.marca if ref.marca else "",
            "url_imagen": ref.url_imagen or "",
            "categoria": ref.categoria or ""
        })
    
    return jsonify({"results": results})

##############################
# ENDPOINT: /api/buscar-imagen
##############################
@app.route('/api/buscar-imagen', methods=['POST'])
@login_requerido
def api_buscar_imagen():
    data = request.get_json() or {}
    nombre = data.get("nombre", "").strip()
    codigo = data.get("codigo", "").strip()
    force_search = data.get("forceSearch", False)  # Nuevo parámetro

    if not nombre:
        return jsonify({
            "status": "error",
            "message": "Debes proporcionar al menos el nombre del producto para generar la imagen."
        })

    print(f"DEBUG API: Buscando imagen para nombre='{nombre}', codigo='{codigo}', force_search={force_search}")
    
    # Si force_search es True, saltamos directamente a buscar con SerpAPI
    if force_search:
        imagen_filename = buscar_imagen_google_images(nombre, codigo)
        if imagen_filename:
            image_url = url_for('static', filename=f'uploads/{imagen_filename}', _external=False)
            return jsonify({
                "status": "success",
                "image_url": image_url,
                "filename": imagen_filename,
                "message": "Imagen encontrada correctamente (búsqueda forzada)"
            })
        else:
            return jsonify({
                "status": "error",
                "message": "No se pudo obtener la imagen. Verifica tu conexión o intenta con otro nombre/código."
            })
    
    # Proceso normal si no es forzado
    # 1. Primero buscar en productos similares (rápido)
    producto_similar = find_similar_product_image(nombre, db.session, Producto)
    
    if producto_similar and producto_similar.foto:
        image_url = url_for('static', filename=f'uploads/{producto_similar.foto}', _external=False)
        return jsonify({
            "status": "success",
            "image_url": image_url,
            "filename": producto_similar.foto,
            "message": "Imagen encontrada en base de datos"
        })
    
    # 2. Segundo, buscar en catálogo global
    catalogo_similar = find_similar_catalog_image(nombre, db.session, CatalogoGlobal)
    
    if catalogo_similar and catalogo_similar.url_imagen:
        # Intentar obtener solo el nombre de archivo
        if "/uploads/" in catalogo_similar.url_imagen:
            filename = catalogo_similar.url_imagen.split("/uploads/")[-1]
        else:
            # Generar nombre aleatorio
            filename = f"{uuid.uuid4().hex}.jpg"
            
        # Descargar en segundo plano si es URL externa
        if catalogo_similar.url_imagen.startswith(("http://", "https://")):
            async_download_image(
                catalogo_similar.url_imagen, 
                os.path.join(UPLOAD_FOLDER, filename)
            )
            
        return jsonify({
            "status": "success",
            "image_url": catalogo_similar.url_imagen,
            "filename": filename,
            "message": "Imagen encontrada en catálogo global"
        })
    
    # 3. AQUÍ usamos SerpAPI - solo cuando el usuario hizo clic en "Generar con IA"
    # Esto es lo que consume recursos de SerpAPI, pero solo sucede cuando se solicita explícitamente
    imagen_filename = buscar_imagen_google_images(nombre, codigo)
    if imagen_filename:
        image_url = url_for('static', filename=f'uploads/{imagen_filename}', _external=False)
        return jsonify({
            "status": "success",
            "image_url": image_url,
            "filename": imagen_filename,
            "message": "Imagen encontrada correctamente"
        })
    
    # Si no se encontró imagen
    return jsonify({
        "status": "error",
        "message": "No se pudo obtener la imagen. Verifica tu conexión o intenta con otro nombre/código."
    })

##############################
# ENDPOINT: /api/check_barcode_exists
##############################
@app.route('/api/check_barcode_exists', methods=['GET'])
@login_requerido
def check_barcode_exists():
    codigo = request.args.get('codigo', '').strip()
    if not codigo:
        return jsonify({"exists": False})
    
    producto_existente = Producto.query.filter_by(
        empresa_id=session['user_id'],
        codigo_barras_externo=codigo
    ).first()
    
    return jsonify({"exists": producto_existente is not None})

##############################################
# FUNCIÓN DE SINCRONIZACIÓN
##############################################
def sync_gsheet_to_catalogo():
    data = leer_hoja("Hoja 1!A2:F")
    if not data:
        print("No hay filas en la hoja de Google.")
        return

    codigos_sheet = set()
    for row in data:
        if len(row) < 6:
            continue
        codigo = row[0].strip()
        nombre = row[1].strip()
        marca = row[2].strip()
        categoria = row[3].strip()
        unidad = row[4].strip()
        url_img = row[5].strip()

        # Truncar URL si es demasiado larga para evitar errores
        url_img = truncar_url(url_img, 295)  # Usar 295 para la columna de 300 caracteres

        codigos_sheet.add(codigo)

        ref = CatalogoGlobal.query.filter_by(codigo_barras=codigo).first()
        if ref:
            ref.nombre = nombre
            ref.marca = marca
            ref.categoria = categoria
            ref.url_imagen = url_img
            print(f"Actualizando {codigo} -> {nombre}")
        else:
            nuevo = CatalogoGlobal(
                codigo_barras=codigo,
                nombre=nombre,
                marca=marca,
                categoria=categoria,
                url_imagen=url_img
            )
            db.session.add(nuevo)
            print(f"Creando {codigo} -> {nombre}")

    db.session.commit()

    # Elimina lo que ya no existe en la Hoja
    todos = CatalogoGlobal.query.all()
    for ref in todos:
        if ref.codigo_barras not in codigos_sheet:
            db.session.delete(ref)
            print(f"Borrando {ref.codigo_barras}, ya no está en la hoja.")
    db.session.commit()

@app.route('/debug-endpoint')
@login_requerido
def debug_endpoint():
    """Muestra todos los endpoints registrados"""
    import pprint
    endpoints = []
    for rule in app.url_map.iter_rules():
        endpoints.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'rule': str(rule)
        })
    
    # Filtrar solo los relacionados con ajuste
    ajuste_endpoints = [e for e in endpoints if 'ajuste' in e['endpoint'].lower() or 'new' in e['endpoint'].lower()]
    
    return f"""
    <h2>Endpoints relacionados con ajuste:</h2>
    <pre>{pprint.pformat(ajuste_endpoints, indent=2)}</pre>
    
    <h2>Todos los endpoints:</h2>
    <pre>{pprint.pformat(endpoints, indent=2)}</pre>
    """

##############################################
# MAIN
##############################################
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)

##############################
# ENDPOINT: /api/autocomplete_by_name
##############################
@app.route('/api/autocomplete_by_name', methods=['GET'])
@login_requerido
def api_autocomplete_by_name():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({"results": []})

    # Consulta optimizada con límite y selección específica de campos
    referencias = (
        CatalogoGlobal.query
        .with_entities(
            CatalogoGlobal.id,
            CatalogoGlobal.codigo_barras,
            CatalogoGlobal.nombre,
            CatalogoGlobal.marca,
            CatalogoGlobal.url_imagen,
            CatalogoGlobal.categoria
        )
        .filter(CatalogoGlobal.nombre.ilike(f"%{q}%"))
        .limit(10)
        .all()
    )

    results = []
    for ref in referencias:
        results.append({
            "id": ref.id,
            "codigo_barras": ref.codigo_barras,
            "nombre": ref.nombre,
            "marca": ref.marca if ref.marca else "",
            "url_imagen": ref.url_imagen or "",
            "categoria": ref.categoria or ""
        })
    return jsonify({"results": results})

##############################
# ENDPOINT /api/find-by-code
##############################
@app.route('/api/find_by_code', methods=['GET'])
@login_requerido
def api_find_by_code():
    code = request.args.get("codigo", "").strip()
    if not code:
        return jsonify({"found": False})

    # Optimizado: usar 'get' para búsqueda exacta por clave
    ref = CatalogoGlobal.query.filter_by(codigo_barras=code).first()
    
    # Búsqueda parcial si no encontramos coincidencia exacta y el código es suficientemente largo
    if not ref and len(code) > 7:
        ref = CatalogoGlobal.query.filter(
            CatalogoGlobal.codigo_barras.startswith(code[:7])
        ).first()
    
    if not ref:
        return jsonify({"found": False})

    return jsonify({
        "found": True,
        "codigo_barras": ref.codigo_barras,
        "nombre": ref.nombre,
        "marca": ref.marca if ref.marca else "",
        "url_imagen": ref.url_imagen or "",
        "categoria": ref.categoria or ""
    })